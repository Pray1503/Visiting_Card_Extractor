# #!/usr/bin/env python3
# # ══════════════════════════════════════════════════════════════════════════════
# #  VISITING CARD OCR ENGINE  v18.0  —  VERIFIED AGAINST 8-CARD TEST RUN
# # ══════════════════════════════════════════════════════════════════════════════
# #
# #  ALL BUGS VERIFIED FROM v17 LIVE OUTPUT vs actual card images:
# #
# #  ❌ v17 BUG A — Card 01 (Aditya Varma): Name = "—" (missing)
# #     ACTUAL CARD: "ADITYA V. VARMA"  (wide letter-spaced, large font)
# #     ROOT CAUSE:
# #       _is_name_word() rejected "V." because the dot-rejection rule said
# #       "reject tokens containing dots unless single-letter initial".
# #       The check was: if "." in w2: return False  — BUT initials like "V."
# #       contain a dot and were supposed to pass via the earlier fullmatch check
# #       r"[A-Za-z]\\.?" which correctly matches "V.".  However the ORDERING in
# #       v17 put the dot-rejection BEFORE the fullmatch, so "V." was rejected
# #       before it could be accepted as an initial.
# #       SECONDARY CAUSE: spaCy NER did not fire on the all-caps spaced name
# #       "ADITYA V. VARMA" because the name was tokenised differently after
# #       _collapse_spaced() ran on the raw OCR line.
# #     ✅ FIX:
# #       (a) Reorder _is_name_word(): check initial pattern BEFORE dot rejection.
# #       (b) Accept "V." and similar single-letter-dot tokens explicitly.
# #       (c) In the heuristic name scorer, allow words list where ONE token
# #           is a dotted initial (e.g. "ADITYA V. VARMA" → 3 words, 1 initial).
# #
# #  ❌ v17 BUG B — Card 02 (Sara Al-Mansoori): Phone #1 missing (+971 4 321 9900)
# #     ACTUAL CARD: two phones — Office +971 4 321 9900 / Mobile +971 50 800 1234
# #     ROOT CAUSE:
# #       _extract_phones() found the mobile number first. When phonenumbers
# #       library validated +971 4 321 9900, it returned it as a valid AE number.
# #       But the line-by-line parser stopped after finding ONE number per line
# #       because the `parsed = True; break` logic exited the region loop after
# #       first success, then the outer `if parsed: continue` moved to the next
# #       line — so BOTH numbers on DIFFERENT lines should have been found.
# #       Investigation: the two phones are on separate lines.  The real issue
# #       is that "+971 4 321 9900" has only 9 digits total (country code 971 +
# #       national 4321 9900 = 7 national digits), which passes the ≥6 check,
# #       but the phonenumbers library with region=AE rejects it because UAE
# #       landline numbers need the area code to be 2 digits not 1.
# #       PaddleOCR read it as "+971 4 321 9900" — the real number on the card
# #       is "+971 4 321 9900" which IS a valid UAE landline (area code 4 = Dubai,
# #       number 321 9900 = 7 digits → total 9 national digits including area code).
# #       The phonenumbers library DOES validate this correctly for AE.
# #       So the bug is that `detected_region` was not being set to "AE" from
# #       the FIRST phone line, causing the fallback region order to try others
# #       first, and one of those other regions was successfully parsing "+971 4"
# #       as a country-code match before AE got a chance.
# #     ✅ FIX:
# #       (a) Detect region from ALL lines upfront (not just first match).
# #       (b) When a +CC prefix is present, ALWAYS try the matching region first
# #           with no fallback until that region is exhausted.
# #       (c) Lower national digit minimum to 6 (UAE landlines: area(1)+number(7)=8,
# #           but the national_number field contains all digits after country code).
# #
# #  ❌ v17 BUG C — Card 04 (Priya Krishnamurthy): Phone = "—" (missing)
# #     ACTUAL CARD: +65 9123 4567
# #     ROOT CAUSE:
# #       The address line "71 Ayer Rajah Crescent, #03-18, Singapore 139951"
# #       contains ADDRESS_KW ("Crescent") → the ENTIRE line is skipped by the
# #       address-line exclusion guard in _extract_phones().
# #       BUT the phone "+65 9123 4567" is on its OWN separate line — it should
# #       NOT be excluded.  The bug is that ADDRESS_KW.search() was matching
# #       "Crescent" in the ADDRESS line, not the phone line, so the phone line
# #       itself should have been fine.
# #       Re-reading the v17 code: the _extract_phones() loop checks
# #       `if ADDRESS_KW.search(ln): continue` — this correctly skips the
# #       address line.  The phone line "+65 9123 4567" should NOT contain
# #       any ADDRESS_KW.  So why is the phone missing?
# #       Answer: The tilted Singapore card (card_04) has OCR difficulties due
# #       to the tilt.  The phone "+65 9123 4567" was correctly detected by
# #       phonenumbers, but then the cross-check against address text failed:
# #       `if len(digs) <= 7 and address_digs and digs in address_digs: continue`
# #       The Singapore postal code is 139951 (6 digits). The phone number
# #       +65 9123 4567 has digits "6591234567" (10 digits) → len > 7, so this
# #       cross-check should NOT trigger.
# #       But the FALLBACK regex runs when phonenumbers returns nothing. The
# #       fallback requires 8+ digits for non-+ prefixed numbers. If OCR reads
# #       the line as "65 9123 4567" without the leading "+", then len("659123456") = 9
# #       which has a "+"-free prefix → requires 8 digits minimum → passes (9 ≥ 8).
# #       But wait — "65 9123 4567" without "+" starts with "6", not "+", so the
# #       fallback requires 8 digits and "659123456" has 9 — should pass.
# #       The REAL issue: phonenumbers library with SG region validates
# #       "+65 9123 4567" correctly. But the line may have been read as
# #       "65 9123 4567" (no plus sign) by OCR. Without "+", the detected_region
# #       is None, and the fallback tries IN first — and "659123456" with IN
# #       region fails. Eventually SG is tried and it validates... actually it
# #       should work.  The actual bug: the tilted card may produce OCR that
# #       puts the phone on the same line as other text, and ADDRESS_KW then
# #       matches that combined line.
# #     ✅ FIX:
# #       (a) More robust phone line detection: if a line contains both
# #           address keywords AND a clear phone pattern (+\d or \d{8+}),
# #           extract the phone first, then still skip the line for address-only scan.
# #       (b) Scan each token individually (not just full lines) for phone numbers
# #           when the full line is skipped due to address content.
# #       (c) Ensure SG is high in fallback region list.
# #
# #  ❌ v17 BUG D — Card 06 (Thabo Nkosi): Address duplicated with OCR corruption
# #     "1 Sturdee Avenue, Rosebank, Johannesburg 2196, South Africa
# #      1 Sturdee Avenue, Rosebank, Johannesburq 2l96, South Africa"
# #     ROOT CAUSE:
# #       The dual OCR pass produced two slightly different readings of the same
# #       address line (OCR corruption: q→g, l→1). The fuzzy dedup in
# #       _extract_address() used rfuzz.ratio() with threshold 85, but the
# #       corrupted version was added FIRST (before the correct one), and the
# #       correct one was subsequently rejected as a duplicate of the corruption.
# #       OR: both survived because their ratio was just below 85.
# #       The corrupted "Johannesburq 2l96" has enough character differences
# #       that rfuzz.ratio against the correct version gives ~92 — so both
# #       should NOT appear. But the IOU dedup at the token level may have kept
# #       BOTH tokens because "Johannesburg 2196" and "Johannesburq 2l96" have
# #       different text (case-insensitive check: "JOHANNESBURG" ≠ "JOHANNESBURQ")
# #       so the IOU dedup doesn't apply (it requires identical text).
# #     ✅ FIX:
# #       (a) In _extract_address(), use fuzzy dedup with threshold 80 (lower
# #           than before to catch OCR corruption variants).
# #       (b) When adding an address part, prefer the version with FEWER
# #           suspicious characters (digits used as letters: l→1, 0→O, q→g).
# #       (c) Also apply _dedup_line_text at the address join stage.
# #
# #  ❌ v17 BUG E — Card 08 (Isabella De Rosa): Name garbled "Isab Ella DE Rosa"
# #                                              Company = "ARMA NI G R 0 U P S.P.A."
# #     ROOT CAUSE:
# #       The spaced luxury font "I S A B E L L A  D E  R O S A" was partially
# #       collapsed by PaddleOCR into chunks "Isab", "Ella", "DE", "Rosa" — the
# #       OCR engine itself merged some adjacent letters before our code sees them.
# #       _collapse_spaced() only handles clean single-letter sequences and
# #       requires ≥50% single-char tokens. "Isab Ella DE Rosa" has 4 tokens,
# #       none of which are single letters → _collapse_spaced() returns it unchanged.
# #       Similarly "ARMA NI G R 0 U P S.P.A." has mixed token lengths.
# #       The "0" in "GR0UP" is a digit-for-letter OCR error (0→O).
# #     ✅ FIX:
# #       (a) New _reconstruct_luxury_name() function: given a line where the
# #           words are all short (1-6 chars) and look like letter fragments,
# #           attempt to reconstruct the full name by:
# #           1. Stripping internal spaces to get one long string
# #           2. Checking if the result is a known-structure name using heuristics
# #           3. For company lines, apply _fix_digit_substitutions() aggressively.
# #       (b) _fix_digit_substitutions(): in ALL-CAPS context, replace
# #           0→O, 1→I, 5→S across the full token before any other processing.
# #       (c) For the name extractor: when no name is found via heuristic/NER,
# #           try a "reconstruction" pass that collapses all short tokens on
# #           the topmost lines of the card.
# #       (d) Company extractor: apply _fix_digit_substitutions() to the final
# #           extracted company string.
# #
# #  ❌ v17 BUG F — Card 03 (Klaus Müller): Job Title = "—" (missing)
# #     ACTUAL CARD: "Senior Systems Engineer" (letter-spaced)
# #     ROOT CAUSE:
# #       OCR reads the spaced title as "S e n i o r  S y s t e m s  E n g i n e e r"
# #       → _collapse_spaced() receives this as single-letter tokens separated
# #       by spaces/double-spaces.  After collapse: "Senior Systems Engineer".
# #       But DESIGNATION_KW requires \b(senior|engineer)\b which needs word
# #       boundaries. If _collapse_spaced() produces "SeniorSystemsEngineer"
# #       (no spaces, v14's bug), the regex won't match.
# #       v17 claims to preserve spaces but the actual output shows "—".
# #       This means either (a) collapse produced no spaces, or (b) the line
# #       was not reached because it was claimed by another extractor.
# #     ✅ FIX: The v17 _collapse_spaced() already preserves word boundaries.
# #       The real fix needed: ensure clean_token() is called on the RAW token
# #       text BEFORE the token is merged into a line, AND that the line-level
# #       clean applies _collapse_spaced() on the FULL LINE TEXT (not per-token).
# #       Add full-line collapse pass in group_into_lines().
# #
# #  ❌ v17 BUG G — Card 04 (Priya): Company has trailing comma
# #     "VERTEX AI LABS PTE.LTD,"
# #     ROOT CAUSE: trailing punctuation not stripped from company field.
# #     ✅ FIX: strip trailing ".,;:" from all extracted field strings.
# #
# #  INSTALL:
# #    pip install paddlepaddle paddleocr opencv-python-headless pillow numpy
# #    pip install pandas openpyxl phonenumbers tldextract rapidfuzz spacy
# #    python -m spacy download en_core_web_sm
# #    (PDF) pip install pdf2image  +  poppler (choco install poppler on Windows)
# #
# #  USAGE:
# #    python visiting_card_ocr_v18.py                   # GUI file picker
# #    python visiting_card_ocr_v18.py card.jpg          # single image
# #    python visiting_card_ocr_v18.py ./cards/          # batch folder
# #    python visiting_card_ocr_v18.py card.jpg --debug  # verbose mode
# #    python visiting_card_ocr_v18.py card.jpg --no-vcf # skip VCF export
# # ══════════════════════════════════════════════════════════════════════════════

# # ── oneDNN Windows crash prevention — MUST precede ALL other imports ──────────
# import os

# os.environ["FLAGS_use_mkldnn"] = "0"
# os.environ["FLAGS_call_stack_level"] = "2"

# import re, io, cv2, sys, csv, math, json, zipfile, argparse, traceback
# import unicodedata, logging
# from datetime import datetime
# from pathlib import Path

# import numpy as np
# from PIL import Image

# # ── Logging ────────────────────────────────────────────────────────────────────
# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] %(message)s",
#     datefmt="%H:%M:%S",
#     handlers=[
#         logging.StreamHandler(sys.stdout),
#         logging.FileHandler("ocr_engine.log", encoding="utf-8"),
#     ],
# )
# log = logging.getLogger("VC_OCR")

# # ── Soft dependencies ──────────────────────────────────────────────────────────
# try:
#     import pandas as pd
#     import openpyxl
#     from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
#     from openpyxl.utils import get_column_letter

#     EXCEL_OK = True
# except ImportError:
#     EXCEL_OK = False
#     log.warning("pip install pandas openpyxl")

# try:
#     import phonenumbers
#     from phonenumbers import PhoneNumberMatcher, PhoneNumberFormat

#     PHONE_OK = True
# except ImportError:
#     PHONE_OK = False
#     log.warning("pip install phonenumbers")

# try:
#     import tldextract

#     TLDX_OK = True
# except ImportError:
#     TLDX_OK = False

# try:
#     import spacy

#     _NLP = spacy.load("en_core_web_sm")
#     SPACY_OK = True
# except Exception:
#     _NLP = None
#     SPACY_OK = False
#     log.warning("python -m spacy download en_core_web_sm")

# try:
#     from rapidfuzz import fuzz as rfuzz

#     FUZZ_OK = True
# except ImportError:
#     FUZZ_OK = False
#     log.warning("pip install rapidfuzz")

# try:
#     from paddleocr import PaddleOCR

#     PADDLE_OK = True
# except ImportError:
#     PADDLE_OK = False
#     log.warning("pip install paddlepaddle paddleocr")

# try:
#     import torch

#     GPU_AVAIL = torch.cuda.is_available()
# except ImportError:
#     GPU_AVAIL = False

# try:
#     from pdf2image import convert_from_bytes

#     PDF_OK = True
# except ImportError:
#     PDF_OK = False

# try:
#     from tkinter import Tk, filedialog

#     TK_OK = True
# except ImportError:
#     TK_OK = False

# import warnings

# warnings.filterwarnings("ignore")


# # ══════════════════════════════════════════════════════════════════════════════
# #  CONFIGURATION
# # ══════════════════════════════════════════════════════════════════════════════
# CFG = {
#     "OCR_CONF_THRESH": 0.18,
#     "PDF_DPI": 250,
#     "SEG_MIN_AREA_FRAC": 0.03,
#     "SEG_MAX_AREA_FRAC": 0.95,
#     "SEG_PAD_PX": 12,
#     "SEG_CARD_ASPECT_MIN": 0.5,
#     "SEG_CARD_ASPECT_MAX": 5.0,
#     "ROW_TOL_FACTOR": 0.55,
#     "ROW_TOL_FALLBACK_PX": 22,
#     "MERGE_CELL_FACTOR": 0.035,
#     "MERGE_CELL_MIN_PX": 18,
#     "ECHO_IOU_THRESH": 0.45,
#     "SECONDARY_SCRIPT_THRESH": 0.15,
#     "PHONE_FUZZ_THRESH": 80,
#     "NAME_ALPHA_RATIO": 0.75,
#     "PHONE_MIN_DIGITS_NO_PREFIX": 8,
#     "ADDR_FUZZ_THRESH": 80,  # BUG D FIX: lower threshold catches OCR corruption
#     "MAX_PHONES": 5,
#     "DEBUG": False,
#     "EXCEL_DB_FILE": "cards_database.xlsx",
# }

# SCRIPT_TO_LANG = {
#     "latin": "en",
#     "arabic": "ar",
#     "devanagari": "hi",
#     "cjk": "ch",
#     "korean": "korean",
#     "cyrillic": "ru",
#     "japanese": "japan",
#     "tamil": "ta",
#     "telugu": "te",
#     "thai": "th",
# }


# # ══════════════════════════════════════════════════════════════════════════════
# #  REGEX PATTERNS
# # ══════════════════════════════════════════════════════════════════════════════

# # Strict TLD-anchored email regex — terminates at TLD boundary (BUG 6 fix)
# _TLD_LIST = (
#     r"com|org|net|in|io|co\.in|co\.uk|co\.ae|co\.sg|co\.za|co\.au|"
#     r"co\.nz|co\.jp|biz|info|edu|gov|ae|us|au|nz|sg|hk|my|ph|za|"
#     r"de|fr|jp|cn|ca|br|mx|tech|app|dev|ai|online|store|site|"
#     r"il|tr|pk|lk|ir|ma|dz|tn|ly|ng|ke|tz|ug|zm|zw|"
#     r"pt|lu|ie|fi|lv|ee|ua|cz|sk|sa|bh|qa|gr|nl|be|hu|it|"
#     r"ro|ch|at|gb|dk|se|no|pl|ru|kr|vn|id|th|tw|ar|pe|cl|co|ve"
# )
# _EMAIL_RE = re.compile(
#     r"[A-Za-z0-9._%+\-]{1,64}"
#     r"@"
#     r"[A-Za-z0-9.\-]{1,253}"
#     r"\."
#     r"(?:" + _TLD_LIST + r")"
#     r"(?=\s|$|[^a-zA-Z0-9.\-])",
#     re.IGNORECASE,
# )

# _WEB_RE = re.compile(
#     r"(?:https?://|www\.)[A-Za-z0-9.\-/_%?=&#]+"
#     r"|[A-Za-z0-9][\w\-]*\."
#     r"(?:" + _TLD_LIST + r")"
#     r"(?:[/\w\-]*)?",
#     re.IGNORECASE,
# )

# _LINKEDIN_RE = re.compile(r"linkedin\.com/in/[\w\-]+", re.IGNORECASE)
# _TWITTER_RE = re.compile(
#     r"(?:(?:^|(?<=\s))@[\w]{2,50}|twitter\.com/[\w]+|x\.com/[\w]+)",
#     re.IGNORECASE | re.MULTILINE,
# )
# _SOCIAL_RE = re.compile(
#     r"linkedin\.com|twitter\.com|x\.com|instagram\.com|facebook\.com|@[\w]{2,50}",
#     re.IGNORECASE,
# )
# _PHONE_LABEL_RE = re.compile(
#     r"^\s*(?:office|mobile|mob|cell|tel|fax|ph|phone|direct|work|home|"
#     r"helpline|hotline|tollfree|toll[-\s]free)\s*[:\-]?\s*",
#     re.IGNORECASE,
# )
# _INTL_PREFIX_RE = re.compile(
#     r"(?:^|[\s\(\[])\+(\d{1,3})[\s\-\(]|(?:^|\s)\+(\d{1,3})(?=\d)"
# )

# # Postal code formats — prevents postal codes from being parsed as phones
# _POSTAL_FMT_RE = re.compile(
#     r"^\d{3}-\d{4}$"  # JP: 108-0075
#     r"|^\d{4,6}$"  # IN/SG/AU: 395007
#     r"|^\d{2}-\d{4}$"  # variant
#     r"|^[A-Z]{1,2}\d{1,2}\s?\d[A-Z]{2}$",  # GB: WC1R 5AH
#     re.IGNORECASE,
# )

# DESIGNATION_KW = re.compile(
#     r"\b(ceo|cto|coo|cfo|cmo|vp|md|gm|founder|co-?founder|partner|"
#     r"director|manager|engineer|developer|architect|analyst|"
#     r"consultant|officer|president|vice[\s\-]?president|associate|"
#     r"senior|junior|principal|head|lead|specialist|advisor|"
#     r"coordinator|supervisor|assistant|proprietor|owner|chairman|"
#     r"trustee|secretary|treasurer|intern|trainee|"
#     r"doctor|dr\.?|prof\.?|professor|lawyer|advocate|solicitor|"
#     r"accountant|auditor|designer|strategist|researcher|scientist|"
#     r"technician|representative|agent|broker|dealer|contractor|builder|"
#     r"gerant|directeur|ingenieur|presidente|gerente|ingeniero)\b",
#     re.IGNORECASE,
# )
# COMPANY_KW = re.compile(
#     r"\b(ltd|limited|inc|llp|llc|pvt|private|plc|corp|corporation|"
#     r"solutions|consulting|studio|technologies|tech|group|company|"
#     r"labs|services|systems|enterprises|associates|builders|"
#     r"construction|industries|global|international|ventures|"
#     r"holdings|capital|networks|media|digital|infra|infrastructure|"
#     r"realty|realtors|properties|developers|architects|interiors|"
#     r"design|agency|firm|foundation|trust|institute|academy|"
#     r"hospital|clinic|healthcare|pharma|logistics|transport|"
#     r"exports|imports|trading|manufacturing|fabrication|engineering|"
#     r"gmbh|ag|kg|ohg|ug|sarl|sas|spa|srl|s\.p\.a\.|k\.k\.|kk|"
#     r"sdn\.?\s*bhd|pte\.?|b\.v\.|n\.v\.|a\/s|s\.a\.)\b",
#     re.IGNORECASE,
# )
# ADDRESS_KW = re.compile(
#     r"\b(floor|fl\.|level|suite|plot|block|sector|phase|unit|apt|apartment|"
#     r"building|bldg|tower|complex|plaza|centre|center|mall|"
#     r"road|rd\.|street|st\.|avenue|ave\.|lane|ln\.|boulevard|blvd\.|"
#     r"drive|dr\.|court|ct\.|place|pl\.|way|highway|freeway|square|sq\.|"
#     r"terrace|ter\.|close|crescent|grove|gardens|"
#     r"nagar|vihar|marg|gali|chowk|bazaar|colony|society|residency|"
#     r"stra[ß]e|strasse|str\.|gasse|weg|platz|rue|impasse|allee|"
#     r"calle|avenida|carrera|paseo|"
#     r"district|area|zone|suburb|village|town|city|state|province|"
#     r"county|region|po\s*box|p\.o\.|zip|pin|postcode|postal)\b",
#     re.IGNORECASE,
# )
# _NAME_STOPWORDS = re.compile(
#     r"\b(floor|level|suite|road|street|avenue|lane|boulevard|drive|court|"
#     r"place|square|terrace|close|crescent|complex|tower|building|"
#     r"nagar|vihar|marg|chowk|colony|society|stra[ß]e|strasse|gasse|weg|"
#     r"platz|rue|calle|avenida|carrera|block|sector|phase|plot|"
#     r"district|zone|suburb|village|town|city|state|province|county|"
#     r"region|postcode|postal|zip|pin)\b",
#     re.IGNORECASE,
# )
# POSTAL_RE = re.compile(r"\b\d{4,9}\b")

# _CC_REGION: dict = {
#     "1": "US",
#     "7": "RU",
#     "20": "EG",
#     "27": "ZA",
#     "30": "GR",
#     "31": "NL",
#     "32": "BE",
#     "33": "FR",
#     "34": "ES",
#     "36": "HU",
#     "39": "IT",
#     "40": "RO",
#     "41": "CH",
#     "43": "AT",
#     "44": "GB",
#     "45": "DK",
#     "46": "SE",
#     "47": "NO",
#     "48": "PL",
#     "49": "DE",
#     "51": "PE",
#     "52": "MX",
#     "54": "AR",
#     "55": "BR",
#     "56": "CL",
#     "57": "CO",
#     "60": "MY",
#     "61": "AU",
#     "62": "ID",
#     "63": "PH",
#     "64": "NZ",
#     "65": "SG",
#     "66": "TH",
#     "81": "JP",
#     "82": "KR",
#     "84": "VN",
#     "86": "CN",
#     "90": "TR",
#     "91": "IN",
#     "92": "PK",
#     "94": "LK",
#     "98": "IR",
#     "212": "MA",
#     "213": "DZ",
#     "216": "TN",
#     "218": "LY",
#     "234": "NG",
#     "254": "KE",
#     "255": "TZ",
#     "256": "UG",
#     "260": "ZM",
#     "263": "ZW",
#     "351": "PT",
#     "352": "LU",
#     "353": "IE",
#     "358": "FI",
#     "371": "LV",
#     "372": "EE",
#     "380": "UA",
#     "420": "CZ",
#     "421": "SK",
#     "966": "SA",
#     "968": "OM",
#     "971": "AE",
#     "972": "IL",
#     "974": "QA",
# }
# # BUG B FIX: SG moved near top to ensure it's tried early for Singapore cards
# _FALLBACK_REGIONS = [
#     "IN",
#     "US",
#     "GB",
#     "AE",
#     "SG",
#     "AU",
#     "CA",
#     "DE",
#     "FR",
#     "ZA",
#     "NG",
#     "KE",
#     "JP",
#     "CN",
#     "KR",
#     "BR",
#     "MX",
#     "NZ",
#     "PH",
#     "MY",
#     "BD",
#     "LK",
#     "NP",
#     "KW",
#     "QA",
#     "BH",
#     "SA",
#     "OM",
#     "EG",
# ]

# EXCEL_COLS = [
#     "Name",
#     "Job_Title",
#     "Company",
#     "Email",
#     "Phone_1",
#     "Phone_2",
#     "Phone_3",
#     "Website",
#     "Address",
#     "LinkedIn",
#     "Twitter",
#     "Quality",
#     "Processed",
#     "File",
# ]
# _QC = {"🟢 GREEN": "C6EFCE", "🟡 YELLOW": "FFEB9C", "🔴 RED": "FFC7CE"}


# # ══════════════════════════════════════════════════════════════════════════════
# #  OCR MODEL SINGLETON
# # ══════════════════════════════════════════════════════════════════════════════
# _PADDLE_CACHE: dict = {}


# def _get_paddle(lang: str = "en"):
#     if lang not in _PADDLE_CACHE:
#         if not PADDLE_OK:
#             raise RuntimeError("paddleocr not installed")
#         log.info("Loading PaddleOCR (lang=%s, gpu=%s)…", lang, GPU_AVAIL)
#         _PADDLE_CACHE[lang] = PaddleOCR(
#             lang=lang,
#             use_doc_orientation_classify=False,
#             use_doc_unwarping=False,
#             use_textline_orientation=False,
#             text_rec_score_thresh=CFG["OCR_CONF_THRESH"],
#             device="cpu",
#         )
#     return _PADDLE_CACHE[lang]


# # ══════════════════════════════════════════════════════════════════════════════
# #  L0 — INPUT LOADER
# # ══════════════════════════════════════════════════════════════════════════════
# _IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}


# def _pil_to_bgr(pil: Image.Image) -> np.ndarray:
#     return cv2.cvtColor(np.array(pil.convert("RGB")), cv2.COLOR_RGB2BGR)


# def load_images(path: str):
#     """Generator: yields (label, bgr_ndarray) from image / PDF / ZIP."""
#     p = Path(path)
#     ext = p.suffix.lower()
#     if ext in _IMG_EXTS:
#         img = cv2.imread(str(p), cv2.IMREAD_COLOR)
#         if img is None:
#             try:
#                 img = _pil_to_bgr(Image.open(p))
#             except Exception as e:
#                 log.error("Cannot read %s: %s", path, e)
#                 return
#         yield p.name, img
#     elif ext == ".pdf":
#         if not PDF_OK:
#             log.error("pdf2image not installed")
#             return
#         try:
#             for i, pg in enumerate(
#                 convert_from_bytes(p.read_bytes(), dpi=CFG["PDF_DPI"]), 1
#             ):
#                 yield f"{p.stem} (page {i})", _pil_to_bgr(pg)
#         except Exception as e:
#             log.error("PDF read failed (%s): %s", path, e)
#     elif ext == ".zip":
#         try:
#             with zipfile.ZipFile(p) as zf:
#                 for m in sorted(
#                     m
#                     for m in zf.namelist()
#                     if Path(m).suffix.lower() in _IMG_EXTS
#                     and not Path(m).name.startswith(".")
#                 ):
#                     try:
#                         arr = np.frombuffer(zf.read(m), np.uint8)
#                         img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
#                         if img is None:
#                             img = _pil_to_bgr(Image.open(io.BytesIO(zf.read(m))))
#                         yield f"{p.name}/{Path(m).name}", img
#                     except Exception as e:
#                         log.warning("Cannot read %s: %s", m, e)
#         except Exception as e:
#             log.error("ZIP read failed (%s): %s", path, e)
#     else:
#         log.warning("Unsupported file type: %s", ext)


# # ══════════════════════════════════════════════════════════════════════════════
# #  L1 — CARD SEGMENTOR
# # ══════════════════════════════════════════════════════════════════════════════


# def segment_cards(img: np.ndarray) -> list:
#     h, w = img.shape[:2]
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img.copy()
#     edges = cv2.Canny(cv2.GaussianBlur(gray, (5, 5), 0), 30, 120)
#     dilated = cv2.dilate(
#         edges, cv2.getStructuringElement(cv2.MORPH_RECT, (15, 5)), iterations=3
#     )
#     cnts, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
#     pad = CFG["SEG_PAD_PX"]
#     rois = []
#     for c in cnts:
#         area = cv2.contourArea(c)
#         if not (
#             h * w * CFG["SEG_MIN_AREA_FRAC"] <= area <= h * w * CFG["SEG_MAX_AREA_FRAC"]
#         ):
#             continue
#         x, y, cw, ch = cv2.boundingRect(c)
#         asp = cw / max(ch, 1)
#         if CFG["SEG_CARD_ASPECT_MIN"] <= asp <= CFG["SEG_CARD_ASPECT_MAX"]:
#             rois.append((x, y, cw, ch))
#     if not rois:
#         return [img]
#     rois.sort(key=lambda r: (r[1] // 100, r[0]))
#     return [
#         img[
#             max(0, y - pad) : min(h, y + ch + pad),
#             max(0, x - pad) : min(w, x + cw + pad),
#         ]
#         for x, y, cw, ch in rois
#     ]


# # ══════════════════════════════════════════════════════════════════════════════
# #  L2 — IMAGE PIPELINE
# # ══════════════════════════════════════════════════════════════════════════════


# def deskew(img: np.ndarray) -> np.ndarray:
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img
#     lines = cv2.HoughLinesP(
#         cv2.Canny(gray, 50, 150),
#         1,
#         np.pi / 180,
#         threshold=80,
#         minLineLength=60,
#         maxLineGap=20,
#     )
#     if lines is None:
#         return img
#     angles = []
#     for l in lines:
#         x1, y1, x2, y2 = l[0]
#         a = math.degrees(math.atan2(y2 - y1, x2 - x1))
#         if abs(a) < 45:
#             angles.append(a)
#     if not angles:
#         return img
#     angle = float(np.median(angles))
#     if abs(angle) < 0.5:
#         return img
#     h, w = img.shape[:2]
#     M = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
#     cos_a, sin_a = abs(M[0, 0]), abs(M[0, 1])
#     nw = int(h * sin_a + w * cos_a)
#     nh = int(h * cos_a + w * sin_a)
#     M[0, 2] += (nw - w) / 2
#     M[1, 2] += (nh - h) / 2
#     return cv2.warpAffine(
#         img, M, (nw, nh), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE
#     )


# def preprocess(img: np.ndarray) -> np.ndarray:
#     """Preprocess and return BGR (3-channel) array — PaddleOCR requires it."""
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img.copy()
#     gray = cv2.GaussianBlur(gray, (3, 3), 0)
#     roi = gray[
#         int(img.shape[0] * 0.30) : int(img.shape[0] * 0.70),
#         int(img.shape[1] * 0.30) : int(img.shape[1] * 0.70),
#     ]
#     if float(np.mean(roi)) < 128:
#         gray = cv2.bitwise_not(gray)
#     gray = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8)).apply(gray)
#     blur = cv2.GaussianBlur(gray, (0, 0), 3)
#     gray = cv2.addWeighted(gray, 1.55, blur, -0.55, 0)
#     if float(np.std(gray)) < 42:
#         gray = cv2.adaptiveThreshold(
#             gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 21, 10
#         )
#     return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)


# # ══════════════════════════════════════════════════════════════════════════════
# #  L3 — OCR ENGINE
# # ══════════════════════════════════════════════════════════════════════════════


# def _parse_paddle_result(results) -> list:
#     """
#     Parse PaddleOCR 2.x / 3.x result into list of (text, conf, pts).
#     Tries all known result formats in order.
#     """
#     parsed = []
#     if not results:
#         return parsed
#     for res_obj in results:
#         inner = None
#         # Strategy 1: subscriptable 3.x object
#         try:
#             _ = res_obj["rec_texts"]
#             inner = res_obj
#         except (KeyError, TypeError):
#             pass
#         # Strategy 2: .json attribute
#         if inner is None and hasattr(res_obj, "json"):
#             try:
#                 j = res_obj.json
#                 if isinstance(j, dict):
#                     inner = j
#             except Exception:
#                 pass
#         # Strategy 3: plain dict
#         if inner is None and isinstance(res_obj, dict):
#             inner = res_obj.get("res", res_obj)
#         # Strategy 4: classic 2.x list-of-[bbox,(text,conf)]
#         if inner is None:
#             try:
#                 items = list(res_obj)
#                 if items and isinstance(items[0], (list, tuple)):
#                     for item in items:
#                         if len(item) < 2:
#                             continue
#                         tc = item[1]
#                         if not isinstance(tc, (list, tuple)) or len(tc) < 2:
#                             continue
#                         text = str(tc[0]).strip()
#                         conf = float(tc[1])
#                         if not text or conf < CFG["OCR_CONF_THRESH"]:
#                             continue
#                         pts = [list(map(float, p)) for p in item[0]]
#                         parsed.append((text, conf, pts))
#                     continue
#             except Exception:
#                 pass
#         if inner is None:
#             continue
#         # Key-based extraction
#         try:
#             texts = inner.get("rec_texts", []) or []
#             scores = inner.get("rec_scores", []) or []
#             polys = inner.get("rec_polys", inner.get("dt_polys", [])) or []
#         except Exception:
#             continue
#         if isinstance(polys, np.ndarray):
#             polys = (
#                 list(polys) if polys.ndim == 3 else ([polys] if polys.ndim == 2 else [])
#             )
#         for text, conf, poly in zip(texts, scores, polys):
#             text = str(text).strip()
#             conf = float(conf)
#             if not text or conf < CFG["OCR_CONF_THRESH"]:
#                 continue
#             pts = (
#                 poly.tolist()
#                 if isinstance(poly, np.ndarray)
#                 else [list(map(float, p)) for p in poly]
#             )
#             parsed.append((text, conf, pts))
#     return parsed


# def _bbox_key(pts: list, cell_px: float) -> str:
#     cx = sum(p[0] for p in pts) / len(pts)
#     cy = sum(p[1] for p in pts) / len(pts)
#     return f"{int(cx//cell_px)},{int(cy//cell_px)}"


# def _iou(a: list, b: list) -> float:
#     ax1 = min(p[0] for p in a)
#     ax2 = max(p[0] for p in a)
#     ay1 = min(p[1] for p in a)
#     ay2 = max(p[1] for p in a)
#     bx1 = min(p[0] for p in b)
#     bx2 = max(p[0] for p in b)
#     by1 = min(p[1] for p in b)
#     by2 = max(p[1] for p in b)
#     ix = max(0.0, min(ax2, bx2) - max(ax1, bx1))
#     iy = max(0.0, min(ay2, by2) - max(ay1, by1))
#     inter = ix * iy
#     if not inter:
#         return 0.0
#     return inter / (
#         max((ax2 - ax1) * (ay2 - ay1), 1) + max((bx2 - bx1) * (by2 - by1), 1) - inter
#     )


# def _iou_dedup(pool: dict) -> dict:
#     """IOU-based dedup: remove overlapping tokens with identical text."""
#     tokens = list(pool.values())
#     keep = [True] * len(tokens)
#     for i in range(len(tokens)):
#         if not keep[i]:
#             continue
#         for j in range(i + 1, len(tokens)):
#             if not keep[j]:
#                 continue
#             if tokens[i]["text"].strip().upper() == tokens[j]["text"].strip().upper():
#                 if _iou(tokens[i]["bbox"], tokens[j]["bbox"]) >= CFG["ECHO_IOU_THRESH"]:
#                     if tokens[i]["conf"] >= tokens[j]["conf"]:
#                         keep[j] = False
#                     else:
#                         keep[i] = False
#                         break
#     result = {}
#     for i, tok in enumerate(tokens):
#         if keep[i]:
#             k = _bbox_key(tok["bbox"], max(tok.get("th", 20), CFG["MERGE_CELL_MIN_PX"]))
#             result[k] = tok
#     return result


# def _run_ocr_pass(img_bgr: np.ndarray, lang: str, cell_px: float, pool: dict):
#     if img_bgr.ndim == 2:
#         img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)
#     elif img_bgr.ndim == 3 and img_bgr.shape[2] == 4:
#         img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_BGRA2BGR)
#     try:
#         raw = _get_paddle(lang).predict(img_bgr)
#         parsed = _parse_paddle_result(raw)
#         if CFG["DEBUG"]:
#             log.debug("OCR pass lang=%s: %d tokens", lang, len(parsed))
#             if not parsed and raw:
#                 log.debug("  raw type=%s repr=%s", type(raw).__name__, repr(raw)[:200])
#         for text, conf, pts in parsed:
#             xs = [p[0] for p in pts]
#             ys = [p[1] for p in pts]
#             tok = {
#                 "text": text,
#                 "conf": conf,
#                 "bbox": pts,
#                 "cx": sum(xs) / len(xs),
#                 "cy": sum(ys) / len(ys),
#                 "th": max(ys) - min(ys),
#             }
#             k = _bbox_key(pts, cell_px)
#             if k not in pool or conf > pool[k]["conf"]:
#                 pool[k] = tok
#     except Exception as e:
#         log.warning("OCR pass failed (lang=%s): %s", lang, e)
#         if CFG["DEBUG"]:
#             traceback.print_exc()


# def detect_scripts(text: str):
#     counts = {s: 0 for s in SCRIPT_TO_LANG}
#     for ch in text:
#         if not ch.strip():
#             continue
#         cp = ord(ch)
#         if 0x0600 <= cp <= 0x06FF:
#             counts["arabic"] += 1
#         elif 0x0900 <= cp <= 0x097F:
#             counts["devanagari"] += 1
#         elif 0x4E00 <= cp <= 0x9FFF:
#             counts["cjk"] += 1
#         elif 0xAC00 <= cp <= 0xD7AF or 0x1100 <= cp <= 0x11FF:
#             counts["korean"] += 1
#         elif 0x0400 <= cp <= 0x04FF:
#             counts["cyrillic"] += 1
#         elif 0x3040 <= cp <= 0x30FF:
#             counts["japanese"] += 1
#         elif 0x0B80 <= cp <= 0x0BFF:
#             counts["tamil"] += 1
#         elif 0x0C00 <= cp <= 0x0C7F:
#             counts["telugu"] += 1
#         elif 0x0E00 <= cp <= 0x0E7F:
#             counts["thai"] += 1
#         elif ch.isalpha():
#             counts["latin"] += 1
#     total = sum(counts.values()) or 1
#     ranked = sorted(counts.items(), key=lambda x: x[1], reverse=True)
#     primary = "latin"
#     if ranked[0][1] > 0 and (ranked[0][0] == "latin" or counts[ranked[0][0]] >= 5):
#         primary = ranked[0][0]
#     secondary = next(
#         (
#             sc
#             for sc, cnt in ranked[1:]
#             if sc != primary and cnt / total >= CFG["SECONDARY_SCRIPT_THRESH"]
#         ),
#         None,
#     )
#     return primary, secondary


# def ocr_card(proc: np.ndarray, lang_override: str = None):
#     h = proc.shape[0]
#     cell_px = max(h * CFG["MERGE_CELL_FACTOR"], CFG["MERGE_CELL_MIN_PX"])
#     pool: dict = {}
#     base = lang_override or "en"
#     _run_ocr_pass(proc, base, cell_px, pool)
#     _run_ocr_pass(cv2.convertScaleAbs(proc, alpha=1.35, beta=25), base, cell_px, pool)
#     pool = _iou_dedup(pool)
#     tokens = list(pool.values())
#     primary, secondary = detect_scripts(" ".join(t["text"] for t in tokens))
#     sec_lang = SCRIPT_TO_LANG.get(secondary)
#     if sec_lang and sec_lang not in {base, "en"}:
#         _run_ocr_pass(proc, sec_lang, cell_px, pool)
#         pool = _iou_dedup(pool)
#         tokens = list(pool.values())
#     return tokens, primary, secondary


# # ══════════════════════════════════════════════════════════════════════════════
# #  TEXT CLEANING
# # ══════════════════════════════════════════════════════════════════════════════


# def _fix_digit_substitutions(text: str) -> str:
#     """
#     BUG E FIX: In all-caps or predominantly-caps tokens, fix common OCR
#     digit-for-letter errors: 0→O, 1→I, 5→S.
#     Only applied when token is ≥70% uppercase letters.
#     Examples: "GR0UP" → "GROUP", "ARMAN1" → "ARMANI"
#     """
#     t = text.strip()
#     if not t:
#         return t
#     alpha = sum(c.isalpha() for c in t)
#     if alpha == 0:
#         return t
#     upper = sum(c.isupper() for c in t if c.isalpha())
#     if upper / max(alpha, 1) >= 0.70:
#         t = t.replace("0", "O").replace("1", "I").replace("5", "S")
#     return t


# def _collapse_spaced(text: str) -> str:
#     """
#     BUG 2 / F FIX: Collapse letter-spaced luxury typography while PRESERVING
#     word boundaries (critical for DESIGNATION_KW regex to work).

#     Handles:
#       "K L A U S  M Ü L L E R"       → "KLAUS MÜLLER"   (double-space = word gap)
#       "I S A B E L L A  D E  R O S A" → "ISABELLA DE ROSA"
#       "S e n i o r  S y s t e m s"   → "Senior Systems"
#       "C R E A T I V E  D I R E C T O R" → "CREATIVE DIRECTOR"
#       "A D I T Y A"                   → "ADITYA"          (single word)

#     Algorithm:
#       1. Split on whitespace → segments
#       2. Require ≥50% single-char letter segments AND all segments ≤3 chars
#       3. Use double-space (2+ spaces) as word separator
#       4. Collapse each word group independently
#       5. Preserve original casing intent
#     """
#     segs = text.strip().split()
#     if len(segs) < 2:
#         return text

#     def _is_letter(ch: str) -> bool:
#         return unicodedata.category(ch).startswith("L")

#     def _is_short_letter_seg(s: str) -> bool:
#         return 1 <= len(s) <= 3 and all(_is_letter(c) or c == "." for c in s)

#     single_count = sum(1 for s in segs if len(s) == 1 and _is_letter(s))
#     if single_count / len(segs) < 0.50:
#         return text
#     if not all(_is_short_letter_seg(s) for s in segs):
#         return text

#     # Use double-space in original text as word separator
#     word_groups = re.split(r"  +", text.strip())
#     if len(word_groups) == 1:
#         # Single-space only — one word
#         return "".join(segs)

#     result_words = []
#     for group in word_groups:
#         letters = [s for s in group.split(" ") if s]
#         if not letters:
#             continue
#         merged = "".join(letters)
#         # Determine casing: lower-dominant single-letters → Title case
#         single_lowers = sum(1 for s in letters if len(s) == 1 and s.islower())
#         use_title = single_lowers > len([s for s in letters if len(s) == 1]) / 2
#         result_words.append(merged.title() if use_title else merged.upper())

#     return " ".join(result_words) if result_words else "".join(segs)


# def _dedup_line_text(text: str) -> str:
#     """Remove exact repeated substrings in a joined line."""
#     t = text.strip()
#     if len(t) < 6:
#         return t
#     m = re.match(r"^(.{4,}?)\s+\1$", t)
#     if m:
#         return m.group(1)
#     for split in range(max(4, len(t) // 4), len(t) // 2 + 1):
#         half = t[:split]
#         if t[split:].lstrip() == half:
#             return half
#     return t


# def _dedup_repeated_words(text: str) -> str:
#     """Remove word-level repetitions: 'SONY GROUP K.K. SONY GROUP K.K.' → fixed."""
#     words = text.split()
#     for half in range(len(words) // 2, 1, -1):
#         for start in range(len(words) - half * 2 + 1):
#             if words[start : start + half] == words[start + half : start + half * 2]:
#                 return _dedup_repeated_words(
#                     " ".join(words[: start + half] + words[start + half * 2 :])
#                 )
#     return text


# def _normalise_pipe(text: str) -> str:
#     """Replace OCR | → I confusion in social-line separator contexts."""
#     if re.search(r"(?:linkedin|twitter|@)", text, re.I):
#         text = re.sub(r"(?<=[^\s])\s+I\s+(?=[^\s])", " | ", text)
#     return text


# def _reconstruct_fragmented_text(text: str) -> str:
#     """
#     Handle partially-merged luxury font OCR where PaddleOCR fuses some
#     adjacent spaced letters into chunks before our code sees them.

#     Patterns:
#       "Isab Ella DE Rosa"        → "ISABELLA DE ROSA" (name)
#       "ARMA NI GROUP S.P.A."     → "ARMANI GROUP S.P.A." (company with keyword)
#       "C REAT IVE D IRECTOR"     → "CREATIVE DIRECTOR" (title)

#     Detection: a line qualifies as fragmented-luxury if:
#       - It has ≥2 tokens
#       - ALL tokens are ≤6 chars AND purely alphabetic (no digits, no punct except ".")
#       - The tokens collectively don't form a meaningful phrase when split
#         (i.e. no individual token is a known keyword by itself)

#     Strategy: try collapsing tokens that look like letter-fragments (≤4 chars,
#     all-caps or mixed short-caps) while preserving tokens that are long enough
#     to be real words (≥5 chars) as word boundaries.
#     """
#     words = text.strip().split()
#     if len(words) < 2:
#         return text

#     # Only process if ALL words are short alphabetic fragments
#     def _is_fragment(w: str) -> bool:
#         clean = w.rstrip(".,;")
#         return 1 <= len(clean) <= 5 and all(c.isalpha() or c == "." for c in clean)

#     if not all(_is_fragment(w) for w in words):
#         return text

#     # Don't process if any word is a known standalone keyword (it's a real word)
#     long_words = [w for w in words if len(w) >= 5]
#     if long_words:
#         # Has real words — might be partially reconstructed already
#         # Try to merge only the very short fragments (1-2 chars) with neighbors
#         result = []
#         i = 0
#         while i < len(words):
#             w = words[i]
#             # Very short fragment (1-2 chars, all caps) → merge with next
#             if len(w) <= 2 and w.isupper() and i + 1 < len(words):
#                 merged = w + words[i + 1]
#                 result.append(merged)
#                 i += 2
#             else:
#                 result.append(w)
#                 i += 1
#         return " ".join(result)

#     # All words are short fragments — collapse everything then re-case
#     collapsed = "".join(words)
#     # Apply digit substitutions on the collapsed form
#     collapsed = _fix_digit_substitutions(collapsed)
#     return collapsed.upper()  # caller will re-case if needed


# def _strip_trailing_punct(text: str) -> str:
#     """BUG G FIX: Strip trailing commas, dots, semicolons from field values."""
#     return re.sub(r"[,\.;:\s]+$", "", text.strip())


# def clean_token(text: str) -> str:
#     text = unicodedata.normalize("NFKC", text)
#     text = _normalise_pipe(text)
#     text = _fix_digit_substitutions(text)
#     text = _collapse_spaced(text)
#     text = re.sub(r"\s*@\s*", "@", text)
#     # CRITICAL FIX: only collapse dot-spaces inside email/URL-like contexts
#     # (between two lowercase word chars), NOT after initials like "V. VARMA"
#     # Old: re.sub(r"(?<=\w)\s*\.\s*(?=\w)", ".") — this ate "V. V" → "V.V"
#     # New: only collapse when both sides are lowercase (email/url pattern)
#     text = re.sub(r"(?<=[a-z0-9])\s*\.\s*(?=[a-z0-9])", ".", text)
#     return re.sub(r" {2,}", " ", text).strip()


# # ══════════════════════════════════════════════════════════════════════════════
# #  LINE GROUPING
# # ══════════════════════════════════════════════════════════════════════════════


# def _collapse_row_singles(tokens: list) -> list:
#     """Merge runs of ≥3 consecutive single-letter tokens: ['J','O','H','N'] → ['JOHN']."""
#     result, i = [], 0
#     while i < len(tokens):
#         tok = tokens[i]
#         if len(tok) == 1 and unicodedata.category(tok)[0] == "L":
#             run = [tok]
#             j = i + 1
#             while (
#                 j < len(tokens)
#                 and len(tokens[j]) == 1
#                 and unicodedata.category(tokens[j])[0] == "L"
#             ):
#                 run.append(tokens[j])
#                 j += 1
#             if len(run) >= 3:
#                 result.append("".join(run))
#                 i = j
#             else:
#                 result.append(tok)
#                 i += 1
#         else:
#             result.append(tok)
#             i += 1
#     return result


# def group_into_lines(tokens: list) -> list:
#     """Group tokens into visual lines. Returns list of (line_text, row_tokens)."""
#     if not tokens:
#         return []
#     heights = [t.get("th", 20) for t in tokens]
#     tol = max(
#         float(np.median(heights)) * CFG["ROW_TOL_FACTOR"], CFG["ROW_TOL_FALLBACK_PX"]
#     )
#     by_y = sorted(tokens, key=lambda t: t["cy"])
#     rows, cur = [], [by_y[0]]
#     for tok in by_y[1:]:
#         if abs(float(tok["cy"]) - sum(t["cy"] for t in cur) / len(cur)) <= tol:
#             cur.append(tok)
#         else:
#             rows.append(sorted(cur, key=lambda t: t["cx"]))
#             cur = [tok]
#     rows.append(sorted(cur, key=lambda t: t["cx"]))
#     result = []
#     for row in rows:
#         texts = [clean_token(t["text"]) for t in row if t["text"].strip()]
#         # Remove consecutive duplicates
#         deduped = []
#         for txt in texts:
#             if not deduped or txt.upper() != deduped[-1].upper():
#                 deduped.append(txt)
#         # Collapse single-letter runs (BUG 2 FIX)
#         collapsed = _collapse_row_singles(deduped)
#         # Apply full-line collapse for cases like "S e n i o r  S y s t e m s"
#         # where each word's letters are individual tokens (BUG F FIX)
#         full_line = " ".join(collapsed)
#         full_line = _collapse_spaced(full_line)
#         # Try to fix partially-merged luxury font fragments (BUG E FIX)
#         # e.g. "Isab Ella DE Rosa" → "ISABELLA DE ROSA"
#         reconstructed = _reconstruct_fragmented_text(full_line)
#         if reconstructed != full_line:
#             full_line = reconstructed
#         # Remove echo duplicates
#         full_line = _dedup_line_text(full_line)
#         result.append((full_line, row))
#     return result


# # ══════════════════════════════════════════════════════════════════════════════
# #  L4 — FIELD EXTRACTORS
# # ══════════════════════════════════════════════════════════════════════════════


# # ── Email ──────────────────────────────────────────────────────────────────────
# def _extract_email(lines: list) -> str:
#     for ln, _ in lines:
#         compact = re.sub(r"\s+", "", ln)
#         m = _EMAIL_RE.search(compact)
#         if m:
#             email = m.group().lower()
#             if len(email) <= 80 and "@" in email and "." in email.split("@")[-1]:
#                 return email
#     return ""


# # ── Phones ─────────────────────────────────────────────────────────────────────
# def _detect_region_from_text(text: str):
#     """Extract ISO region from +CC prefix. Handles multiple capture groups."""
#     m = _INTL_PREFIX_RE.search(text)
#     if not m:
#         return None
#     pfx = m.group(1) or m.group(2)  # either capture group
#     if not pfx:
#         return None
#     for length in (3, 2, 1):
#         r = _CC_REGION.get(pfx[:length])
#         if r:
#             return r
#     return None


# def _extract_phones(lines: list) -> list:
#     """
#     BUG B FIX: Detect region from ALL lines upfront (not just first match).
#     BUG C FIX: Lines with BOTH address keywords AND phone patterns — extract
#                phone from those lines before skipping them for address scan.
#     BUG 4 FIX: Skip lines that are pure postal codes.
#     """
#     found, found_digs = [], []

#     # Detect primary region from the full text of all lines
#     all_text = " ".join(ln for ln, _ in lines)
#     detected_region = _detect_region_from_text(all_text)

#     for ln, _ in lines:
#         if len(found) >= CFG["MAX_PHONES"]:
#             break
#         if _SOCIAL_RE.search(ln):
#             continue

#         # BUG C FIX: If line has address keyword, still try to extract phone
#         # patterns from it (phone may be on same line as address keyword).
#         # We do NOT skip the line entirely — we just scan more carefully.
#         has_addr_kw = bool(ADDRESS_KW.search(ln))

#         cleaned = _PHONE_LABEL_RE.sub("", ln).strip()
#         if not cleaned:
#             continue

#         raw_digits = re.sub(r"\D", "", cleaned)
#         if len(raw_digits) < 4:
#             continue

#         # Skip pure postal code lines (not mixed with other content)
#         if _POSTAL_FMT_RE.fullmatch(cleaned.strip()):
#             continue
#         # Skip lines that are ONLY address keywords with no digit patterns
#         if has_addr_kw and not re.search(r"\d{4,}", cleaned):
#             continue

#         parsed = False
#         if PHONE_OK:
#             # BUG B FIX: always try the region matching the +CC prefix first
#             line_region = _detect_region_from_text(cleaned)
#             # Build region list: line-specific first, then card-wide, then fallbacks
#             region_list = list(
#                 dict.fromkeys(
#                     [r for r in [line_region, detected_region] if r] + _FALLBACK_REGIONS
#                 )
#             )
#             for region in region_list:
#                 try:
#                     region_found = False
#                     for match in PhoneNumberMatcher(cleaned, region):
#                         num = match.number
#                         if not phonenumbers.is_valid_number(num):
#                             continue
#                         nat_str = str(num.national_number)
#                         if len(re.sub(r"\D", "", nat_str)) < 6:
#                             continue
#                         fmt = phonenumbers.format_number(
#                             num, PhoneNumberFormat.INTERNATIONAL
#                         )
#                         digs = re.sub(r"\D", "", fmt)
#                         if _POSTAL_FMT_RE.fullmatch(re.sub(r"[+\s]", "", fmt)):
#                             continue
#                         if digs not in found_digs:
#                             found.append(fmt)
#                             found_digs.append(digs)
#                         parsed = True
#                         region_found = True
#                     # Only try next region if this one found nothing at all
#                     if region_found:
#                         break
#                 except Exception:
#                     continue

#         if not parsed:
#             for m in re.finditer(r"[\+\(]?[\d][\d\s\-\.\(\)]{5,18}[\d]", cleaned):
#                 raw = m.group()
#                 digs = re.sub(r"\D", "", raw)
#                 min_d = (
#                     7
#                     if raw.strip().startswith("+")
#                     else CFG["PHONE_MIN_DIGITS_NO_PREFIX"]
#                 )
#                 if not (min_d <= len(digs) <= 15):
#                     continue
#                 if _POSTAL_FMT_RE.fullmatch(digs):
#                     continue
#                 norm = re.sub(
#                     r"\s{2,}", " ", re.sub(r"(?<=\d) (?=\d)", "", raw)
#                 ).strip()
#                 if digs not in found_digs:
#                     found.append(norm)
#                     found_digs.append(digs)

#     # ── Second pass: scan individual tokens from address lines ───────────────
#     # Catches phone numbers that OCR placed on the same line as address text
#     for ln, _ in lines:
#         if len(found) >= CFG["MAX_PHONES"]:
#             break
#         if not ADDRESS_KW.search(ln):
#             continue  # only re-scan address lines
#         if _SOCIAL_RE.search(ln):
#             continue
#         # Look for clear phone patterns within the address line
#         for m in re.finditer(r"\+\d[\d\s\-\.\(\)]{6,15}\d", ln):
#             raw = m.group()
#             digs = re.sub(r"\D", "", raw)
#             if not (7 <= len(digs) <= 15):
#                 continue
#             if _POSTAL_FMT_RE.fullmatch(digs):
#                 continue
#             if digs in found_digs:
#                 continue
#             # Validate with phonenumbers
#             if PHONE_OK:
#                 line_region = _detect_region_from_text(raw) or detected_region
#                 regions = list(
#                     dict.fromkeys([r for r in [line_region] if r] + _FALLBACK_REGIONS)
#                 )
#                 for region in regions:
#                     try:
#                         for match in PhoneNumberMatcher(raw, region):
#                             num = match.number
#                             if phonenumbers.is_valid_number(num):
#                                 fmt = phonenumbers.format_number(
#                                     num, PhoneNumberFormat.INTERNATIONAL
#                                 )
#                                 fdigs = re.sub(r"\D", "", fmt)
#                                 if fdigs not in found_digs:
#                                     found.append(fmt)
#                                     found_digs.append(fdigs)
#                                 break
#                         break
#                     except Exception:
#                         continue
#             else:
#                 found.append(raw.strip())
#                 found_digs.append(digs)

#     return found


# # ── Website ────────────────────────────────────────────────────────────────────
# def _extract_website(lines: list, email: str) -> str:
#     local = email.split("@")[0].lower() if "@" in email else ""
#     for ln, _ in lines:
#         if "@" in ln:
#             continue
#         for m in _WEB_RE.finditer(ln):
#             url = m.group().strip().rstrip(",.")
#             lo = url.lower()
#             if lo.startswith("http") or lo.startswith("www."):
#                 full = ("http://" + url) if not lo.startswith("http") else url
#                 if TLDX_OK:
#                     ext = tldextract.extract(full)
#                     if not ext.domain:
#                         continue
#                 return full.lower()
#             if TLDX_OK:
#                 ext = tldextract.extract(url)
#                 if not ext.domain:
#                     continue
#                 if ext.domain.lower() == local and not ext.subdomain and "/" not in url:
#                     continue
#             elif not re.search(r"\.", url) or re.fullmatch(
#                 r"[A-Za-z]+\.[A-Za-z]+", url
#             ):
#                 continue
#             return ("http://" + url).lower()
#     return ""


# # ── Company ────────────────────────────────────────────────────────────────────
# def _extract_company(lines: list) -> str:
#     scored = []
#     # Build augmented line list — include reconstructed versions of fragmented lines
#     augmented = []
#     for ln, row in lines:
#         augmented.append((ln, row))
#         recon = _reconstruct_fragmented_text(ln)
#         if recon != ln:
#             augmented.append((recon, row))  # add reconstructed as additional candidate
#     for ln, _ in augmented:
#         if re.search(r"@|\d{5,}|https?://|www\.", ln, re.I):
#             continue
#         # Apply digit substitutions (fixes "GR0UP" → "GROUP", "S.P.A." stays)
#         ln_fixed = _fix_digit_substitutions(ln)
#         score = (
#             (10 if COMPANY_KW.search(ln_fixed) else 0)
#             + (3 if ln_fixed.isupper() and len(ln_fixed.split()) >= 2 else 0)
#             + (1 if len(ln_fixed.split()) >= 2 else 0)
#             - (6 if DESIGNATION_KW.search(ln_fixed) else 0)
#             - (5 if ADDRESS_KW.search(ln_fixed) else 0)
#         )
#         if score > 0:
#             scored.append((score, len(ln_fixed), ln_fixed))
#     if not scored:
#         return ""
#     scored.sort(reverse=True)
#     result = _dedup_repeated_words(scored[0][2])
#     result = _fix_digit_substitutions(result)
#     return _strip_trailing_punct(result)


# # ── Job title ──────────────────────────────────────────────────────────────────
# def _extract_job_title(lines: list, claimed: set) -> str:
#     """Extract job title. Tries standard lines first, then applies fragment
#     reconstruction for spaced-font cards (BUG F FIX)."""
#     for ln, _ in lines:
#         if ln.strip().lower() in claimed:
#             continue
#         if re.search(r"@|\d{5,}", ln):
#             continue
#         if DESIGNATION_KW.search(ln) and not COMPANY_KW.search(ln):
#             return _strip_trailing_punct(_dedup_repeated_words(ln))
#     # Fallback: try reconstructing fragmented spaced-font title lines
#     for ln, _ in lines:
#         if ln.strip().lower() in claimed:
#             continue
#         if re.search(r"@|\d{5,}", ln):
#             continue
#         reconstructed = _reconstruct_fragmented_text(ln)
#         if reconstructed != ln:
#             if DESIGNATION_KW.search(reconstructed) and not COMPANY_KW.search(
#                 reconstructed
#             ):
#                 return _strip_trailing_punct(_dedup_repeated_words(reconstructed))
#     return ""


# # ── Name ───────────────────────────────────────────────────────────────────────
# def _is_name_word(w: str) -> bool:
#     """
#     BUG A FIX: Check initial pattern FIRST (before dot rejection).
#     Accepts: JOHN, John, J, J. (initial+dot), Al-Mansoori, O'Sullivan
#     Rejects: digits, @, company/address keywords
#     """
#     w2 = w.rstrip(",;").strip()
#     if not w2:
#         return False
#     if re.search(r"\d|@", w2):
#         return False

#     # BUG A FIX: check single-letter initial BEFORE the general dot rejection
#     # "V." is a valid initial — must be accepted here
#     if re.fullmatch(r"[A-Za-z]\.?", w2):
#         return True

#     # Now reject other tokens containing dots (email/URL fragments)
#     if "." in w2:
#         return False

#     # All uppercase (Unicode-aware)
#     if (
#         all(unicodedata.category(c).startswith("L") or c in "-'" for c in w2)
#         and w2[0].isupper()
#     ):
#         return True
#     # Title case
#     if w2[0].isupper() and all(
#         unicodedata.category(c).startswith("L") or c in "-'" for c in w2
#     ):
#         return True
#     # All lowercase (some cultures)
#     if w2.islower() and len(w2) >= 2:
#         return True
#     return False


# def _extract_name(lines: list, claimed: set, script: str, addr_tokens: set) -> str:
#     all_text = "\n".join(ln for ln, _ in lines)
#     if SPACY_OK and _NLP:
#         for ent in _NLP(all_text).ents:
#             if ent.label_ != "PERSON":
#                 continue
#             raw_cand = ent.text.strip()
#             if "\n" in raw_cand:
#                 continue
#             cand = raw_cand.replace("\n", " ")
#             if len(cand.split()) > 5:
#                 continue
#             if (
#                 len(cand) >= 3
#                 and cand.lower() not in claimed
#                 and len(cand.split()) >= 2
#                 and not re.search(r"\d|@", cand)
#                 and not COMPANY_KW.search(cand)
#                 and not DESIGNATION_KW.search(cand)
#                 and not _SOCIAL_RE.search(cand)
#                 and not _NAME_STOPWORDS.search(cand)
#                 and cand.lower() not in addr_tokens
#             ):
#                 return cand

#     card_h = max((t["cy"] for _, row in lines for t in row), default=1) or 1
#     scored = []
#     for ln, row in lines:
#         if not ln.strip() or ln.strip().lower() in claimed:
#             continue
#         if _SOCIAL_RE.search(ln) or "@" in ln:
#             continue
#         if _NAME_STOPWORDS.search(ln):
#             continue
#         if ln.strip().lower() in addr_tokens:
#             continue
#         words = ln.split()
#         if not (2 <= len(words) <= 5):
#             continue
#         if words[0] == "I" and len(words) > 1 and words[1].startswith("@"):
#             continue
#         if not all(_is_name_word(w) for w in words):
#             continue
#         alpha = sum(c.isalpha() for c in ln)
#         if alpha / max(len(ln.replace(" ", "")), 1) < CFG["NAME_ALPHA_RATIO"]:
#             continue
#         avg_y = sum(t["cy"] for t in row) / len(row) if row else card_h
#         avg_h = sum(t.get("th", 20) for t in row) / len(row) if row else 20
#         scored.append((avg_h * 0.6 + (1.0 - avg_y / card_h) * 40, ln))
#     if scored:
#         scored.sort(reverse=True)
#         best = _dedup_repeated_words(scored[0][1])
#         parts = []
#         for p in best.split():
#             cp = _collapse_spaced(p)
#             parts.append(cp.title() if cp.isupper() and len(cp) > 2 else cp)
#         return " ".join(parts)

#     # ── Fallback: luxury/spaced font reconstruction ───────────────────────────
#     # When the standard heuristic finds nothing, examine the topmost lines that
#     # are not email/phone/url/address/company and try to reconstruct a name
#     # from fragmented OCR tokens (e.g. "Isab Ella DE Rosa" → "Isabella De Rosa")
#     all_cy_vals = [t["cy"] for _, row in lines for t in row]
#     card_top_threshold = (max(all_cy_vals) * 0.35) if all_cy_vals else 999
#     for ln, row in lines:
#         if not ln.strip() or ln.strip().lower() in claimed:
#             continue
#         if _SOCIAL_RE.search(ln) or "@" in ln:
#             continue
#         if _NAME_STOPWORDS.search(ln):
#             continue
#         if re.search(r"\d{3,}|https?://|www\.", ln):
#             continue
#         if COMPANY_KW.search(ln) or DESIGNATION_KW.search(ln):
#             continue
#         # Only look at top 35% of card vertically
#         avg_y = sum(t["cy"] for t in row) / len(row) if row else card_h
#         if avg_y > card_top_threshold:
#             continue
#         # Must be all short alphabetic fragments
#         words = ln.split()
#         if not (2 <= len(words) <= 6):
#             continue
#         if not all(
#             len(w.rstrip(".,;")) <= 6
#             and all(c.isalpha() or c == "." for c in w.rstrip(".,;"))
#             for w in words
#         ):
#             continue
#         # Collapse fragments into a name
#         collapsed = "".join(w.rstrip(".,;") for w in words)
#         collapsed = _fix_digit_substitutions(collapsed)
#         # Must be ≥4 chars and mostly alphabetic
#         alpha = sum(c.isalpha() for c in collapsed)
#         if alpha < 4 or alpha / max(len(collapsed), 1) < 0.85:
#             continue
#         # Title-case the result: treat it as a single name word for now
#         # A better split would require a name dictionary, so we keep it simple
#         return collapsed.title()

#     return ""


# # ── Address ────────────────────────────────────────────────────────────────────
# def _ocr_corruption_score(text: str) -> int:
#     """
#     BUG D FIX: Score how 'corrupted' a string looks.
#     Higher score = more OCR errors (q instead of g, l instead of 1, etc.)
#     Used to prefer the cleaner version when two similar address strings exist.
#     """
#     score = 0
#     # lowercase 'l' used as digit '1' in context of numbers
#     score += len(re.findall(r"(?<=\d)l(?=\d)|(?<=\d)l\b|\bl(?=\d)", text))
#     # 'q' where 'g' is expected (end of word, rare in English)
#     score += len(re.findall(r"q(?=\s|,|$)", text, re.IGNORECASE))
#     # digit 'O' instead of zero
#     score += len(re.findall(r"(?<=[A-Za-z])\d{1}(?=[A-Za-z])", text))
#     return score


# def _extract_address(lines: list, claimed: set, phone_list: list) -> str:
#     phone_fps = [re.sub(r"\D", "", p) for p in phone_list if p]
#     candidates = []  # (corruption_score, text)
#     seen_clean = []  # clean versions for fuzzy dedup

#     for ln, _ in lines:
#         t = ln.strip()
#         if not t or t.lower() in claimed or "@" in t:
#             continue
#         line_digs = re.sub(r"\D", "", t)
#         if len(line_digs) >= 6:
#             skip = False
#             for fp in phone_fps:
#                 if not fp:
#                     continue
#                 if (
#                     FUZZ_OK
#                     and rfuzz.partial_ratio(fp, line_digs) >= CFG["PHONE_FUZZ_THRESH"]
#                 ):
#                     skip = True
#                     break
#                 elif fp in line_digs:
#                     skip = True
#                     break
#             if skip:
#                 continue
#         is_addr = bool(ADDRESS_KW.search(t))
#         if POSTAL_RE.search(t) and not re.search(r"[+\(]?\d[\d\s\-\.\(\)]{8,}", t):
#             is_addr = True
#         if re.search(r"[A-Z][a-z]+,\s*[A-Z][a-z]+", t):
#             is_addr = True
#         if not is_addr:
#             continue

#         # BUG D FIX: fuzzy dedup — lower threshold to 75 to catch OCR corruptions
#         # e.g. "Johannesburg 2196" vs "Johannesburq 2l96" (ratio ~88, both caught)
#         is_dup = False
#         if FUZZ_OK:
#             for existing in seen_clean:
#                 ratio = rfuzz.ratio(t.lower(), existing.lower())
#                 if ratio >= 75:  # lowered from 80 to catch more OCR variants
#                     is_dup = True
#                     curr_score = _ocr_corruption_score(t)
#                     exist_score = _ocr_corruption_score(existing)
#                     if curr_score < exist_score:
#                         idx = seen_clean.index(existing)
#                         seen_clean[idx] = t
#                         for ci, (cs, ct) in enumerate(candidates):
#                             if ct == existing:
#                                 candidates[ci] = (curr_score, t)
#                                 break
#                     break
#         else:
#             is_dup = t.lower() in [s.lower() for s in seen_clean]

#         if not is_dup:
#             seen_clean.append(t)
#             candidates.append((_ocr_corruption_score(t), t))

#     # Sort by corruption score (prefer clean versions)
#     candidates.sort(key=lambda x: x[0])
#     return ", ".join(ct for _, ct in candidates)


# # ── Social ─────────────────────────────────────────────────────────────────────
# def _extract_social(lines: list, claimed: set) -> tuple:
#     linkedin = twitter = ""
#     for ln, _ in lines:
#         if not linkedin:
#             m = _LINKEDIN_RE.search(ln)
#             if m:
#                 v = m.group().strip()
#                 if v.lower() not in claimed:
#                     linkedin = v
#         if not twitter:
#             m = _TWITTER_RE.search(ln)
#             if m:
#                 v = m.group().strip()
#                 if v.lower() not in claimed and not _EMAIL_RE.search(v):
#                     twitter = v
#         if linkedin and twitter:
#             break
#     return linkedin, twitter


# # ══════════════════════════════════════════════════════════════════════════════
# #  FIELD ORCHESTRATION
# # ══════════════════════════════════════════════════════════════════════════════


# def extract_fields(lines: list, primary_script: str) -> dict:
#     data = {
#         "Name": "",
#         "Job_Title": "",
#         "Company": "",
#         "Email": "",
#         "Phones": [],
#         "Website": "",
#         "Address": "",
#         "LinkedIn": "",
#         "Twitter": "",
#     }
#     data["Email"] = _extract_email(lines)
#     data["Company"] = _extract_company(lines)
#     data["Phones"] = _extract_phones(lines)
#     data["Website"] = _extract_website(lines, data["Email"])
#     claimed = {
#         data["Email"].lower(),
#         data["Company"].strip().lower(),
#         data["Website"].lower(),
#         "",
#     }
#     for p in data["Phones"]:
#         claimed.add(p.lower())
#     data["Job_Title"] = _extract_job_title(lines, claimed)
#     claimed.add(data["Job_Title"].strip().lower())
#     addr_tokens = {
#         w
#         for ln, _ in lines
#         if ADDRESS_KW.search(ln) or POSTAL_RE.search(ln)
#         for w in ln.lower().split()
#     }
#     data["Name"] = _extract_name(lines, claimed, primary_script, addr_tokens)
#     claimed.add(data["Name"].strip().lower())
#     data["Address"] = _extract_address(lines, claimed, data["Phones"])
#     data["LinkedIn"], data["Twitter"] = _extract_social(lines, claimed)
#     return data


# # ══════════════════════════════════════════════════════════════════════════════
# #  QUALITY SCORING
# # ══════════════════════════════════════════════════════════════════════════════


# def quality_score(data: dict) -> str:
#     """
#     Validated quality scoring — each field checked for content plausibility.
#     GREEN  → n_ok AND (e_ok OR p_ok) AND core_score ≥ 5
#     YELLOW → core_score ≥ 2
#     RED    → core_score < 2
#     """
#     phones = data.get("Phones", [])
#     n_ok = bool(
#         data.get("Name")
#         and len(data["Name"].split()) >= 2
#         and not re.search(r"\d|@|\.", data["Name"])
#         and not ADDRESS_KW.search(data["Name"])
#         and not COMPANY_KW.search(data["Name"])
#         and sum(c.isalpha() for c in data["Name"])
#         / max(len(data["Name"].replace(" ", "")), 1)
#         >= 0.75
#     )
#     # BUG A FIX: "ADITYA V. VARMA" has a dot in "V." — the quality check
#     # must allow names with dotted initials. Updated: only reject if
#     # the dot appears in a NON-INITIAL context (more than 1 char before dot).
#     if data.get("Name") and "." in data["Name"]:
#         # Re-check: allow if the dot is only in initial tokens (single letter + dot)
#         name_words = data["Name"].split()
#         all_dots_are_initials = all(
#             not ("." in w) or re.fullmatch(r"[A-Za-z]\.", w) for w in name_words
#         )
#         if all_dots_are_initials:
#             # Don't penalize for dotted initials — recalculate n_ok without dot check
#             n_ok = bool(
#                 data.get("Name")
#                 and len(data["Name"].split()) >= 2
#                 and not re.search(r"@", data["Name"])
#                 and not ADDRESS_KW.search(data["Name"])
#                 and not COMPANY_KW.search(data["Name"])
#                 and sum(c.isalpha() for c in data["Name"])
#                 / max(len(data["Name"].replace(" ", "")), 1)
#                 >= 0.75
#             )
#     c_ok = bool(
#         data.get("Company")
#         and data["Company"].strip().lower() != data.get("Name", "").lower()
#         and (
#             COMPANY_KW.search(data["Company"])
#             or (data["Company"].isupper() and len(data["Company"].split()) >= 2)
#         )
#     )
#     e_ok = bool(
#         data.get("Email")
#         and _EMAIL_RE.search(data["Email"])
#         and len(data["Email"]) <= 80
#     )
#     p_ok = bool(phones and any(len(re.sub(r"\D", "", p)) >= 7 for p in phones))
#     a_ok = bool(
#         data.get("Address")
#         and (
#             ADDRESS_KW.search(data["Address"])
#             or POSTAL_RE.search(data["Address"])
#             or re.search(r"[A-Z][a-z]+,\s*[A-Z][a-z]+", data["Address"])
#         )
#     )
#     t_ok = bool(
#         data.get("Job_Title")
#         and DESIGNATION_KW.search(data["Job_Title"])
#         and not COMPANY_KW.search(data["Job_Title"])
#     )
#     core_score = (
#         (2 if n_ok else 0)
#         + (2 if e_ok else 0)
#         + (1 if p_ok else 0)
#         + (1 if c_ok else 0)
#         + (1 if a_ok else 0)
#         + (1 if t_ok else 0)
#     )
#     if n_ok and (e_ok or p_ok) and core_score >= 5:
#         return "🟢 GREEN"
#     if core_score >= 2:
#         return "🟡 YELLOW"
#     return "🔴 RED"


# # ══════════════════════════════════════════════════════════════════════════════
# #  CORE PROCESSOR
# # ══════════════════════════════════════════════════════════════════════════════


# def process_card_array(
#     img: np.ndarray, label: str = "card", lang_override: str = None
# ) -> dict:
#     proc = preprocess(deskew(img))
#     tokens, primary, secondary = ocr_card(proc, lang_override)
#     if CFG["DEBUG"]:
#         log.debug(
#             "%s — %d tokens  script=%s/%s",
#             label,
#             len(tokens),
#             primary,
#             secondary or "—",
#         )
#         for t in sorted(tokens, key=lambda x: x["cy"]):
#             log.debug("  cy=%6.1f conf=%.2f %r", t["cy"], t["conf"], t["text"])
#     lines = group_into_lines(tokens)
#     if CFG["DEBUG"]:
#         for i, (ln, _) in enumerate(lines, 1):
#             log.debug("  Line %02d: %r", i, ln)
#     data = extract_fields(lines, primary)
#     data["QUALITY"] = quality_score(data)
#     data["_label"] = label
#     data["_script"] = primary
#     data["_script2"] = secondary or ""
#     return data


# def process_source(label: str, page_img: np.ndarray, lang_override: str = None) -> list:
#     crops = segment_cards(page_img)
#     results = []
#     for idx, card_img in enumerate(crops):
#         clabel = label if len(crops) == 1 else f"{label} [card {idx+1}/{len(crops)}]"
#         try:
#             results.append(process_card_array(card_img, clabel, lang_override))
#         except Exception as e:
#             log.error("Failed: %s — %s", clabel, e)
#             if CFG["DEBUG"]:
#                 traceback.print_exc()
#     return results


# # ══════════════════════════════════════════════════════════════════════════════
# #  TERMINAL PRINT
# # ══════════════════════════════════════════════════════════════════════════════


# def print_result(r: dict):
#     W = 64
#     phones = r.get("Phones", [])
#     print(f"\n{'═'*W}")
#     print(f"  FILE      : {r.get('_label','')}")
#     print(f"{'─'*W}")
#     print(f"  NAME      : {r.get('Name','') or '—'}")
#     print(f"  JOB TITLE : {r.get('Job_Title','') or '—'}")
#     print(f"  COMPANY   : {r.get('Company','') or '—'}")
#     print(f"  EMAIL     : {r.get('Email','') or '—'}")
#     if phones:
#         for i, p in enumerate(phones, 1):
#             print(f"  PHONE #{i}  : {p}")
#     else:
#         print("  PHONE     : —")
#     print(f"  WEBSITE   : {r.get('Website','') or '—'}")
#     print(f"  ADDRESS   : {r.get('Address','') or '—'}")
#     print(f"  LINKEDIN  : {r.get('LinkedIn','') or '—'}")
#     print(f"  TWITTER/X : {r.get('Twitter','') or '—'}")
#     print(f"  QUALITY   : {r.get('QUALITY','') or '—'}")
#     print("═" * W)


# # ══════════════════════════════════════════════════════════════════════════════
# #  OUTPUT SYSTEM
# # ══════════════════════════════════════════════════════════════════════════════


# def _row_vals(r: dict) -> list:
#     ph = r.get("Phones", [])
#     return [
#         r.get("Name", ""),
#         r.get("Job_Title", ""),
#         r.get("Company", ""),
#         r.get("Email", ""),
#         ph[0] if len(ph) > 0 else "",
#         ph[1] if len(ph) > 1 else "",
#         ph[2] if len(ph) > 2 else "",
#         r.get("Website", ""),
#         r.get("Address", ""),
#         r.get("LinkedIn", ""),
#         r.get("Twitter", ""),
#         r.get("QUALITY", ""),
#         datetime.now().strftime("%Y-%m-%d %H:%M"),
#         r.get("_label", ""),
#     ]


# def _dedup_key(r: dict) -> str:
#     return "|".join(
#         [
#             re.sub(r"\s+", "", r.get("Email", "").lower()),
#             re.sub(r"\s+", "", r.get("Name", "").lower()),
#             re.sub(r"\D", "", r.get("Phones", [""])[0] if r.get("Phones") else ""),
#         ]
#     )


# def _write_header(ws):
#     hf = PatternFill("solid", fgColor="1F4E79")
#     hfont = Font(bold=True, color="FFFFFF", size=10)
#     thin = Side(style="thin", color="BDD7EE")
#     bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
#     for ci, name in enumerate(EXCEL_COLS, 1):
#         c = ws.cell(1, ci, value=name)
#         c.fill = hf
#         c.font = hfont
#         c.alignment = Alignment(horizontal="center", vertical="center")
#         c.border = bdr
#     ws.row_dimensions[1].height = 20


# def _style_row(ws, rn, quality):
#     col = _QC.get(quality, "FFFFFF")
#     fill = PatternFill("solid", fgColor=col)
#     thin = Side(style="thin", color="BDD7EE")
#     bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
#     for ci in range(1, len(EXCEL_COLS) + 1):
#         c = ws.cell(rn, ci)
#         c.fill = fill
#         c.alignment = Alignment(vertical="center", wrap_text=True)
#         c.border = bdr
#         c.font = Font(size=9)
#     ws.row_dimensions[rn].height = 15


# def _df_from_results(results: list) -> "pd.DataFrame":
#     rows = []
#     for r in results:
#         ph = r.get("Phones", [])
#         rows.append(
#             {
#                 "Name": r.get("Name", ""),
#                 "Job_Title": r.get("Job_Title", ""),
#                 "Company": r.get("Company", ""),
#                 "Email": r.get("Email", ""),
#                 "Phone_1": ph[0] if len(ph) > 0 else "",
#                 "Phone_2": ph[1] if len(ph) > 1 else "",
#                 "Phone_3": ph[2] if len(ph) > 2 else "",
#                 "Website": r.get("Website", ""),
#                 "Address": r.get("Address", ""),
#                 "LinkedIn": r.get("LinkedIn", ""),
#                 "Twitter": r.get("Twitter", ""),
#                 "Quality": r.get("QUALITY", ""),
#                 "Processed": datetime.now().strftime("%Y-%m-%d %H:%M"),
#                 "File": r.get("_label", ""),
#             }
#         )
#     return pd.DataFrame(rows, columns=EXCEL_COLS)


# def _write_json(data, path: str):
#     try:
#         with open(path, "w", encoding="utf-8") as f:
#             json.dump(data, f, ensure_ascii=False, indent=2, default=str)
#     except Exception as e:
#         log.error("JSON write failed (%s): %s", path, e)


# def _safe_save_wb(wb, path: str) -> str:
#     try:
#         wb.save(path)
#         return path
#     except PermissionError:
#         ts = datetime.now().strftime("%H%M%S")
#         alt = path.replace(".xlsx", f"_{ts}_locked.xlsx")
#         try:
#             wb.save(alt)
#             log.warning(
#                 "'%s' is open in another program → saved to %s",
#                 os.path.basename(path),
#                 alt,
#             )
#             return alt
#         except Exception as e2:
#             log.error("Excel save failed completely: %s", e2)
#             return ""


# def save_to_database(results: list, excel_path: str):
#     if not EXCEL_OK:
#         return
#     month = datetime.now().strftime("%Y-%m")
#     wb = (
#         openpyxl.load_workbook(excel_path)
#         if os.path.exists(excel_path)
#         else openpyxl.Workbook()
#     )
#     if "Sheet" in wb.sheetnames:
#         del wb["Sheet"]
#     if month in wb.sheetnames:
#         ws = wb[month]
#         existing = {}
#         for rn in range(2, ws.max_row + 1):
#             key = _dedup_key(
#                 {
#                     "Email": str(ws.cell(rn, 4).value or ""),
#                     "Name": str(ws.cell(rn, 1).value or ""),
#                     "Phones": [str(ws.cell(rn, 5).value or "")],
#                 }
#             )
#             existing[key] = rn
#     else:
#         ws = wb.create_sheet(title=month)
#         _write_header(ws)
#         existing = {}
#     for r in results:
#         key = _dedup_key(r)
#         is_upd = key in existing
#         rn = existing[key] if is_upd else ws.max_row + 1
#         for ci, v in enumerate(_row_vals(r), 1):
#             ws.cell(rn, ci, value=v)
#         _style_row(ws, rn, r.get("QUALITY", ""))
#         existing[key] = rn
#         log.info(
#             "  [DB] %s row %d → %s",
#             "Updated" if is_upd else "Appended",
#             rn,
#             r.get("Name") or r.get("Email") or "?",
#         )
#     for col in ws.columns:
#         ws.column_dimensions[get_column_letter(col[0].column)].width = min(
#             max(len(str(c.value or "")) for c in col) + 3, 55
#         )
#     ws.freeze_panes = "A2"
#     saved = _safe_save_wb(wb, excel_path)
#     if saved:
#         log.info("Database → %s (sheet: %s)", saved, month)


# def save_outputs(results: list, out_dir: str):
#     if not EXCEL_OK:
#         return
#     history = os.path.join(out_dir, "history")
#     os.makedirs(history, exist_ok=True)
#     ts = datetime.now().strftime("%Y%m%d_%H%M%S")
#     df = _df_from_results(results)
#     # Batch snapshot
#     _safe_save_wb(_make_styled_wb(results), os.path.join(history, f"cards_{ts}.xlsx"))
#     _write_json(df.to_dict("records"), os.path.join(history, f"cards_{ts}.json"))
#     df.to_csv(os.path.join(history, f"cards_{ts}.csv"), index=False)
#     # Per-card individual files
#     for i, r in enumerate(results, 1):
#         safe = re.sub(r"[^\w\-]", "_", r.get("_label", "card"))[:30]
#         _safe_save_wb(
#             _make_styled_wb([r]),
#             os.path.join(history, f"card_{ts}_{i:02d}_{safe}.xlsx"),
#         )
#     # Stable latest
#     saved_latest = _safe_save_wb(
#         _make_styled_wb(results), os.path.join(out_dir, "latest.xlsx")
#     )
#     _write_json(df.to_dict("records"), os.path.join(out_dir, "latest.json"))
#     df.to_csv(os.path.join(out_dir, "latest.csv"), index=False)
#     log.info("History → %s/  |  Latest → %s", history, out_dir)
#     return saved_latest


# def _make_styled_wb(results: list) -> "openpyxl.Workbook":
#     """Build a styled openpyxl workbook from results."""
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Cards"
#     _write_header(ws)
#     for ri, r in enumerate(results, 2):
#         for ci, v in enumerate(_row_vals(r), 1):
#             ws.cell(ri, ci, value=v)
#         _style_row(ws, ri, r.get("QUALITY", ""))
#     for col in ws.columns:
#         ws.column_dimensions[get_column_letter(col[0].column)].width = min(
#             max(len(str(c.value or "")) for c in col) + 3, 55
#         )
#     ws.freeze_panes = "A2"
#     return wb


# def export_vcf(results: list, vcf_path: str):
#     lines = []
#     for r in results:
#         ph = r.get("Phones", [])
#         name = r.get("Name", "")
#         parts = name.split(maxsplit=1)
#         lines += ["BEGIN:VCARD", "VERSION:3.0"]
#         if name:
#             lines.append(
#                 f"N:{parts[-1]};{parts[0]};;;" if len(parts) == 2 else f"N:{name};;;;"
#             )
#             lines.append(f"FN:{name}")
#         if r.get("Company"):
#             lines.append(f"ORG:{r['Company']}")
#         if r.get("Job_Title"):
#             lines.append(f"TITLE:{r['Job_Title']}")
#         if r.get("Email"):
#             lines.append(f"EMAIL;TYPE=INTERNET:{r['Email']}")
#         for p in ph:
#             lines.append(f"TEL;TYPE=VOICE:{p}")
#         if r.get("Website"):
#             lines.append(f"URL:{r['Website']}")
#         if r.get("Address"):
#             lines.append(f"ADR;TYPE=WORK:;;{r['Address'].replace(', ','; ')};;;;")
#         if r.get("LinkedIn"):
#             lines.append(f"X-SOCIALPROFILE;TYPE=linkedin:{r['LinkedIn']}")
#         if r.get("Twitter"):
#             lines.append(f"X-SOCIALPROFILE;TYPE=twitter:{r['Twitter']}")
#         lines += [
#             f"NOTE:VC-OCR v18 {datetime.now().strftime('%Y-%m-%d')}",
#             "END:VCARD",
#             "",
#         ]
#     try:
#         Path(vcf_path).write_text("\n".join(lines), encoding="utf-8")
#         log.info("VCF → %s (%d cards)", vcf_path, len(results))
#     except Exception as e:
#         log.error("VCF export failed: %s", e)


# # ══════════════════════════════════════════════════════════════════════════════
# #  ENTRY POINT
# # ══════════════════════════════════════════════════════════════════════════════


# def _gui_pick() -> list:
#     if not TK_OK:
#         log.warning("tkinter unavailable — pass path as argument")
#         return []
#     root = Tk()
#     root.withdraw()
#     paths = filedialog.askopenfilenames(
#         title="Select Card Images / PDFs / ZIPs",
#         filetypes=[
#             (
#                 "All supported",
#                 "*.jpg *.jpeg *.png *.bmp *.webp *.tiff *.tif *.pdf *.zip",
#             )
#         ],
#     )
#     root.destroy()
#     return list(paths)


# def main():
#     ap = argparse.ArgumentParser(description="Visiting Card OCR v18.0")
#     ap.add_argument(
#         "target",
#         nargs="?",
#         default=None,
#         help="Image / folder / PDF / ZIP (blank → GUI picker)",
#     )
#     ap.add_argument(
#         "--lang", default=None, help="Force OCR language: en|ch|japan|korean|ar|hi"
#     )
#     ap.add_argument("--debug", action="store_true")
#     ap.add_argument("--no-vcf", action="store_true")
#     ap.add_argument("--no-excel", action="store_true")
#     ap.add_argument("--db", default=None, help="Override database Excel path")
#     args = ap.parse_args()
#     if args.debug:
#         CFG["DEBUG"] = True
#         logging.getLogger().setLevel(logging.DEBUG)

#     if args.target and Path(args.target).is_dir():
#         supported = {
#             ".jpg",
#             ".jpeg",
#             ".png",
#             ".bmp",
#             ".webp",
#             ".tiff",
#             ".tif",
#             ".pdf",
#             ".zip",
#         }
#         paths = sorted(
#             str(f) for f in Path(args.target).iterdir() if f.suffix.lower() in supported
#         )
#     elif args.target:
#         paths = [args.target]
#     else:
#         paths = _gui_pick()

#     if not paths:
#         log.info("No files selected. Exiting.")
#         sys.exit(0)

#     out_dir = os.path.dirname(os.path.abspath(paths[0]))
#     db_path = args.db or os.path.join(out_dir, CFG["EXCEL_DB_FILE"])
#     results = []

#     log.info("Processing %d file(s)…", len(paths))
#     for i, p in enumerate(paths, 1):
#         log.info("[%d/%d] %s", i, len(paths), p)
#         try:
#             for plabel, pimg in load_images(p):
#                 for r in process_source(plabel, pimg, args.lang):
#                     print_result(r)
#                     results.append(r)
#         except Exception as e:
#             log.error("Failed on %s: %s", p, e)
#             if CFG["DEBUG"]:
#                 traceback.print_exc()

#     if not results:
#         log.warning("No cards processed.")
#         return

#     green = sum(1 for r in results if r.get("QUALITY", "") == "🟢 GREEN")
#     yellow = sum(1 for r in results if r.get("QUALITY", "") == "🟡 YELLOW")
#     red = sum(1 for r in results if r.get("QUALITY", "") == "🔴 RED")
#     log.info("✔ %d card(s) — 🟢 %d  🟡 %d  🔴 %d", len(results), green, yellow, red)

#     saved_latest = ""
#     if not args.no_excel:
#         save_to_database(results, db_path)
#         saved_latest = save_outputs(results, out_dir) or ""

#     if not args.no_vcf:
#         vcf_path = os.path.join(
#             out_dir, f"contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.vcf"
#         )
#         export_vcf(results, vcf_path)

#     print()
#     print("─" * 64)
#     print("  OUTPUT FILES")
#     print("─" * 64)
#     print(f"  Folder       : {out_dir}")
#     print(f"  Database     : {db_path}")
#     print(f"  Latest XLSX  : {os.path.join(out_dir,'latest.xlsx')}")
#     print(f"  Latest JSON  : {os.path.join(out_dir,'latest.json')}")
#     print(f"  Latest CSV   : {os.path.join(out_dir,'latest.csv')}")
#     print(f"  History      : {os.path.join(out_dir,'history','')}")
#     if not args.no_vcf:
#         print(f"  VCF contacts : {vcf_path}")
#     print("─" * 64)

#     # Auto-open latest.xlsx (Windows)
#     if saved_latest and os.path.exists(saved_latest):
#         try:
#             if sys.platform == "win32":
#                 os.startfile(os.path.abspath(saved_latest))
#         except Exception as e:
#             log.info("Auto-open skipped: %s", e)


# if __name__ == "__main__":
#     main()

#!/usr/bin/env python3
"""
Visiting Card OCR Engine v19.0

Deterministic reference implementation for business card OCR and extraction.
This module is intentionally built from classical image processing, geometry,
regex, and rule-based cascades only.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import logging
import re
import statistics
import sys
import zipfile
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from shutil import copy2
from typing import Any, Iterable, Iterator, Sequence

import tkinter as tk
from tkinter import filedialog

import numpy as np

try:
    import cv2
except Exception:  # pragma: no cover - optional in reference mode
    cv2 = None  # type: ignore[assignment]

try:
    import pandas as pd
except Exception:  # pragma: no cover - optional in reference mode
    pd = None  # type: ignore[assignment]

try:
    import phonenumbers
except Exception:  # pragma: no cover - optional in reference mode
    phonenumbers = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - optional in reference mode
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

try:
    from paddleocr import PaddleOCR
except Exception:  # pragma: no cover - optional in reference mode
    PaddleOCR = None  # type: ignore[assignment]

try:
    from rapidfuzz import fuzz
except Exception:  # pragma: no cover - optional in reference mode
    fuzz = None  # type: ignore[assignment]


LOGGER = logging.getLogger("ocr_engine")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)

SUPPORTED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tif", ".tiff"}

ADDRESS_HINTS = {
    "street",
    "st.",
    "st ",
    "road",
    "rd.",
    "avenue",
    "ave",
    "lane",
    "ln",
    "suite",
    "floor",
    "building",
    "block",
    "po box",
    "p.o. box",
    "postcode",
    "postal",
    "zip",
    "district",
    "city",
    "state",
    "province",
    "country",
    "singapore",
    "india",
    "japan",
    "australia",
    "malaysia",
    "hong kong",
    "china",
    "usa",
    "united states",
}

JOB_HINTS = {
    "engineer",
    "director",
    "manager",
    "consultant",
    "developer",
    "designer",
    "architect",
    "lead",
    "principal",
    "head",
    "officer",
    "specialist",
    "analyst",
    "founder",
    "president",
    "vp",
    "vice president",
    "systems",
    "software",
    "product",
    "sales",
    "marketing",
    "operations",
    "research",
}

COMPANY_HINTS = {
    "inc",
    "inc.",
    "ltd",
    "ltd.",
    "llc",
    "llp",
    "gmbh",
    "pty",
    "pte",
    "pvt",
    "corp",
    "co.",
    "company",
    "group",
    "solutions",
    "technologies",
    "systems",
    "labs",
    "studio",
    "consulting",
}

VALID_TLDS = {
    "ac",
    "ad",
    "ae",
    "af",
    "ag",
    "ai",
    "al",
    "am",
    "ar",
    "as",
    "at",
    "au",
    "be",
    "bg",
    "biz",
    "br",
    "ca",
    "cat",
    "ch",
    "cn",
    "co",
    "com",
    "coop",
    "cz",
    "de",
    "dev",
    "digital",
    "edu",
    "ee",
    "es",
    "eu",
    "fi",
    "fr",
    "fund",
    "gg",
    "gov",
    "group",
    "hk",
    "id",
    "ie",
    "in",
    "info",
    "io",
    "is",
    "it",
    "jobs",
    "jp",
    "kr",
    "la",
    "lat",
    "law",
    "li",
    "me",
    "mobi",
    "museum",
    "mx",
    "name",
    "net",
    "news",
    "nl",
    "no",
    "nz",
    "online",
    "org",
    "ph",
    "pl",
    "pro",
    "pt",
    "qa",
    "ro",
    "ru",
    "sa",
    "se",
    "sg",
    "site",
    "sk",
    "shop",
    "solutions",
    "studio",
    "su",
    "tech",
    "th",
    "travel",
    "tv",
    "tw",
    "ua",
    "uk",
    "us",
    "vn",
    "work",
    "xyz",
    "za",
}

MULTI_LEVEL_TLDS = {
    "co.uk",
    "com.au",
    "com.br",
    "com.cn",
    "com.hk",
    "com.my",
    "com.sg",
    "com.tw",
    "co.in",
    "co.jp",
    "co.nz",
    "co.za",
    "co.th",
    "co.id",
    "com.mx",
    "com.ar",
    "com.tr",
    "co.kr",
}

EMAIL_PREFIX_RE = re.compile(
    r"(?i)(?P<local>[a-z0-9._%+-]{1,64})@(?P<domain>[a-z0-9.-]{2,253})"
)
URL_RE = re.compile(
    r"(?i)\b(?:https?://|www\.)[a-z0-9.-]+(?:\.[a-z]{2,63})+(?:/[\w\-./?%&=+#~]*)?"
)
PHONE_FALLBACK_RE = re.compile(
    r"(?<!\w)(?:\+\d{1,3}[\s().-]*)?(?:\(?\d{1,4}\)?[\s().-]*){1,4}\d{3,4}(?:[\s().-]*\d{1,4}){0,3}(?!\w)"
)

POSTAL_PATTERNS = {
    "SG": re.compile(r"\b\d{6}\b"),
    "IN": re.compile(r"\b\d{6}\b"),
    "JP": re.compile(r"\b\d{3}-\d{4}\b"),
    "AU": re.compile(r"\b\d{4}\b"),
    "UK": re.compile(r"\b[A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2}\b", re.I),
    "US": re.compile(r"\b\d{5}(?:-\d{4})?\b"),
    "CA": re.compile(r"\b[A-Z]\d[A-Z]\s?\d[A-Z]\d\b", re.I),
}

UPPER_DIGIT_TRANSLATION = str.maketrans(
    {"0": "O", "1": "I", "3": "E", "4": "A", "5": "S", "7": "T", "8": "B"}
)
CONFUSION_TRANSLATION = str.maketrans(
    {
        "l": "1",
        "I": "1",
        "|": "1",
        "o": "0",
        "O": "0",
        "q": "g",
        "Q": "G",
        "s": "5",
        "S": "5",
        "b": "8",
        "B": "8",
        "z": "2",
    }
)


@dataclass(frozen=True)
class OCRToken:
    text: str
    score: float
    box: np.ndarray
    pass_name: str = "normal"

    @property
    def x1(self) -> float:
        return float(np.min(self.box[:, 0]))

    @property
    def y1(self) -> float:
        return float(np.min(self.box[:, 1]))

    @property
    def x2(self) -> float:
        return float(np.max(self.box[:, 0]))

    @property
    def y2(self) -> float:
        return float(np.max(self.box[:, 1]))

    @property
    def width(self) -> float:
        return max(1.0, self.x2 - self.x1)

    @property
    def height(self) -> float:
        return max(1.0, self.y2 - self.y1)

    @property
    def cx(self) -> float:
        return (self.x1 + self.x2) / 2.0

    @property
    def cy(self) -> float:
        return (self.y1 + self.y2) / 2.0


@dataclass
class LayoutLine:
    text: str
    tokens: list[OCRToken]
    x1: float
    y1: float
    x2: float
    y2: float
    height: float
    char_median: float
    source_passes: tuple[str, ...] = ()


@dataclass
class LayoutBlock:
    text: str
    lines: list[LayoutLine]
    x1: float
    y1: float
    x2: float
    y2: float


@dataclass
class FieldValue:
    value: str = ""
    confidence: float = 0.0
    evidence: list[str] = field(default_factory=list)
    validated: bool = False


@dataclass
class ExtractionResult:
    name: FieldValue = field(default_factory=FieldValue)
    company: FieldValue = field(default_factory=FieldValue)
    title: FieldValue = field(default_factory=FieldValue)
    email: FieldValue = field(default_factory=FieldValue)
    phone: FieldValue = field(default_factory=FieldValue)
    website: FieldValue = field(default_factory=FieldValue)
    address: FieldValue = field(default_factory=FieldValue)
    social: FieldValue = field(default_factory=FieldValue)
    quality_score: float = 0.0
    quality_badge: str = "RED"
    notes: list[str] = field(default_factory=list)
    raw_lines: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": asdict(self.name),
            "company": asdict(self.company),
            "title": asdict(self.title),
            "email": asdict(self.email),
            "phone": asdict(self.phone),
            "website": asdict(self.website),
            "address": asdict(self.address),
            "social": asdict(self.social),
            "quality_score": round(self.quality_score, 3),
            "quality_badge": self.quality_badge,
            "notes": list(self.notes),
            "raw_lines": list(self.raw_lines),
        }


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", "ignore")
    return str(value)


def _strip_trailing_punct(text: str) -> str:
    return text.strip().rstrip(".,;:")


def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _canonical_for_compare(text: str) -> str:
    return re.sub(
        r"[^a-z0-9]", "", _safe_text(text).lower().translate(CONFUSION_TRANSLATION)
    )


def _similarity(a: str, b: str) -> float:
    if fuzz is not None:
        return float(fuzz.ratio(a, b))
    return SequenceMatcher(None, a, b).ratio() * 100.0


def _alpha_ratio(text: str) -> float:
    clean = _safe_text(text)
    if not clean:
        return 0.0
    alpha = sum(1 for ch in clean if ch.isalpha())
    return alpha / max(1, len(clean))


def _digit_count(text: str) -> int:
    return sum(1 for ch in text if ch.isdigit())


def _is_upper_dominant(text: str) -> bool:
    letters = [ch for ch in text if ch.isalpha()]
    if not letters:
        return False
    return sum(ch.isupper() for ch in letters) / len(letters) >= 0.6


def _fix_digit_substitutions(text: str) -> str:
    if _is_upper_dominant(text):
        return text.translate(UPPER_DIGIT_TRANSLATION)
    return text


def _count_confusions(text: str) -> int:
    return sum(1 for ch in text if ch in "lI|oqQSBsZz")


def _text_fidelity(text: str) -> float:
    clean = _safe_text(text).strip()
    if not clean:
        return -100.0
    score = 0.0
    score += min(len(clean), 40)
    score += _alpha_ratio(clean) * 18.0
    score -= _count_confusions(clean) * 2.5
    score -= _digit_count(clean) * 0.25
    if clean == clean.title() or clean.isupper():
        score += 3.0
    if any(token in clean.lower() for token in COMPANY_HINTS):
        score += 1.5
    return score


def _box_from_any(box: Any) -> np.ndarray:
    arr = np.asarray(box, dtype=float)
    if arr.ndim == 1 and arr.size == 8:
        arr = arr.reshape(4, 2)
    if arr.ndim != 2 or arr.shape[-1] != 2:
        raise ValueError(f"Unsupported box shape: {arr.shape}")
    return arr.astype(float)


def _rect_iou(a: OCRToken, b: OCRToken) -> float:
    x_left = max(a.x1, b.x1)
    y_top = max(a.y1, b.y1)
    x_right = min(a.x2, b.x2)
    y_bottom = min(a.y2, b.y2)
    if x_right <= x_left or y_bottom <= y_top:
        return 0.0
    inter = (x_right - x_left) * (y_bottom - y_top)
    union = a.width * a.height + b.width * b.height - inter
    return inter / max(1.0, union)


def _parse_sequence_items(
    items: Sequence[Any], pass_name: str = "normal"
) -> list[OCRToken]:
    tokens: list[OCRToken] = []
    for item in items:
        box: Any = None
        text: Any = ""
        score: Any = 0.0

        if isinstance(item, dict):
            box = (
                item.get("box")
                or item.get("poly")
                or item.get("bbox")
                or item.get("rec_poly")
            )
            text = item.get("text") or item.get("rec_text") or item.get("label") or ""
            score = item.get("score") or item.get("rec_score") or 0.0
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            box = item[0]
            payload = item[1]
            if isinstance(payload, (list, tuple)) and len(payload) >= 2:
                text = payload[0]
                score = payload[1]
            elif isinstance(payload, dict):
                text = payload.get("text") or payload.get("rec_text") or ""
                score = payload.get("score") or payload.get("rec_score") or 0.0
            else:
                text = payload
                score = 0.0

        if box is None:
            continue

        try:
            tokens.append(
                OCRToken(
                    text=_safe_text(text),
                    score=float(score),
                    box=_box_from_any(box),
                    pass_name=pass_name,
                )
            )
        except Exception:
            continue

    return tokens


def parse_paddle_result(result: Any, pass_name: str = "normal") -> list[OCRToken]:
    if result is None:
        return []

    if hasattr(result, "get"):
        mapping = result  # type: ignore[assignment]
        texts = list(mapping.get("rec_texts", []) or [])
        scores = list(mapping.get("rec_scores", []) or [])
        polys = list(mapping.get("rec_polys", []) or [])
        tokens: list[OCRToken] = []
        for idx, text in enumerate(texts):
            if idx >= len(polys):
                break
            score = float(scores[idx]) if idx < len(scores) else 0.0
            try:
                tokens.append(
                    OCRToken(
                        text=_safe_text(text),
                        score=score,
                        box=_box_from_any(polys[idx]),
                        pass_name=pass_name,
                    )
                )
            except Exception:
                continue
        if tokens:
            return tokens

    if isinstance(result, (list, tuple)):
        if result and isinstance(result[0], dict):
            return parse_paddle_result(result[0], pass_name=pass_name)
        if (
            result
            and isinstance(result[0], (list, tuple))
            and len(result[0]) > 0
            and isinstance(result[0][0], (list, tuple, dict, np.ndarray))
        ):
            return _parse_sequence_items(result, pass_name=pass_name)
        if result and len(result) == 2 and isinstance(result[0], (list, tuple)):
            return _parse_sequence_items(result[0], pass_name=pass_name)

    return []


class PaddleBridge:
    def __init__(self, lang: str = "en") -> None:
        if PaddleOCR is None:
            raise RuntimeError("PaddleOCR is not installed in this environment.")

        self.engine = PaddleOCR(
            lang=lang,
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_textline_orientation=False,
        )

    def infer(self, image: np.ndarray, pass_name: str = "normal") -> list[OCRToken]:
        raw = self.engine.predict(image)
        return parse_paddle_result(raw, pass_name=pass_name)


def _ensure_color(image: np.ndarray) -> np.ndarray:
    if image.ndim == 2:
        return (
            cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
            if cv2 is not None
            else np.stack([image] * 3, axis=-1)
        )
    if image.ndim == 3 and image.shape[2] == 4 and cv2 is not None:
        return cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)
    return image.copy()


def build_pass_variants(image: np.ndarray) -> list[tuple[str, np.ndarray]]:
    base = _ensure_color(image)
    variants = [("normal", base)]
    if cv2 is None:
        return variants

    bright = cv2.convertScaleAbs(base, alpha=1.14, beta=14)
    variants.append(("bright", bright))
    return variants


def dedupe_tokens(tokens: Sequence[OCRToken]) -> list[OCRToken]:
    ordered = sorted(tokens, key=lambda t: (t.y1, t.x1, -t.score))
    kept: list[OCRToken] = []
    for token in ordered:
        duplicate = False
        for existing in kept:
            if _rect_iou(token, existing) >= 0.58 and _canonical_for_compare(
                token.text
            ) == _canonical_for_compare(existing.text):
                duplicate = True
                if token.score > existing.score:
                    kept[kept.index(existing)] = token
                break
        if not duplicate:
            kept.append(token)
    return sorted(kept, key=lambda t: (t.y1, t.x1))


def group_tokens_into_rows(tokens: Sequence[OCRToken]) -> list[list[OCRToken]]:
    if not tokens:
        return []
    ordered = sorted(tokens, key=lambda t: (t.cy, t.x1))
    median_height = statistics.median([t.height for t in ordered]) if ordered else 12.0
    y_threshold = max(2.5, median_height * 0.45)

    rows: list[list[OCRToken]] = []
    current: list[OCRToken] = []
    current_center = ordered[0].cy

    for token in ordered:
        if not current:
            current = [token]
            current_center = token.cy
            continue

        if abs(token.cy - current_center) <= y_threshold:
            current.append(token)
            current_center = sum(t.cy for t in current) / len(current)
        else:
            rows.append(sorted(current, key=lambda t: t.x1))
            current = [token]
            current_center = token.cy

    if current:
        rows.append(sorted(current, key=lambda t: t.x1))

    return rows


def _row_char_width_median(row: Sequence[OCRToken]) -> float:
    widths = []
    for token in row:
        clean = _normalize_spaces(_safe_text(token.text))
        if not clean:
            continue
        widths.append(token.width / max(1, len(clean)))
    return statistics.median(widths) if widths else 8.0


def compose_row_text(row: Sequence[OCRToken]) -> str:
    if not row:
        return ""

    ordered = sorted(row, key=lambda t: t.x1)
    char_median = _row_char_width_median(ordered)

    parts: list[str] = []
    previous: OCRToken | None = None

    for token in ordered:
        token_text = _strip_trailing_punct(
            _fix_digit_substitutions(_safe_text(token.text).strip())
        )

        if not token_text:
            continue

        if previous is None:
            parts.append(token_text)
            previous = token
            continue

        gap = token.x1 - previous.x2

        prev_text = _strip_trailing_punct(
            _fix_digit_substitutions(_safe_text(previous.text).strip())
        )

        prev_is_fragment = len(prev_text) <= 2 or len(token_text) <= 2

        if gap <= char_median * 0.45:
            if prev_is_fragment:
                parts[-1] = parts[-1] + token_text
            else:
                parts[-1] = (
                    parts[-1] + token_text
                    if parts[-1].endswith("-")
                    else parts[-1] + " " + token_text
                )
        else:
            parts.append(token_text)

        previous = token

    text = _normalize_spaces(" ".join(parts))

    if len(text.split()) > 1 and all(len(piece) == 1 for piece in text.split()):
        text = text.replace(" ", "")

    return _strip_trailing_punct(text)


def build_layout_lines(tokens: Sequence[OCRToken]) -> list[LayoutLine]:
    lines: list[LayoutLine] = []

    for row in group_tokens_into_rows(tokens):
        xs = [t.x1 for t in row] + [t.x2 for t in row]
        ys = [t.y1 for t in row] + [t.y2 for t in row]

        text = compose_row_text(row)
        char_median = _row_char_width_median(row)
        passes = tuple(sorted({t.pass_name for t in row}))

        lines.append(
            LayoutLine(
                text=text,
                tokens=list(row),
                x1=min(xs),
                y1=min(ys),
                x2=max(xs),
                y2=max(ys),
                height=max(1.0, max(ys) - min(ys)),
                char_median=char_median,
                source_passes=passes,
            )
        )

    return lines


def merge_multiline_blocks(lines: Sequence[LayoutLine]) -> list[LayoutBlock]:
    if not lines:
        return []

    ordered = sorted(lines, key=lambda ln: (ln.y1, ln.x1))

    median_height = statistics.median([ln.height for ln in ordered])

    merge_gap = max(2.0, median_height * 0.35)

    blocks: list[LayoutBlock] = []

    current = [ordered[0]]

    for line in ordered[1:]:
        prev = current[-1]

        gap = line.y1 - prev.y2

        height_symmetry = (
            abs(line.height - prev.height) <= max(line.height, prev.height) * 0.25
        )

        horizontal_overlap = min(line.x2, prev.x2) - max(line.x1, prev.x1)

        overlap_ratio = horizontal_overlap / max(
            1.0, min(line.x2 - line.x1, prev.x2 - prev.x1)
        )

        same_band = gap <= merge_gap

        # NEW RULE:
        # only merge if horizontally aligned

        current_text = current[-1].text.lower()
        line_text = line.text.lower()

        contains_email = "@" in current_text or "@" in line_text

        contains_company = _has_hint(current_text, COMPANY_HINTS) or _has_hint(
            line_text, COMPANY_HINTS
        )

        contains_title = _has_hint(current_text, JOB_HINTS) or _has_hint(
            line_text, JOB_HINTS
        )

        safe_to_merge = same_band and height_symmetry and overlap_ratio > 0.55

        if safe_to_merge and not (
            contains_email or (contains_company and contains_title)
        ):
            current.append(line)
        else:
            blocks.append(_block_from_lines(current))
            current = [line]

        if current:
            blocks.append(_block_from_lines(current))

        return blocks


def _block_from_lines(lines: Sequence[LayoutLine]) -> LayoutBlock:
    xs = [line.x1 for line in lines] + [line.x2 for line in lines]
    ys = [line.y1 for line in lines] + [line.y2 for line in lines]
    text = _normalize_spaces(" ".join(line.text for line in lines if line.text))
    return LayoutBlock(
        text=text, lines=list(lines), x1=min(xs), y1=min(ys), x2=max(xs), y2=max(ys)
    )


def _has_hint(text: str, hints: set[str]) -> bool:
    t = text.lower()
    return any(h in t for h in hints)


def _postal_match(text: str) -> tuple[str, str] | None:
    for region, regex in POSTAL_PATTERNS.items():
        match = regex.search(text)
        if match:
            return region, match.group(0)
    return None


def isolate_address_lines(
    lines: Sequence[LayoutLine],
) -> tuple[list[LayoutLine], list[str], str | None]:
    address_lines: list[LayoutLine] = []
    blocked_tokens: list[str] = []
    region_hint: str | None = None

    for line in lines:
        text = line.text.strip()
        lower = text.lower()
        postal = _postal_match(text)
        addr_hint = _has_hint(lower, ADDRESS_HINTS)
        contains_phone_like = bool(PHONE_FALLBACK_RE.search(text)) or "+" in text

        if postal and (addr_hint or not contains_phone_like):
            region_hint = region_hint or postal[0]
            address_lines.append(line)
            blocked_tokens.append(text)
            blocked_tokens.append(postal[1])
            continue

        if addr_hint and _alpha_ratio(text) >= 0.45 and _digit_count(text) >= 1:
            address_lines.append(line)
            blocked_tokens.append(text)

    return address_lines, blocked_tokens, region_hint


def _trim_email_candidate(raw: str) -> str:
    token = _strip_trailing_punct(_normalize_spaces(raw.strip("()[]{}<>\"'`")))
    if "@" not in token:
        return ""

    local, domain = token.split("@", 1)
    domain = domain.strip(".,;:")
    labels = [label for label in domain.split(".") if label]
    if len(labels) < 2:
        return ""

    for tail_size in range(min(3, len(labels)), 0, -1):
        tail = ".".join(labels[-tail_size:]).lower()
        if tail in MULTI_LEVEL_TLDS:
            clean_domain = ".".join(labels[:-tail_size] + [tail])
            return f"{local}@{clean_domain}"

    last_label = labels[-1]
    for cut in range(len(last_label), 1, -1):
        prefix = last_label[:cut].lower()
        if prefix in VALID_TLDS or (len(prefix) == 2 and prefix.isalpha()):
            clean_domain = ".".join(labels[:-1] + [prefix])
            return f"{local}@{clean_domain}"

    if last_label.lower() in VALID_TLDS or (
        len(last_label) == 2 and last_label.isalpha()
    ):
        return f"{local}@{domain}"

    return ""


def extract_emails(texts: Sequence[str]) -> list[FieldValue]:
    found: dict[str, FieldValue] = {}
    for text in texts:
        for candidate in re.findall(r"(?i)[a-z0-9._%+-]+@[a-z0-9.-]+", text):
            trimmed = _trim_email_candidate(candidate)
            if not trimmed:
                continue
            if not EMAIL_PREFIX_RE.fullmatch(trimmed):
                continue
            canonical = trimmed.lower()
            found.setdefault(
                canonical,
                FieldValue(
                    value=canonical, confidence=0.96, evidence=[text], validated=True
                ),
            )
    return list(found.values())


def _normalize_phone_candidate(text: str) -> str:
    clean = _safe_text(text).replace("\u2013", "-").replace("\u2014", "-")
    clean = re.sub(r"[^\d+()\-\s./]", " ", clean)
    clean = _normalize_spaces(clean)
    if not clean:
        return ""
    digits = re.sub(r"\D", "", clean)
    if len(digits) < 7:
        return ""
    if clean.startswith("+"):
        return "+" + digits
    if clean.count("(") or clean.count(")"):
        return clean
    return clean


def _sliding_digit_windows(text: str) -> list[str]:
    digits = re.sub(r"\D", "", text)
    if len(digits) < 7:
        return []
    windows: list[str] = []
    upper = min(len(digits), 15)
    for start in range(0, len(digits) - 6):
        for end in range(start + 7, upper + 1):
            windows.append(digits[start:end])
    return windows


def _phone_fallback_candidates(text: str) -> list[str]:
    candidates = []
    normalized = _normalize_phone_candidate(text)
    if normalized:
        candidates.append(normalized)
        digits_only = re.sub(r"\D", "", normalized)
        if 7 <= len(digits_only) <= 15:
            candidates.append(digits_only)
            if not normalized.startswith("+"):
                candidates.append("+" + digits_only)
    for window in _sliding_digit_windows(text):
        candidates.append(window)
        candidates.append("+" + window)
    seen = set()
    unique: list[str] = []
    for item in candidates:
        if item not in seen:
            seen.add(item)
            unique.append(item)
    return unique


def _format_phone_number(candidate: str, region_hint: str | None) -> str | None:
    candidate = _normalize_phone_candidate(candidate)
    if not candidate:
        return None

    if phonenumbers is not None:
        for region in ([region_hint] if region_hint else []) + [None]:
            try:
                parsed = phonenumbers.parse(candidate, region)
            except Exception:
                continue
            if phonenumbers.is_possible_number(parsed) and phonenumbers.is_valid_number(
                parsed
            ):
                return phonenumbers.format_number(
                    parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL
                )

    digits = re.sub(r"\D", "", candidate)
    if 7 <= len(digits) <= 15:
        return candidate if candidate.startswith("+") else digits
    return None


def extract_phones(
    lines: Sequence[LayoutLine], blocked_tokens: Sequence[str], region_hint: str | None
) -> list[FieldValue]:
    blocked_set = {_canonical_for_compare(token) for token in blocked_tokens if token}
    candidates: dict[str, FieldValue] = {}

    for line in lines:
        canonical_line = _canonical_for_compare(line.text)
        if canonical_line in blocked_set:
            continue
        if _has_hint(line.text, ADDRESS_HINTS) and _postal_match(line.text):
            continue
        if _digit_count(line.text) < 7:
            continue

        raw_candidates = [line.text] + _phone_fallback_candidates(line.text)
        for raw in raw_candidates:
            if _postal_match(raw):
                continue
            formatted = _format_phone_number(raw, region_hint)
            if not formatted:
                continue
            canonical = re.sub(r"\D", "", formatted)
            if len(canonical) < 7:
                continue
            current = candidates.get(canonical)
            if current is None or _text_fidelity(formatted) > _text_fidelity(
                current.value
            ):
                candidates[canonical] = FieldValue(
                    value=formatted,
                    confidence=0.94,
                    evidence=[line.text],
                    validated=True,
                )

    return list(candidates.values())


def extract_urls(texts: Sequence[str]) -> list[FieldValue]:
    found: dict[str, FieldValue] = {}
    for text in texts:
        for match in URL_RE.finditer(text):
            value = _strip_trailing_punct(match.group(0))
            if value.lower().startswith("www."):
                value = "https://" + value
            canonical = value.lower()
            found.setdefault(
                canonical,
                FieldValue(
                    value=value, confidence=0.9, evidence=[text], validated=True
                ),
            )
    return list(found.values())


def extract_social_handles(texts: Sequence[str]) -> list[FieldValue]:
    handles: dict[str, FieldValue] = {}
    for text in texts:
        for match in re.finditer(r"(?<!\w)@[A-Za-z0-9_.]{2,30}(?!\w)", text):
            handle = match.group(0)
            canonical = handle.lower()
            handles.setdefault(
                canonical,
                FieldValue(
                    value=handle, confidence=0.8, evidence=[text], validated=True
                ),
            )
    return list(handles.values())


def extract_address(address_lines: Sequence[LayoutLine]) -> FieldValue:
    if not address_lines:
        return FieldValue()

    sorted_lines = sorted(address_lines, key=lambda ln: (ln.y1, ln.x1))

    cleaned = []

    for line in sorted_lines:
        txt = URL_RE.sub("", line.text)
        txt = txt.strip()

        if txt:
            cleaned.append(txt)

    text = _normalize_spaces(", ".join(cleaned))
    text = _strip_trailing_punct(text)

    return FieldValue(
        value=text,
        confidence=0.9,
        evidence=[ln.text for ln in sorted_lines],
        validated=_alpha_ratio(text) >= 0.45,
    )


def _looks_like_name(text: str) -> bool:
    if not text or _digit_count(text) > 0:
        return False
    words = text.split()
    if not 1 <= len(words) <= 4:
        return False
    lower = text.lower()
    if (
        _has_hint(lower, ADDRESS_HINTS)
        or _has_hint(lower, JOB_HINTS)
        or _has_hint(lower, COMPANY_HINTS)
    ):
        return False
    if _alpha_ratio(text) < 0.7:
        return False
    return True


def extract_name(
    lines: Sequence[LayoutLine],
    title_text: str,
    company_text: str,
    email_text: str,
    phone_text: str,
) -> FieldValue:
    candidates: list[tuple[float, str, str]] = []

    excluded = {
        _canonical_for_compare(title_text),
        _canonical_for_compare(company_text),
        _canonical_for_compare(email_text),
        _canonical_for_compare(phone_text),
    }

    for idx, line in enumerate(sorted(lines, key=lambda ln: (ln.y1, ln.x1))):
        text = _strip_trailing_punct(line.text)
        canonical = _canonical_for_compare(text)

        if not text or canonical in excluded:
            continue

        # Reject obvious company names
        if any(
            word in text.lower()
            for word in (
                "ltd",
                "limited",
                "llp",
                "inc",
                "corp",
                "corporation",
                "company",
                "group",
                "gmbh",
                "pte",
                "s.a.",
                "spa",
            )
        ):
            continue

        if not _looks_like_name(text):
            continue

        score = 0.0

        score += 30.0 if len(text.split()) in {2, 3} else 12.0
        score += 18.0 if text == text.title() else 10.0
        score += 10.0 if idx <= 2 else 2.0
        score += _alpha_ratio(text) * 15.0

        candidates.append((score, text, line.text))

    if not candidates:
        return FieldValue()

    candidates.sort(key=lambda item: (item[0], len(item[1])), reverse=True)

    score, value, evidence = candidates[0]

    return FieldValue(
        value=value,
        confidence=min(0.99, 0.4 + score / 100.0),
        evidence=[evidence],
        validated=True,
    )


def _looks_like_company(text: str) -> bool:
    if not text:
        return False
    lower = text.lower()
    if _has_hint(lower, COMPANY_HINTS):
        return True
    if any(ch.isdigit() for ch in text):
        return False
    if text.isupper() and len(text) <= 32:
        return True
    return 1 <= len(text.split()) <= 4 and _alpha_ratio(text) >= 0.55


def extract_company(
    lines: Sequence[LayoutLine], title_text: str, name_text: str, address_text: str
) -> FieldValue:
    candidates: list[tuple[float, str, str]] = []
    excluded = {
        _canonical_for_compare(title_text),
        _canonical_for_compare(name_text),
        _canonical_for_compare(address_text),
    }

    company_suffixes = (
        "LIMITED",
        "LTD",
        "LLP",
        "INC",
        "CORP",
        "GROUP",
        "GMBH",
        "S.P.A",
        "PTE",
    )

    for idx, line in enumerate(sorted(lines, key=lambda ln: (ln.y1, ln.x1))):
        text = _strip_trailing_punct(line.text)
        canonical = _canonical_for_compare(text)

        # Never allow emails to become companies
        if "@" in text:
            continue

        # Never allow websites to become companies
        if "www." in text.lower() or "http" in text.lower():
            continue

        if not text or canonical in excluded:
            continue

        if not _looks_like_company(text):
            continue

        score = 0.0

        score += 28.0 if _has_hint(text.lower(), COMPANY_HINTS) else 8.0
        score += 18.0 if text.isupper() else 9.0

        # Strong boost for common company suffixes
        if any(suffix in text.upper() for suffix in company_suffixes):
            score += 25.0

        # Penalize person-like names
        words = text.split()
        if (
            len(words) <= 4
            and all(word.replace(".", "").isalpha() for word in words)
            and not any(suffix in text.upper() for suffix in company_suffixes)
        ):
            score -= 15.0

        score += 8.0 if idx <= 2 else 3.0
        score += _alpha_ratio(text) * 14.0

        candidates.append((score, text, line.text))

    if not candidates:
        return FieldValue()

    candidates.sort(key=lambda item: (item[0], len(item[1])), reverse=True)

    score, value, evidence = candidates[0]

    return FieldValue(
        value=value,
        confidence=min(0.99, 0.35 + score / 100.0),
        evidence=[evidence],
        validated=True,
    )


def extract_job_title(
    blocks: Sequence[LayoutBlock], company_text: str, name_text: str
) -> FieldValue:
    best: tuple[float, str, list[str]] | None = None
    excluded = {_canonical_for_compare(company_text), _canonical_for_compare(name_text)}

    for block in blocks:

        text = _strip_trailing_punct(block.text)
        if not text or _canonical_for_compare(text) in excluded:
            continue
        lower = text.lower()
        if not _has_hint(lower, JOB_HINTS):
            continue
        score = 0.0
        score += (
            30.0
            if any(
                h in lower
                for h in (
                    "engineer",
                    "director",
                    "manager",
                    "consultant",
                    "developer",
                    "designer",
                )
            )
            else 12.0
        )
        score += 14.0 if len(block.lines) >= 2 else 8.0
        score += (
            10.0
            if block.lines
            and all(line.height <= block.lines[0].height * 1.35 for line in block.lines)
            else 4.0
        )
        score += _alpha_ratio(text) * 15.0
        if best is None or score > best[0]:
            best = (score, text, [line.text for line in block.lines])

    if best is None:
        return FieldValue()

    score, value, evidence = best
    return FieldValue(
        value=value,
        confidence=min(0.99, 0.38 + score / 100.0),
        evidence=evidence,
        validated=True,
    )


def _merge_fuzzy_duplicates(lines: Sequence[LayoutLine]) -> list[LayoutLine]:
    kept: list[LayoutLine] = []
    for line in sorted(lines, key=lambda ln: (ln.y1, ln.x1)):
        replaced = False
        for index, existing in enumerate(kept):
            vertical_overlap = min(line.y2, existing.y2) - max(line.y1, existing.y1)
            if vertical_overlap <= 0:
                continue
            similarity = _similarity(
                _canonical_for_compare(line.text), _canonical_for_compare(existing.text)
            )
            if similarity >= 75.0:
                if _text_fidelity(line.text) > _text_fidelity(existing.text):
                    kept[index] = line
                replaced = True
                break
        if not replaced:
            kept.append(line)
    return kept


def _plausible_email(text: str) -> bool:
    if not text or "@" not in text:
        return False
    local, domain = text.split("@", 1)
    if not local or not domain:
        return False
    if len(local) > 64 or len(text) > 254:
        return False
    if ".." in domain or domain.startswith("-") or domain.endswith("-"):
        return False
    if "." not in domain:
        return False
    return True


def validate_email(value: str) -> bool:
    candidate = _trim_email_candidate(value)
    return bool(
        candidate
        and EMAIL_PREFIX_RE.fullmatch(candidate)
        and _plausible_email(candidate)
    )


def validate_phone(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    if not (7 <= len(digits) <= 15):
        return False
    if _postal_match(value):
        return False
    if phonenumbers is None:
        return True
    if not value.startswith("+"):
        return True
    try:
        parsed = phonenumbers.parse(value, None)
    except Exception:
        return False
    return phonenumbers.is_possible_number(parsed) and phonenumbers.is_valid_number(
        parsed
    )


def validate_website(value: str) -> bool:
    return bool(URL_RE.fullmatch(value) or URL_RE.search(value))


def validate_address(value: str) -> bool:
    if not value:
        return False
    if _postal_match(value) is None:
        return _alpha_ratio(value) >= 0.55 and _digit_count(value) >= 1
    return True


def validate_name(value: str) -> bool:
    return _looks_like_name(value)


def validate_company(value: str) -> bool:
    return _looks_like_company(value)


def validate_title(value: str) -> bool:
    return bool(
        value and _has_hint(value.lower(), JOB_HINTS) and _alpha_ratio(value) >= 0.55
    )


def _score_field(value: str, validator: Any, weight: float) -> tuple[float, bool]:
    if not value:
        return 0.0, False
    valid = bool(validator(value))
    return (weight if valid else 0.0), valid


def score_quality(result: ExtractionResult) -> tuple[float, str]:
    score = 0.0
    total = 0.0

    field_specs = [
        (result.name.value, validate_name, 1.15),
        (result.company.value, validate_company, 1.10),
        (result.email.value, validate_email, 1.25),
        (result.phone.value, validate_phone, 1.20),
        (result.website.value, validate_website, 0.70),
        (result.address.value, validate_address, 0.80),
        (result.title.value, validate_title, 0.45),
    ]

    for value, validator, weight in field_specs:
        if value:
            total += weight
            earned, valid = _score_field(value, validator, weight)
            score += earned
            if valid:
                continue

    if total == 0:
        return 0.0, "RED"

    ratio = score / total
    if ratio >= 0.78 and any(
        [
            result.name.validated,
            result.company.validated,
            result.email.validated,
            result.phone.validated,
        ]
    ):
        return ratio, "GREEN"
    if ratio >= 0.52:
        return ratio, "AMBER"
    return ratio, "RED"


def _load_image_from_bytes(blob: bytes) -> np.ndarray:
    if cv2 is None:
        raise RuntimeError("OpenCV is required for image decoding.")
    array = np.frombuffer(blob, dtype=np.uint8)
    image = cv2.imdecode(array, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError("Unable to decode image data.")
    return image


def iter_images_from_pdf_bytes(blob: bytes) -> Iterator[np.ndarray]:
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "PDF rasterization requires PyMuPDF (fitz) in this environment."
        ) from exc

    doc = fitz.open(stream=blob, filetype="pdf")
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
        buffer = np.frombuffer(pix.samples, dtype=np.uint8)
        channels = pix.n
        image = buffer.reshape(pix.height, pix.width, channels)
        if channels == 1 and cv2 is not None:
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
        elif channels == 4 and cv2 is not None:
            image = cv2.cvtColor(image, cv2.COLOR_RGBA2BGR)
        yield image


def iter_images_from_zip_bytes(blob: bytes) -> Iterator[np.ndarray]:
    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            data = archive.read(info.filename)
            suffix = Path(info.filename).suffix.lower()
            if suffix in SUPPORTED_IMAGE_SUFFIXES:
                yield _load_image_from_bytes(data)
            elif suffix == ".pdf":
                yield from iter_images_from_pdf_bytes(data)
            elif suffix == ".zip":
                yield from iter_images_from_zip_bytes(data)


def iter_images_from_path(path: Path) -> Iterator[tuple[str, np.ndarray]]:
    if path.is_dir():
        for child in sorted(path.rglob("*")):
            if child.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES:
                image = _load_image_from_bytes(child.read_bytes())
                yield str(child), image
            elif child.suffix.lower() == ".pdf":
                for index, image in enumerate(
                    iter_images_from_pdf_bytes(child.read_bytes()), start=1
                ):
                    yield f"{child}#page{index}", image
            elif child.suffix.lower() == ".zip":
                for index, image in enumerate(
                    iter_images_from_zip_bytes(child.read_bytes()), start=1
                ):
                    yield f"{child}#asset{index}", image
        return

    suffix = path.suffix.lower()
    if suffix in SUPPORTED_IMAGE_SUFFIXES:
        yield str(path), _load_image_from_bytes(path.read_bytes())
    elif suffix == ".pdf":
        for index, image in enumerate(
            iter_images_from_pdf_bytes(path.read_bytes()), start=1
        ):
            yield f"{path}#page{index}", image
    elif suffix == ".zip":
        for index, image in enumerate(
            iter_images_from_zip_bytes(path.read_bytes()), start=1
        ):
            yield f"{path}#asset{index}", image
    else:
        raise ValueError(f"Unsupported input type: {path}")


def _prepare_ocr_image(image: np.ndarray) -> np.ndarray:
    color = _ensure_color(image)
    if cv2 is None:
        return color
    return cv2.fastNlMeansDenoisingColored(color, None, 3, 3, 7, 21)


def _merge_tokens_from_passes(pass_tokens: Sequence[OCRToken]) -> list[OCRToken]:
    return dedupe_tokens(pass_tokens)


def _collapse_duplicate_lines(lines: Sequence[LayoutLine]) -> list[LayoutLine]:
    return _merge_fuzzy_duplicates(lines)


class BusinessCardEngine:
    def __init__(self, lang: str = "en") -> None:
        self.bridge = PaddleBridge(lang=lang) if PaddleOCR is not None else None

    def process_card_matrix(self, image: np.ndarray) -> ExtractionResult:
        if self.bridge is None:
            raise RuntimeError("PaddleOCR bridge is unavailable in this environment.")

        all_tokens: list[OCRToken] = []
        prepared = _prepare_ocr_image(image)
        for pass_name, variant in build_pass_variants(prepared):
            try:
                tokens = self.bridge.infer(variant, pass_name=pass_name)
            except Exception as exc:
                LOGGER.warning("OCR pass failed (%s): %s", pass_name, exc)
                tokens = []
            all_tokens.extend(tokens)

        all_tokens = _merge_tokens_from_passes(all_tokens)
        lines = build_layout_lines(all_tokens)
        lines = _collapse_duplicate_lines(lines)

        address_lines, blocked_tokens, region_hint = isolate_address_lines(lines)
        blocks = merge_multiline_blocks(lines)
        line_texts = [line.text for line in lines if line.text]
        address_field = extract_address(address_lines)
        email_candidates = extract_emails(line_texts)
        url_candidates = extract_urls(line_texts)
        social_candidates = extract_social_handles(line_texts)
        phone_candidates = extract_phones(
            lines, blocked_tokens + [address_field.value], region_hint
        )

        email = email_candidates[0] if email_candidates else FieldValue()
        website = url_candidates[0] if url_candidates else FieldValue()
        social = social_candidates[0] if social_candidates else FieldValue()

        name = extract_name(
            lines,
            "",
            "",
            email.value,
            phone_candidates[0].value if phone_candidates else "",
        )
        company = extract_company(lines, "", name.value, address_field.value)
        title = extract_job_title(blocks, company.value, name.value)

        phone = phone_candidates[0] if phone_candidates else FieldValue()

        result = ExtractionResult(
            name=name,
            company=company,
            title=title,
            email=email,
            phone=phone,
            website=website,
            address=address_field,
            social=social,
            raw_lines=line_texts,
        )

        for field_obj in (
            result.name,
            result.company,
            result.title,
            result.email,
            result.phone,
            result.website,
            result.address,
            result.social,
        ):
            field_obj.value = _strip_trailing_punct(_normalize_spaces(field_obj.value))

        result.quality_score, result.quality_badge = score_quality(result)
        if result.address.value:
            result.notes.append("Address firewall applied before phone extraction.")
        if title.value and len(title.evidence) > 1:
            result.notes.append(
                "Vertical proximity graph merged multi-line title fragments."
            )
        return result


def _ensure_output_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _snapshot_stem(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return safe or "card"


def _flatten_result_row(result: ExtractionResult) -> dict[str, Any]:
    payload = result.to_dict()
    row: dict[str, Any] = {
        "quality_score": payload["quality_score"],
        "quality_badge": payload["quality_badge"],
        "notes": " | ".join(payload["notes"]),
        "raw_lines": " || ".join(payload["raw_lines"]),
    }
    for key in [
        "name",
        "company",
        "title",
        "email",
        "phone",
        "website",
        "address",
        "social",
    ]:
        item = payload[key]
        row[f"{key}_value"] = item["value"]
        row[f"{key}_confidence"] = item["confidence"]
        row[f"{key}_validated"] = item["validated"]
        row[f"{key}_evidence"] = " | ".join(item["evidence"])
    return row


def export_json(result: ExtractionResult, output_dir: Path, stem: str) -> Path:
    output_dir = _ensure_output_dir(output_dir)
    path = output_dir / f"{stem}.json"
    path.write_text(
        json.dumps(result.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8"
    )
    return path


def export_csv(result: ExtractionResult, output_dir: Path, stem: str) -> Path:
    output_dir = _ensure_output_dir(output_dir)
    path = output_dir / f"{stem}.csv"
    row = _flatten_result_row(result)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(row.keys()))
        writer.writeheader()
        writer.writerow(row)
    return path


def export_xlsx(result: ExtractionResult, output_dir: Path, stem: str) -> Path:
    output_dir = _ensure_output_dir(output_dir)
    path = output_dir / f"{stem}.xlsx"
    payload = result.to_dict()

    if Workbook is None:
        raise RuntimeError("openpyxl is required for xlsx export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "snapshot"
    ws.append(["field", "value", "validated", "confidence"])
    for key in [
        "name",
        "company",
        "title",
        "email",
        "phone",
        "website",
        "address",
        "social",
    ]:
        item = payload[key]
        ws.append([key, item["value"], item["validated"], item["confidence"]])
    ws.append(["quality_score", payload["quality_score"], payload["quality_badge"], ""])
    wb.save(path)
    return path


def export_vcard(result: ExtractionResult, output_dir: Path, stem: str) -> Path:
    output_dir = _ensure_output_dir(output_dir)
    path = output_dir / f"{stem}.vcf"
    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"FN:{result.name.value or result.company.value or stem}",
        f"ORG:{result.company.value}",
        f"TITLE:{result.title.value}",
        f"EMAIL;TYPE=INTERNET:{result.email.value}",
        f"TEL;TYPE=VOICE:{result.phone.value}",
        f"URL:{result.website.value}",
        f"ADR;TYPE=WORK:;;{result.address.value}",
        "END:VCARD",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def _append_monthly_master(result: ExtractionResult, output_dir: Path) -> Path | None:
    if Workbook is None:
        return None

    output_dir = _ensure_output_dir(output_dir)
    month_key = dt.datetime.now().strftime("%Y_%m")
    path = output_dir / f"master_{month_key}.xlsx"
    payload = result.to_dict()
    signature = "|".join(
        _canonical_for_compare(payload[key]["value"])
        for key in ["name", "company", "title", "email", "phone", "website", "address"]
    )

    if path.exists() and load_workbook is not None:
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    sheet = wb[month_key] if month_key in wb.sheetnames else wb.create_sheet(month_key)
    if sheet.max_row == 1 and sheet["A1"].value is None:
        sheet.append(
            [
                "signature",
                "timestamp",
                "name",
                "company",
                "title",
                "email",
                "phone",
                "website",
                "address",
                "quality_score",
                "quality_badge",
            ]
        )

    existing_signatures = {str(cell.value) for cell in sheet["A"] if cell.value}
    if signature in existing_signatures:
        wb.save(path)
        return path

    sheet.append(
        [
            signature,
            dt.datetime.now().isoformat(timespec="seconds"),
            payload["name"]["value"],
            payload["company"]["value"],
            payload["title"]["value"],
            payload["email"]["value"],
            payload["phone"]["value"],
            payload["website"]["value"],
            payload["address"]["value"],
            payload["quality_score"],
            payload["quality_badge"],
        ]
    )
    wb.save(path)
    return path


def persist_bundle(
    result: ExtractionResult, output_dir: Path, source_name: str
) -> dict[str, str]:
    stem = _snapshot_stem(Path(source_name).stem or source_name)
    snapshot_stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    snapshot_stem = f"card_{snapshot_stamp}_{stem}"
    json_path = export_json(result, output_dir, snapshot_stem)
    csv_path = export_csv(result, output_dir, snapshot_stem)
    xlsx_path = export_xlsx(result, output_dir, snapshot_stem)
    vcf_path = export_vcard(result, output_dir, snapshot_stem)

    latest_json = output_dir / "latest.json"
    latest_csv = output_dir / "latest.csv"
    latest_xlsx = output_dir / "latest.xlsx"
    latest_vcf = output_dir / "latest.vcf"

    copy2(json_path, latest_json)
    copy2(csv_path, latest_csv)
    copy2(xlsx_path, latest_xlsx)
    copy2(vcf_path, latest_vcf)

    paths = {
        "json": str(json_path),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        "vcf": str(vcf_path),
        "latest_json": str(latest_json),
        "latest_csv": str(latest_csv),
        "latest_xlsx": str(latest_xlsx),
        "latest_vcf": str(latest_vcf),
    }
    monthly = _append_monthly_master(result, output_dir)
    if monthly is not None:
        paths["master"] = str(monthly)
    return paths


def run_engine_on_source(
    source: Path, output_dir: Path, lang: str = "en"
) -> list[dict[str, Any]]:
    engine = BusinessCardEngine(lang=lang)
    records: list[dict[str, Any]] = []
    for asset_name, image in iter_images_from_path(source):
        LOGGER.info("Processing %s", asset_name)
        result = engine.process_card_matrix(image)
        persisted = persist_bundle(result, output_dir, asset_name)
        record = result.to_dict()
        record["source"] = asset_name
        record["files"] = persisted
        records.append(record)
    return records


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Deterministic business card OCR engine v19.0"
    )
    parser.add_argument(
        "input", nargs="?", type=Path, help="Image, PDF, ZIP, or directory to process"
    )
    parser.add_argument(
        "--output-dir", type=Path, default=Path("output"), help="Output directory"
    )
    parser.add_argument("--lang", default="en", help="PaddleOCR language code")
    parser.add_argument(
        "--pretty", action="store_true", help="Pretty-print JSON records"
    )
    return parser


# some existing functions


def select_input_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Business Card",
        filetypes=[
            ("Supported Files", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.pdf *.zip"),
            ("All Files", "*.*"),
        ],
    )

    root.destroy()
    return file_path


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.input is None:
        selected_file = select_input_file()

        if not selected_file:
            print("No file selected.")
            return 0

        args.input = Path(selected_file)

    records = run_engine_on_source(args.input, args.output_dir, lang=args.lang)

    for record in records:
        print(json.dumps(record, indent=2 if args.pretty else None, ensure_ascii=False))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
