# #!/usr/bin/env python3
# """
# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║           VISITING CARD OCR EXTRACTOR  v5.0  —  FINAL EDITION              ║
# ║                                                                              ║
# ║  ► No hardcoded countries, cities, or regions — works on ANY card globally  ║
# ║  ► Multi-engine: EasyOCR  +  Tesseract fallback for maximum coverage        ║
# ║  ► Full deskew pipeline  — handles tilted / rotated cards                   ║
# ║  ► CLAHE + adaptive preprocessing — works on dark, light, and grey cards    ║
# ║  ► Triple-pass OCR — normal, brightened, darkened — catches all text tones  ║
# ║  ► Fuzzy reconstruction  — handles spaced letters, OCR artefacts            ║
# ║  ► Auto Y-tolerance — adapts to font size, no manual tuning needed          ║
# ║  ► Exports JSON + CSV per run                                               ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

# REQUIREMENTS  (install once):
#     pip install easyocr opencv-python-headless numpy pillow pytesseract

#     pytesseract also needs the Tesseract binary:
#         Windows : https://github.com/UB-Mannheim/tesseract/wiki
#         Linux   : sudo apt install tesseract-ocr
#         macOS   : brew install tesseract

# USAGE:
#     python visiting_card_ocr.py                  # GUI file picker (single image)
#     python visiting_card_ocr.py path/to/img.jpg  # CLI single image
#     python visiting_card_ocr.py path/to/folder/  # CLI batch — all images in folder
#     python visiting_card_ocr.py img.jpg --debug  # also save debug images
# """

# # ══════════════════════════════════════════════════════════════════════════════
# #  IMPORTS
# # ══════════════════════════════════════════════════════════════════════════════
# import os
# import re
# import sys
# import csv
# import json
# import math
# import warnings
# import argparse
# import unicodedata
# from pathlib import Path
# from datetime import datetime

# import cv2
# import numpy as np
# from PIL import Image

# # ── Optional imports handled gracefully ───────────────────────────────────────
# try:
#     import easyocr

#     EASYOCR_AVAILABLE = True
# except ImportError:
#     EASYOCR_AVAILABLE = False
#     print("  [WARN] easyocr not installed — pip install easyocr")

# try:
#     import pytesseract

#     TESSERACT_AVAILABLE = True
# except ImportError:
#     TESSERACT_AVAILABLE = False
#     print("  [WARN] pytesseract not installed — pip install pytesseract")

# try:
#     import torch

#     GPU_AVAILABLE = torch.cuda.is_available()
# except ImportError:
#     GPU_AVAILABLE = False

# warnings.filterwarnings("ignore")


# # ══════════════════════════════════════════════════════════════════════════════
# #  TUNEABLE CONSTANTS
# # ══════════════════════════════════════════════════════════════════════════════

# # Preprocessing
# DARK_BG_THRESHOLD = 128  # centre-pixel mean below this → dark card
# DARK_PIXEL_RATIO = 0.50  # fraction of centre pixels below threshold
# RESIZE_MAX_SIDE = 2200  # upscale to this; EasyOCR hates tiny images
# RESIZE_MIN_SIDE = 900  # downscale huge images to save RAM
# CLAHE_CLIP = 2.5
# CLAHE_TILE = (8, 8)
# UNSHARP_AMOUNT = 0.55
# STD_DEV_ADAPTIVE = 42  # σ below this → use adaptive threshold

# # OCR
# OCR_CONF_THRESHOLD = 0.28  # lower than default to catch faint grey text
# OCR_WIDTH_THS = 0.75
# OCR_DECODER = "greedy"

# # Line grouping
# LINE_Y_BASE_TOL = 28  # pixels; auto-scaled by median font height

# # Phone digit count validation (international)
# PHONE_MIN_DIGITS = 6
# PHONE_MAX_DIGITS = 15


# # ══════════════════════════════════════════════════════════════════════════════
# #  COMPILED REGEX
# # ══════════════════════════════════════════════════════════════════════════════

# _TLD = (
#     r"(?:com|in|org|net|co\.in|co\.uk|io|biz|info|edu|gov|"
#     r"ae|us|au|nz|sg|hk|my|ph|za|de|fr|jp|cn|ca|br|mx|"
#     r"co\.ae|co\.sg|co\.za|co\.au|co\.nz|co\.jp|co\.ca|"
#     r"tech|app|dev|ai)"
# )

# # Email — allows OCR-inserted spaces around @ and .
# _EMAIL_RE = re.compile(
#     r"[A-Za-z0-9._%+\-]+\s*@\s*[A-Za-z0-9.\-]+\s*\.\s*" + _TLD,
#     re.IGNORECASE,
# )

# # Website
# _WEB_RE = re.compile(
#     r"(?:https?://|www\.)[A-Za-z0-9.\-/_%?=&#]+"
#     r"|[A-Za-z0-9\-]+\.(?:com|in|org|net|io|co\.in|co\.uk|biz)[/\w\-]*",
#     re.IGNORECASE,
# )

# # Phone — liberal; digit-count validated afterwards
# _PHONE_RE = re.compile(
#     r"(?:\+\s*\d{1,4}[\s\-\.]*)?"
#     r"(?:\([\d\s]+\)[\s\-\.]*)?"
#     r"[\d][\d\s\-\.]{4,18}[\d]",
# )

# # LinkedIn / Twitter
# _LINKEDIN_RE = re.compile(r"linkedin\.com/in/[A-Za-z0-9_\-]+", re.IGNORECASE)
# _TWITTER_RE = re.compile(
#     r"(?:twitter\.com/|x\.com/|@)[A-Za-z0-9_]{2,50}", re.IGNORECASE
# )

# # Job title keywords — purely structural, no country assumption
# _TITLE_KW = re.compile(
#     r"\b(?:manager|director|engineer|developer|architect|founder|partner|"
#     r"consultant|analyst|officer|executive|president|associate|senior|"
#     r"junior|principal|head|lead|specialist|advisor|coordinator|"
#     r"supervisor|assistant|proprietor|owner|chairman|trustee|secretary|"
#     r"treasurer|intern|trainee|ceo|cto|coo|cfo|cmo|md|vp|"
#     r"vice\s*president|doctor|dr\.?|prof\.?|professor|lawyer|advocate|"
#     r"solicitor|barrister|accountant|auditor|designer|strategist|"
#     r"planner|researcher|scientist|technician|operator|representative|"
#     r"agent|broker|dealer|trader|contractor|builder)\b",
#     re.IGNORECASE,
# )

# # Company / org keywords
# _COMPANY_KW = re.compile(
#     r"\b(?:ltd|limited|inc|llp|llc|pvt|private|plc|corp|corporation|"
#     r"solutions|consulting|studio|technologies|tech|group|company|co\.|"
#     r"labs|services|systems|enterprises|associates|builders|construction|"
#     r"industries|global|international|ventures|holdings|capital|networks|"
#     r"media|digital|infra|infrastructure|realty|realtors|properties|"
#     r"developers|architects|interiors|design|agency|firm|bureau|office|"
#     r"foundation|trust|institute|academy|school|college|hospital|clinic|"
#     r"healthcare|pharma|logistics|transport|exports|imports|trading|"
#     r"manufacturing|fabrication|engineering)\b",
#     re.IGNORECASE,
# )

# # Address structural cues — works globally without any city list
# _ADDR_KW = re.compile(
#     r"\b(?:floor|fl\.|level|suite|no\.|plot|block|sector|phase|unit|"
#     r"road|rd\.|street|st\.|avenue|ave\.|lane|ln\.|boulevard|blvd\.|"
#     r"nagar|vihar|marg|gali|chowk|bazaar|market|colony|society|"
#     r"residency|plaza|centre|center|mall|complex|tower|building|"
#     r"district|area|zone|suburb|village|town|city|state|province|"
#     r"county|region|po\s*box|p\.o\.|zip|pin)\b",
#     re.IGNORECASE,
# )


# # ══════════════════════════════════════════════════════════════════════════════
# #  BBOX HELPERS
# # ══════════════════════════════════════════════════════════════════════════════


# def mid_y(bbox) -> float:
#     return (bbox[0][1] + bbox[2][1]) / 2.0


# def top_y(bbox) -> float:
#     return float(bbox[0][1])


# def left_x(bbox) -> float:
#     return float(bbox[0][0])


# def height_px(bbox) -> float:
#     return float(bbox[2][1] - bbox[0][1])


# def width_px(bbox) -> float:
#     return float(bbox[1][0] - bbox[0][0])


# def _bbox_key(bbox) -> str:
#     """Coarse grid key for deduplication across OCR engines."""
#     return f"{int(left_x(bbox)//14)},{int(top_y(bbox)//14)}"


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 1 — IMAGE LOADING
# # ══════════════════════════════════════════════════════════════════════════════


# def load_image(path: str) -> np.ndarray:
#     img = cv2.imread(path)
#     if img is None:
#         try:
#             pil = Image.open(path).convert("RGB")
#             img = cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)
#         except Exception as e:
#             raise ValueError(f"Cannot open image: {path}\n{e}")
#     return img


# def resize_for_ocr(img: np.ndarray) -> np.ndarray:
#     h, w = img.shape[:2]
#     max_s = max(h, w)
#     min_s = min(h, w)
#     if max_s < RESIZE_MIN_SIDE:
#         scale = RESIZE_MIN_SIDE / min_s
#     elif max_s > RESIZE_MAX_SIDE:
#         scale = RESIZE_MAX_SIDE / max_s
#     else:
#         scale = 1.0
#     if abs(scale - 1.0) > 0.02:
#         interp = cv2.INTER_CUBIC if scale > 1 else cv2.INTER_AREA
#         img = cv2.resize(img, None, fx=scale, fy=scale, interpolation=interp)
#     return img


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 2 — DESKEW
# # ══════════════════════════════════════════════════════════════════════════════


# def deskew(img: np.ndarray) -> tuple:
#     """
#     Detect and correct card tilt using Hough lines on Canny edges.
#     Returns (corrected_image, angle_degrees).
#     Safe: returns original image unchanged if tilt < 0.5° or detection fails.
#     """
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#     blur = cv2.GaussianBlur(gray, (5, 5), 0)
#     otsu_t, _ = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
#     edges = cv2.Canny(blur, otsu_t * 0.5, otsu_t)

#     lines = cv2.HoughLinesP(
#         edges, 1, math.pi / 180, threshold=80, minLineLength=60, maxLineGap=20
#     )
#     if lines is None or len(lines) < 3:
#         return img, 0.0

#     angles = []
#     for line in lines:
#         x1, y1, x2, y2 = line[0]
#         a = math.degrees(math.atan2(y2 - y1, x2 - x1))
#         if abs(a) < 45:
#             angles.append(a)

#     if not angles:
#         return img, 0.0

#     angle = float(np.median(angles))
#     if abs(angle) < 0.5:
#         return img, 0.0

#     h, w = img.shape[:2]
#     cx, cy = w // 2, h // 2
#     M = cv2.getRotationMatrix2D((cx, cy), angle, 1.0)
#     cos_a = abs(M[0, 0])
#     sin_a = abs(M[0, 1])
#     nw = int(h * sin_a + w * cos_a)
#     nh = int(h * cos_a + w * sin_a)
#     M[0, 2] += (nw - w) / 2
#     M[1, 2] += (nh - h) / 2
#     rotated = cv2.warpAffine(
#         img, M, (nw, nh), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE
#     )
#     return rotated, angle


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 3 — PREPROCESSING
# # ══════════════════════════════════════════════════════════════════════════════


# def _detect_dark_bg(gray: np.ndarray) -> bool:
#     h, w = gray.shape
#     y1, y2 = int(h * 0.30), int(h * 0.70)
#     x1, x2 = int(w * 0.30), int(w * 0.70)
#     c = gray[y1:y2, x1:x2]
#     return (
#         float(np.mean(c)) < DARK_BG_THRESHOLD
#         and float(np.sum(c < DARK_BG_THRESHOLD)) / c.size > DARK_PIXEL_RATIO
#     )


# def preprocess_image(img: np.ndarray) -> tuple:
#     """
#     Returns (processed_gray, is_dark, mode_string).

#     Pipeline:
#       1. Greyscale
#       2. Invert if dark background (text → dark on white)
#       3. CLAHE — local contrast, does NOT over-expose faint grey text
#       4. Unsharp mask — sharpens text edges
#       5. Low σ → adaptive threshold; else keep greyscale
#     """
#     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
#     is_dark = _detect_dark_bg(gray)

#     if is_dark:
#         gray = cv2.bitwise_not(gray)

#     clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP, tileGridSize=CLAHE_TILE)
#     gray = clahe.apply(gray)

#     # Unsharp mask
#     blur = cv2.GaussianBlur(gray, (0, 0), 3)
#     gray = cv2.addWeighted(gray, 1 + UNSHARP_AMOUNT, blur, -UNSHARP_AMOUNT, 0)

#     std = float(np.std(gray))

#     if std < STD_DEV_ADAPTIVE:
#         mode = f"ADAPTIVE THRESHOLD (σ={std:.1f})"
#         gray = cv2.adaptiveThreshold(
#             gray,
#             255,
#             cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
#             cv2.THRESH_BINARY,
#             blockSize=21,
#             C=10,
#         )
#     else:
#         mode = f"CLAHE+UNSHARP (σ={std:.1f}, dark={is_dark})"

#     return gray, is_dark, mode


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 4 — OCR ENGINES
# # ══════════════════════════════════════════════════════════════════════════════

# _EASYOCR_READER = None


# def _get_reader():
#     global _EASYOCR_READER
#     if _EASYOCR_READER is None and EASYOCR_AVAILABLE:
#         _EASYOCR_READER = easyocr.Reader(["en"], gpu=GPU_AVAILABLE, verbose=False)
#     return _EASYOCR_READER


# def run_easyocr(processed: np.ndarray) -> list:
#     """
#     Triple-pass EasyOCR:
#       Pass 1 — image as-is          (normal contrast)
#       Pass 2 — brightened +30       (catches dark/faint text)
#       Pass 3 — darkened  -25        (catches light-on-slightly-grey backgrounds)
#     All three results are merged; highest-confidence token wins per grid cell.
#     """
#     reader = _get_reader()
#     if reader is None:
#         return []

#     pool = {}

#     def _pass(arr):
#         for bbox, text, conf in reader.readtext(
#             arr, width_ths=OCR_WIDTH_THS, decoder=OCR_DECODER, paragraph=False
#         ):
#             k = _bbox_key(bbox)
#             if k not in pool or conf > pool[k][2]:
#                 pool[k] = (bbox, text, conf)

#     _pass(processed)
#     _pass(cv2.convertScaleAbs(processed, alpha=1.35, beta=25))
#     _pass(cv2.convertScaleAbs(processed, alpha=0.72, beta=-15))

#     return list(pool.values())


# def run_tesseract(processed: np.ndarray) -> list:
#     """
#     Tesseract supplement — PSM 6 (block), 11 (sparse), 3 (auto).
#     Uses the same (bbox, text, conf) structure as EasyOCR for unified handling.
#     """
#     if not TESSERACT_AVAILABLE:
#         return []

#     pil = Image.fromarray(processed)
#     pool = {}

#     for psm in [6, 11, 3]:
#         try:
#             data = pytesseract.image_to_data(
#                 pil,
#                 config=f"--psm {psm} --oem 3",
#                 output_type=pytesseract.Output.DICT,
#             )
#         except Exception:
#             continue

#         for i in range(len(data["text"])):
#             text = data["text"][i].strip()
#             if not text:
#                 continue
#             raw_conf = data["conf"][i]
#             if isinstance(raw_conf, str):
#                 raw_conf = (
#                     float(raw_conf) if raw_conf.strip() not in ("-1", "") else 0.0
#                 )
#             conf = max(0.0, float(raw_conf) / 100.0)

#             x, y, w, h = (
#                 data["left"][i],
#                 data["top"][i],
#                 data["width"][i],
#                 data["height"][i],
#             )
#             if w < 4 or h < 4:
#                 continue

#             bbox = [[x, y], [x + w, y], [x + w, y + h], [x, y + h]]
#             k = _bbox_key(bbox)
#             if k not in pool or conf > pool[k][2]:
#                 pool[k] = (bbox, text, conf)

#     return list(pool.values())


# def merge_ocr(easy: list, tess: list) -> list:
#     merged = {}
#     for bbox, text, conf in easy + tess:
#         k = _bbox_key(bbox)
#         if k not in merged or conf > merged[k][2]:
#             merged[k] = (bbox, text, conf)
#     return list(merged.values())


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 5 — TOKEN CLEANING
# # ══════════════════════════════════════════════════════════════════════════════


# def collapse_spaced(text: str) -> str:
#     """'A D I T Y A'  →  'ADITYA'  |  'V A R M A'  →  'VARMA'"""
#     segs = text.strip().split()
#     if len(segs) >= 2 and all(re.fullmatch(r"[A-Za-z]\.?", s) for s in segs):
#         return "".join(segs)
#     return text


# _DIGIT_MAP = str.maketrans(
#     {
#         "O": "0",
#         "o": "0",
#         "I": "1",
#         "l": "1",
#         "S": "5",
#         "s": "5",
#         "B": "8",
#         "G": "6",
#         "g": "9",
#         "Z": "2",
#         "z": "2",
#     }
# )


# def fix_digits(text: str) -> str:
#     return text.translate(_DIGIT_MAP)


# def clean_text(text: str) -> str:
#     text = unicodedata.normalize("NFKC", text)
#     text = re.sub(r" {2,}", " ", text).strip()
#     return text


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 6 — LINE GROUPING
# # ══════════════════════════════════════════════════════════════════════════════


# def group_lines(tokens: list) -> list:
#     """
#     Sort by Y then X; group tokens whose mid-Y values are within a dynamic
#     tolerance (based on median font height) into the same line.
#     """
#     if not tokens:
#         return []

#     heights = [height_px(b) for b, _, _ in tokens]
#     med_h = float(np.median(heights)) if heights else 20
#     y_tol = max(LINE_Y_BASE_TOL, int(med_h * 0.65))

#     sorted_t = sorted(tokens, key=lambda r: (mid_y(r[0]), left_x(r[0])))

#     lines = []
#     current = [sorted_t[0]]

#     for item in sorted_t[1:]:
#         if abs(mid_y(item[0]) - mid_y(current[0][0])) <= y_tol:
#             current.append(item)
#         else:
#             lines.append(sorted(current, key=lambda r: left_x(r[0])))
#             current = [item]
#     lines.append(sorted(current, key=lambda r: left_x(r[0])))
#     return lines


# def line_text(line: list, sep: str = " ") -> str:
#     parts = [collapse_spaced(clean_text(t)) for _, t, _ in line if t.strip()]
#     return sep.join(p for p in parts if p)


# def line_conf(line: list) -> float:
#     return sum(c for _, _, c in line) / len(line) if line else 0.0


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 7 — FIELD EXTRACTION
# # ══════════════════════════════════════════════════════════════════════════════

# # ── Email ─────────────────────────────────────────────────────────────────────


# def extract_email(lines: list) -> str:
#     hits = []
#     for line in lines:
#         # Full joined line (no spaces)
#         full = re.sub(r"\s+", "", line_text(line))
#         m = _EMAIL_RE.search(full)
#         if m:
#             hits.append((line_conf(line), re.sub(r"\s+", "", m.group()).lower()))
#         # Per-token
#         for _, t, c in line:
#             t2 = re.sub(r"\s+", "", t)
#             m2 = _EMAIL_RE.search(t2)
#             if m2:
#                 hits.append((c, re.sub(r"\s+", "", m2.group()).lower()))
#     return max(hits, key=lambda x: x[0])[1] if hits else ""


# # ── Phones ────────────────────────────────────────────────────────────────────


# def extract_phones(lines: list) -> list:
#     found = []
#     for line in lines:
#         txt = fix_digits(line_text(line))
#         for m in _PHONE_RE.finditer(txt):
#             raw = m.group()
#             digits = re.sub(r"\D", "", raw)
#             if PHONE_MIN_DIGITS <= len(digits) <= PHONE_MAX_DIGITS:
#                 # Normalise: collapse single-digit-spaces like '2 2 0 0' → '2200'
#                 norm = re.sub(r"(?<=\d) (?=\d)", "", raw)
#                 norm = re.sub(r"\s{2,}", " ", norm).strip()
#                 if norm and norm not in found:
#                     found.append(norm)
#     return found


# # ── Website ───────────────────────────────────────────────────────────────────


# def extract_website(lines: list) -> str:
#     for line in lines:
#         m = _WEB_RE.search(line_text(line))
#         if m:
#             url = m.group().strip()
#             return ("http://" + url if not url.startswith("http") else url).lower()
#     return ""


# # ── Name ─────────────────────────────────────────────────────────────────────


# def extract_name(lines: list, raw_tokens: list) -> str:
#     """
#     Name extraction — purely structural, no hardcoded names or locations:

#     1. Reject tokens with digits, @, company/job/address keywords, <60% alpha.
#     2. From survivors, group into lines.
#     3. Score each line: font_height × confidence × vertical_position_bonus.
#     4. Pick the highest-scoring line.
#     5. Title-case if all-caps.
#     """
#     if not raw_tokens:
#         return ""

#     all_bottom_y = [top_y(b) + height_px(b) for b, _, _ in raw_tokens]
#     card_h = max(all_bottom_y) if all_bottom_y else 1

#     def _ok(bbox, text, conf):
#         if conf < OCR_CONF_THRESHOLD:
#             return False
#         t = text.strip()
#         if len(t) < 2:
#             return False
#         if re.search(r"\d", t) or "@" in t:
#             return False
#         if _COMPANY_KW.search(t) or _TITLE_KW.search(t) or _ADDR_KW.search(t):
#             return False
#         alpha_r = sum(c.isalpha() for c in t) / max(len(t), 1)
#         if alpha_r < 0.60:
#             return False
#         return True

#     cands = [(b, t, c) for b, t, c in raw_tokens if _ok(b, t, c)]
#     if not cands:
#         return ""

#     cand_lines = group_lines(cands)

#     def _score(ln):
#         avg_h = float(np.mean([height_px(b) for b, _, _ in ln]))
#         avg_c = line_conf(ln)
#         avg_y = float(np.mean([top_y(b) for b, _, _ in ln]))
#         pos_b = 1.0 - (avg_y / card_h)
#         return avg_h * avg_c * (1 + 0.35 * pos_b)

#     best = max(cand_lines, key=_score)
#     parts = []
#     for _, t, _ in sorted(best, key=lambda r: left_x(r[0])):
#         p = collapse_spaced(clean_text(t))
#         if p.isupper() and len(p) > 2:
#             p = p.title()
#         parts.append(p)

#     name = " ".join(parts)
#     name = re.sub(r"^[^A-Za-z]+|[^A-Za-z.'\-]+$", "", name).strip()
#     return name


# # ── Company ───────────────────────────────────────────────────────────────────


# def extract_company(lines: list) -> str:
#     scored = []
#     for line in lines:
#         t = line_text(line).strip()
#         if not t:
#             continue
#         score = 0
#         if _COMPANY_KW.search(t):
#             score += 10
#         if t.isupper() and len(t.split()) >= 2:
#             score += 3
#         if len(t.split()) >= 2:
#             score += 1
#         if "@" in t or re.search(r"\d{4,}", t):
#             score -= 10
#         if _ADDR_KW.search(t):
#             score -= 5
#         if _TITLE_KW.search(t):
#             score -= 4
#         if score > 0:
#             scored.append((score, line_conf(line), t))

#     if not scored:
#         return ""
#     scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
#     return scored[0][2]


# # ── Job title ─────────────────────────────────────────────────────────────────


# def extract_job_title(lines: list) -> str:
#     for line in lines:
#         t = line_text(line).strip()
#         if _TITLE_KW.search(t) and not _COMPANY_KW.search(t):
#             return t
#     return ""


# # ── Address ───────────────────────────────────────────────────────────────────


# def extract_address(lines: list) -> str:
#     """
#     Detect address fragments using ONLY structural cues (floor/road/complex
#     keywords and standalone pincodes).  Zero hardcoded city / country names.
#     Works globally.
#     """
#     parts = []
#     for line in lines:
#         t = line_text(line).strip()
#         if not t:
#             continue
#         is_addr = False
#         if _ADDR_KW.search(t):
#             is_addr = True
#         # Standalone numeric block that looks like a pincode / ZIP
#         if re.search(r"\b\d{4,9}\b", t) and not re.search(r"[+\-]\s*\d{6,}", t):
#             is_addr = True
#         # "Something, Something" with initial caps — city/country pattern
#         if re.search(r"[A-Z][a-z]+,\s*[A-Z][a-z]+", t):
#             is_addr = True
#         if is_addr:
#             parts.append(t)
#     return ", ".join(parts) if parts else ""


# # ── Social ────────────────────────────────────────────────────────────────────


# def extract_social(lines: list) -> dict:
#     r = {"linkedin": "", "twitter": ""}
#     for line in lines:
#         t = line_text(line)
#         if not r["linkedin"]:
#             m = _LINKEDIN_RE.search(t)
#             if m:
#                 r["linkedin"] = m.group().strip()
#         if not r["twitter"]:
#             m = _TWITTER_RE.search(t)
#             if m:
#                 r["twitter"] = m.group().strip()
#     return r


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 8 — FULL PIPELINE
# # ══════════════════════════════════════════════════════════════════════════════


# def process_card(image_path: str, debug: bool = False) -> dict:
#     result = {
#         "file": os.path.basename(image_path),
#         "name": "",
#         "job_title": "",
#         "company": "",
#         "email": "",
#         "phones": [],
#         "website": "",
#         "address": "",
#         "linkedin": "",
#         "twitter": "",
#         "_debug": {},
#     }

#     # 1. Load + resize
#     img = load_image(image_path)
#     img = resize_for_ocr(img)

#     # 2. Deskew
#     img, skew = deskew(img)

#     # 3. Preprocess
#     proc, is_dark, mode = preprocess_image(img)

#     if debug:
#         result["_debug"].update({"skew": round(skew, 2), "dark": is_dark, "mode": mode})
#         d = os.path.dirname(os.path.abspath(image_path))
#         cv2.imwrite(os.path.join(d, "DEBUG_preprocessed_v5.png"), proc)

#     # 4. OCR (EasyOCR + Tesseract merged)
#     easy = run_easyocr(proc)
#     tess = run_tesseract(proc)
#     tokens = merge_ocr(easy, tess)
#     kept = [(b, t, c) for b, t, c in tokens if c >= OCR_CONF_THRESHOLD]

#     if debug:
#         result["_debug"]["total"] = len(tokens)
#         result["_debug"]["kept"] = len(kept)
#         result["_debug"]["raw"] = [
#             {"text": t, "conf": round(c, 3), "y": round(top_y(b), 1)}
#             for b, t, c in sorted(tokens, key=lambda r: top_y(r[0]))
#         ]

#     # 5. Line grouping
#     lines = group_lines(kept)

#     if debug:
#         result["_debug"]["lines"] = [
#             {"n": i + 1, "text": line_text(ln), "conf": round(line_conf(ln), 3)}
#             for i, ln in enumerate(lines)
#         ]

#     # 6. Extract fields
#     result["email"] = extract_email(lines)
#     result["phones"] = extract_phones(lines)
#     result["website"] = extract_website(lines)
#     result["name"] = extract_name(lines, kept)
#     result["company"] = extract_company(lines)
#     result["job_title"] = extract_job_title(lines)
#     result["address"] = extract_address(lines)
#     soc = extract_social(lines)
#     result["linkedin"] = soc["linkedin"]
#     result["twitter"] = soc["twitter"]

#     # 7. Debug annotation
#     if debug:
#         _annotate(img, tokens, image_path)

#     return result


# def _annotate(img: np.ndarray, tokens: list, src: str):
#     ann = img.copy()
#     for bbox, text, conf in tokens:
#         pts = np.array([[int(p[0]), int(p[1])] for p in bbox], np.int32)
#         color = (0, 200, 0) if conf >= OCR_CONF_THRESHOLD else (0, 0, 200)
#         cv2.polylines(ann, [pts], True, color, 2)
#         label = f"{text[:24]}{'…' if len(text)>24 else ''} [{conf:.2f}]"
#         pos = (max(pts[0][0], 0), max(pts[0][1] - 5, 12))
#         (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.4, 1)
#         cv2.rectangle(
#             ann, (pos[0], pos[1] - lh - 2), (pos[0] + lw, pos[1] + 2), (20, 20, 20), -1
#         )
#         cv2.putText(
#             ann, label, pos, cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1, cv2.LINE_AA
#         )
#     out = os.path.join(os.path.dirname(os.path.abspath(src)), "DEBUG_annotated_v5.png")
#     cv2.imwrite(out, ann)
#     print(f"  [DEBUG] Annotated → {out}")


# # ══════════════════════════════════════════════════════════════════════════════
# #  STEP 9 — OUTPUT
# # ══════════════════════════════════════════════════════════════════════════════


# def print_result(r: dict):
#     w = 60
#     print()
#     print("═" * w)
#     print(f"  FILE      : {r['file']}")
#     print("─" * w)
#     print(f"  NAME      : {r['name']      or '—'}")
#     print(f"  JOB TITLE : {r['job_title'] or '—'}")
#     print(f"  COMPANY   : {r['company']   or '—'}")
#     print(f"  EMAIL     : {r['email']     or '—'}")
#     for i, ph in enumerate(r["phones"], 1):
#         label = f"  PHONE {'#'+str(i):<4}"
#         print(f"{label}: {ph}")
#     if not r["phones"]:
#         print(f"  PHONE     : —")
#     print(f"  WEBSITE   : {r['website']   or '—'}")
#     print(f"  ADDRESS   : {r['address']   or '—'}")
#     print(f"  LINKEDIN  : {r['linkedin']  or '—'}")
#     print(f"  TWITTER/X : {r['twitter']   or '—'}")
#     print("═" * w)

#     if r.get("_debug"):
#         print(
#             f"\n  [DEBUG] skew={r['_debug'].get('skew')}°  "
#             f"dark={r['_debug'].get('dark')}  "
#             f"mode={r['_debug'].get('mode')}"
#         )
#         print(
#             f"  [DEBUG] tokens total={r['_debug'].get('total')}  "
#             f"kept={r['_debug'].get('kept')}"
#         )
#         print()
#         print("  RAW TOKENS:")
#         for tk in r["_debug"].get("raw", []):
#             mark = "  " if tk["conf"] >= OCR_CONF_THRESHOLD else "✗ "
#             print(f"  {mark}y={tk['y']:>6.1f}  conf={tk['conf']:.3f}  {tk['text']!r}")
#         print()
#         print("  GROUPED LINES:")
#         for ln in r["_debug"].get("lines", []):
#             print(f"  Line {ln['n']:02d} (conf={ln['conf']:.2f}) : {ln['text']!r}")
#         print()


# def _ts() -> str:
#     return datetime.now().strftime("%Y%m%d_%H%M%S")


# def save_json(results: list, out_dir: str):
#     path = os.path.join(out_dir, f"cards_{_ts()}.json")
#     clean = [{k: v for k, v in r.items() if k != "_debug"} for r in results]
#     with open(path, "w", encoding="utf-8") as f:
#         json.dump(clean, f, indent=2, ensure_ascii=False)
#     print(f"  [JSON] → {path}")


# def save_csv(results: list, out_dir: str):
#     path = os.path.join(out_dir, f"cards_{_ts()}.csv")
#     fields = [
#         "file",
#         "name",
#         "job_title",
#         "company",
#         "email",
#         "phone_1",
#         "phone_2",
#         "website",
#         "address",
#         "linkedin",
#         "twitter",
#     ]
#     with open(path, "w", newline="", encoding="utf-8") as f:
#         w = csv.DictWriter(f, fieldnames=fields)
#         w.writeheader()
#         for r in results:
#             ph = r.get("phones", [])
#             w.writerow(
#                 {
#                     "file": r.get("file", ""),
#                     "name": r.get("name", ""),
#                     "job_title": r.get("job_title", ""),
#                     "company": r.get("company", ""),
#                     "email": r.get("email", ""),
#                     "phone_1": ph[0] if ph else "",
#                     "phone_2": ph[1] if len(ph) > 1 else "",
#                     "website": r.get("website", ""),
#                     "address": r.get("address", ""),
#                     "linkedin": r.get("linkedin", ""),
#                     "twitter": r.get("twitter", ""),
#                 }
#             )
#     print(f"  [CSV]  → {path}")


# # ══════════════════════════════════════════════════════════════════════════════
# #  ENTRY POINT
# # ══════════════════════════════════════════════════════════════════════════════

# _IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp", ".heic"}


# def collect_paths(target: str) -> list:
#     p = Path(target)
#     if p.is_file():
#         return [str(p)]
#     if p.is_dir():
#         return sorted(str(f) for f in p.iterdir() if f.suffix.lower() in _IMG_EXTS)
#     return []


# def gui_pick() -> list:
#     try:
#         from tkinter import Tk, filedialog

#         root = Tk()
#         root.withdraw()
#         paths = filedialog.askopenfilenames(
#             title="Select Visiting Card Image(s)",
#             filetypes=[("Images", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif *.webp")],
#         )
#         root.destroy()
#         return list(paths)
#     except Exception:
#         print("  [WARN] Tkinter not available. Pass the image path as a CLI arg.")
#         return []


# def main():
#     ap = argparse.ArgumentParser(
#         description="Visiting Card OCR Extractor v5.0 — global, no hardcoded regions"
#     )
#     ap.add_argument(
#         "target",
#         nargs="?",
#         default=None,
#         help="Image file or folder (blank → GUI picker)",
#     )
#     ap.add_argument(
#         "--debug",
#         action="store_true",
#         help="Save DEBUG_preprocessed_v5.png, DEBUG_annotated_v5.png, "
#         "and dump raw token/line info to terminal",
#     )
#     ap.add_argument("--no-json", action="store_true")
#     ap.add_argument("--no-csv", action="store_true")
#     args = ap.parse_args()

#     paths = collect_paths(args.target) if args.target else gui_pick()
#     if not paths:
#         print("  No images found. Exiting.")
#         sys.exit(0 if not args.target else 1)

#     print(f"\n  {len(paths)} image(s) to process.\n")

#     results = []
#     out_dir = os.path.dirname(os.path.abspath(paths[0]))

#     for i, path in enumerate(paths, 1):
#         print(f"  [{i}/{len(paths)}] {os.path.basename(path)}")
#         try:
#             r = process_card(path, debug=args.debug)
#             print_result(r)
#             results.append(r)
#         except Exception as e:
#             print(f"  [ERROR] {path}: {e}")
#             if args.debug:
#                 import traceback

#                 traceback.print_exc()

#     if results:
#         if not args.no_json:
#             save_json(results, out_dir)
#         if not args.no_csv:
#             save_csv(results, out_dir)

#     print("\n  Done.")


# if __name__ == "__main__":
#     main()


#!/usr/bin/env python3


#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║           VISITING CARD OCR EXTRACTOR  v5.0  —  FINAL EDITION              ║
║                                                                              ║
║  ► No hardcoded countries, cities, or regions — works on ANY card globally  ║
║  ► Multi-engine: EasyOCR  +  Tesseract fallback for maximum coverage        ║
║  ► Full deskew pipeline  — handles tilted / rotated cards                   ║
║  ► CLAHE + adaptive preprocessing — works on dark, light, and grey cards    ║
║  ► Triple-pass OCR — normal, brightened, darkened — catches all text tones  ║
║  ► Fuzzy reconstruction  — handles spaced letters, OCR artefacts            ║
║  ► Auto Y-tolerance — adapts to font size, no manual tuning needed          ║
║  ► Exports JSON + XLSX per run                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, re, sys, json, math, warnings, argparse, unicodedata
from pathlib import Path
from datetime import datetime

import cv2
import numpy as np
from PIL import Image

try:
    import easyocr

    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False
    print("  [WARN] easyocr not installed — pip install easyocr")

try:
    import pytesseract

    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False
    print("  [WARN] pytesseract not installed — pip install pytesseract")

try:
    import torch

    GPU_AVAILABLE = torch.cuda.is_available()
except ImportError:
    GPU_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("  [WARN] openpyxl not installed — pip install openpyxl  (no Excel output)")

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
#  TUNEABLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
DARK_BG_THRESHOLD = 128
DARK_PIXEL_RATIO = 0.50
RESIZE_MAX_SIDE = 2200
RESIZE_MIN_SIDE = 900
CLAHE_CLIP = 2.5
CLAHE_TILE = (8, 8)
UNSHARP_AMOUNT = 0.55
STD_DEV_ADAPTIVE = 42
OCR_CONF_THRESHOLD = 0.28
OCR_WIDTH_THS = 0.75
OCR_DECODER = "greedy"
LINE_Y_BASE_TOL = 28
PHONE_MIN_DIGITS = 6
PHONE_MAX_DIGITS = 15

# ══════════════════════════════════════════════════════════════════════════════
#  COMPILED REGEX
# ══════════════════════════════════════════════════════════════════════════════
_TLD = (
    r"(?:com|in|org|net|co\.in|co\.uk|io|biz|info|edu|gov|"
    r"ae|us|au|nz|sg|hk|my|ph|za|de|fr|jp|cn|ca|br|mx|"
    r"co\.ae|co\.sg|co\.za|co\.au|co\.nz|co\.jp|co\.ca|"
    r"tech|app|dev|ai)"
)
_EMAIL_RE = re.compile(
    r"[A-Za-z0-9._%+\-]+\s*@\s*[A-Za-z0-9.\-]+\s*\.\s*" + _TLD, re.IGNORECASE
)
_WEB_RE = re.compile(
    r"(?:https?://|www\.)[A-Za-z0-9.\-/_%?=&#]+"
    r"|[A-Za-z0-9\-]+\.(?:com|in|org|net|io|co\.in|co\.uk|biz)[/\w\-]*",
    re.IGNORECASE,
)
_PHONE_RE = re.compile(
    r"(?:\+\s*\d{1,4}[\s\-\.]*)?(?:\([\d\s]+\)[\s\-\.]*)?[\d][\d\s\-\.]{4,18}[\d]"
)
_LINKEDIN_RE = re.compile(r"linkedin\.com/in/[A-Za-z0-9_\-]+", re.IGNORECASE)
_TWITTER_RE = re.compile(
    r"(?:twitter\.com/|x\.com/|@)[A-Za-z0-9_]{2,50}", re.IGNORECASE
)

_TITLE_KW = re.compile(
    r"\b(?:manager|director|engineer|developer|architect|founder|partner|"
    r"consultant|analyst|officer|executive|president|associate|senior|"
    r"junior|principal|head|lead|specialist|advisor|coordinator|"
    r"supervisor|assistant|proprietor|owner|chairman|trustee|secretary|"
    r"treasurer|intern|trainee|ceo|cto|coo|cfo|cmo|md|vp|"
    r"vice\s*president|doctor|dr\.?|prof\.?|professor|lawyer|advocate|"
    r"solicitor|barrister|accountant|auditor|designer|strategist|"
    r"planner|researcher|scientist|technician|operator|representative|"
    r"agent|broker|dealer|trader|contractor|builder)\b",
    re.IGNORECASE,
)

_COMPANY_KW = re.compile(
    r"\b(?:ltd|limited|inc|llp|llc|pvt|private|plc|corp|corporation|"
    r"solutions|consulting|studio|technologies|tech|group|company|co\.|"
    r"labs|services|systems|enterprises|associates|builders|construction|"
    r"industries|global|international|ventures|holdings|capital|networks|"
    r"media|digital|infra|infrastructure|realty|realtors|properties|"
    r"developers|architects|interiors|design|agency|firm|bureau|office|"
    r"foundation|trust|institute|academy|school|college|hospital|clinic|"
    r"healthcare|pharma|logistics|transport|exports|imports|trading|"
    r"manufacturing|fabrication|engineering)\b",
    re.IGNORECASE,
)

_ADDR_KW = re.compile(
    r"\b(?:floor|fl\.|level|suite|no\.|plot|block|sector|phase|unit|"
    r"road|rd\.|street|st\.|avenue|ave\.|lane|ln\.|boulevard|blvd\.|"
    r"nagar|vihar|marg|gali|chowk|bazaar|market|colony|society|"
    r"residency|plaza|centre|center|mall|complex|tower|building|"
    r"district|area|zone|suburb|village|town|city|state|province|"
    r"county|region|po\s*box|p\.o\.|zip|pin)\b",
    re.IGNORECASE,
)


# ══════════════════════════════════════════════════════════════════════════════
#  BBOX HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def mid_y(bbox):
    return (bbox[0][1] + bbox[2][1]) / 2.0


def top_y(bbox):
    return float(bbox[0][1])


def left_x(bbox):
    return float(bbox[0][0])


def height_px(bbox):
    return float(bbox[2][1] - bbox[0][1])


def width_px(bbox):
    return float(bbox[1][0] - bbox[0][0])


def _bbox_key(bbox):
    return f"{int(left_x(bbox)//14)},{int(top_y(bbox)//14)}"


# ══════════════════════════════════════════════════════════════════════════════
#  IMAGE LOADING
# ══════════════════════════════════════════════════════════════════════════════
def load_image(path):
    img = cv2.imread(path)
    if img is None:
        try:
            pil = Image.open(path).convert("RGB")
            img = cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)
        except Exception as e:
            raise ValueError(f"Cannot open image: {path}\n{e}")
    return img


def resize_for_ocr(img):
    h, w = img.shape[:2]
    max_s, min_s = max(h, w), min(h, w)
    if max_s < RESIZE_MIN_SIDE:
        scale = RESIZE_MIN_SIDE / min_s
    elif max_s > RESIZE_MAX_SIDE:
        scale = RESIZE_MAX_SIDE / max_s
    else:
        scale = 1.0
    if abs(scale - 1.0) > 0.02:
        interp = cv2.INTER_CUBIC if scale > 1 else cv2.INTER_AREA
        img = cv2.resize(img, None, fx=scale, fy=scale, interpolation=interp)
    return img


# ══════════════════════════════════════════════════════════════════════════════
#  DESKEW
# ══════════════════════════════════════════════════════════════════════════════
def deskew(img):

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    blur = cv2.GaussianBlur(gray, (5, 5), 0)

    otsu_t, _ = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    edges = cv2.Canny(blur, otsu_t * 0.5, otsu_t)

    lines = cv2.HoughLinesP(
        edges, 1, math.pi / 180, threshold=80, minLineLength=60, maxLineGap=20
    )

    # No usable lines found
    if lines is None or len(lines) < 3:
        return img, 0.0

    angles = []

    # Extract valid angles
    for line in lines:

        x1, y1, x2, y2 = line[0]

        angle = math.degrees(math.atan2(y2 - y1, x2 - x1))

        # Ignore vertical/extreme lines
        if abs(angle) < 45:
            angles.append(angle)

    # No horizontal-like lines
    if not angles:
        return img, 0.0

    median_angle = float(np.median(angles))

    # Ignore tiny skew
    if abs(median_angle) < 0.5:
        return img, 0.0

    h, w = img.shape[:2]

    center = (w // 2, h // 2)

    M = cv2.getRotationMatrix2D(center, median_angle, 1.0)

    cos_a = abs(M[0, 0])
    sin_a = abs(M[0, 1])

    new_w = int((h * sin_a) + (w * cos_a))
    new_h = int((h * cos_a) + (w * sin_a))

    # Adjust translation
    M[0, 2] += (new_w - w) / 2
    M[1, 2] += (new_h - h) / 2

    rotated = cv2.warpAffine(
        img, M, (new_w, new_h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE
    )

    return rotated, median_angle


# ══════════════════════════════════════════════════════════════════════════════
#  PREPROCESSING
# ══════════════════════════════════════════════════════════════════════════════
def _detect_dark_bg(gray):
    h, w = gray.shape
    y1, y2 = int(h * 0.30), int(h * 0.70)
    x1, x2 = int(w * 0.30), int(w * 0.70)
    c = gray[y1:y2, x1:x2]
    return (
        float(np.mean(c)) < DARK_BG_THRESHOLD
        and float(np.sum(c < DARK_BG_THRESHOLD)) / c.size > DARK_PIXEL_RATIO
    )


def preprocess_image(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    is_dark = _detect_dark_bg(gray)
    if is_dark:
        gray = cv2.bitwise_not(gray)
    clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP, tileGridSize=CLAHE_TILE)
    gray = clahe.apply(gray)
    blur = cv2.GaussianBlur(gray, (0, 0), 3)
    gray = cv2.addWeighted(gray, 1 + UNSHARP_AMOUNT, blur, -UNSHARP_AMOUNT, 0)
    std = float(np.std(gray))
    if std < STD_DEV_ADAPTIVE:
        mode = f"ADAPTIVE THRESHOLD (σ={std:.1f})"
        gray = cv2.adaptiveThreshold(
            gray,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            blockSize=21,
            C=10,
        )
    else:
        mode = f"CLAHE+UNSHARP (σ={std:.1f}, dark={is_dark})"
    return gray, is_dark, mode


# ══════════════════════════════════════════════════════════════════════════════
#  OCR ENGINES
# ══════════════════════════════════════════════════════════════════════════════
_EASYOCR_READER = None


def _get_reader():
    global _EASYOCR_READER
    if _EASYOCR_READER is None and EASYOCR_AVAILABLE:
        _EASYOCR_READER = easyocr.Reader(["en"], gpu=GPU_AVAILABLE, verbose=False)
    return _EASYOCR_READER


def run_easyocr(processed):
    reader = _get_reader()
    if reader is None:
        return []
    pool = {}

    def _pass(arr):
        for bbox, text, conf in reader.readtext(
            arr, width_ths=OCR_WIDTH_THS, decoder=OCR_DECODER, paragraph=False
        ):
            k = _bbox_key(bbox)
            if k not in pool or conf > pool[k][2]:
                pool[k] = (bbox, text, conf)

    _pass(processed)
    _pass(cv2.convertScaleAbs(processed, alpha=1.35, beta=25))
    _pass(cv2.convertScaleAbs(processed, alpha=0.72, beta=-15))
    return list(pool.values())


def run_tesseract(processed):
    if not TESSERACT_AVAILABLE:
        return []
    pil = Image.fromarray(processed)
    pool = {}
    for psm in [6, 11, 3]:
        try:
            data = pytesseract.image_to_data(
                pil, config=f"--psm {psm} --oem 3", output_type=pytesseract.Output.DICT
            )
        except Exception:
            continue
        for i in range(len(data["text"])):
            text = data["text"][i].strip()
            if not text:
                continue
            raw_conf = data["conf"][i]
            if isinstance(raw_conf, str):
                raw_conf = (
                    float(raw_conf) if raw_conf.strip() not in ("-1", "") else 0.0
                )
            conf = max(0.0, float(raw_conf) / 100.0)
            x, y, w, h = (
                data["left"][i],
                data["top"][i],
                data["width"][i],
                data["height"][i],
            )
            if w < 4 or h < 4:
                continue
            bbox = [[x, y], [x + w, y], [x + w, y + h], [x, y + h]]
            k = _bbox_key(bbox)
            if k not in pool or conf > pool[k][2]:
                pool[k] = (bbox, text, conf)
    return list(pool.values())


def merge_ocr(easy, tess):
    merged = {}
    for bbox, text, conf in easy + tess:
        k = _bbox_key(bbox)
        if k not in merged or conf > merged[k][2]:
            merged[k] = (bbox, text, conf)
    return list(merged.values())


# ══════════════════════════════════════════════════════════════════════════════
#  TOKEN CLEANING
# ══════════════════════════════════════════════════════════════════════════════
def collapse_spaced(text):
    segs = text.strip().split()
    if len(segs) >= 2 and all(re.fullmatch(r"[A-Za-z]\.?", s) for s in segs):
        return "".join(segs)
    return text


_DIGIT_MAP = str.maketrans(
    {
        "O": "0",
        "o": "0",
        "I": "1",
        "l": "1",
        "S": "5",
        "s": "5",
        "B": "8",
        "G": "6",
        "g": "9",
        "Z": "2",
        "z": "2",
    }
)


def fix_digits(text):
    return text.translate(_DIGIT_MAP)


def clean_text(text):
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r" {2,}", " ", text).strip()


# ══════════════════════════════════════════════════════════════════════════════
#  LINE GROUPING
# ══════════════════════════════════════════════════════════════════════════════
def group_lines(tokens):
    if not tokens:
        return []
    heights = [height_px(b) for b, _, _ in tokens]
    med_h = float(np.median(heights)) if heights else 20
    y_tol = max(LINE_Y_BASE_TOL, int(med_h * 0.65))
    sorted_t = sorted(tokens, key=lambda r: (mid_y(r[0]), left_x(r[0])))
    lines = []
    current = [sorted_t[0]]
    for item in sorted_t[1:]:
        if abs(mid_y(item[0]) - mid_y(current[0][0])) <= y_tol:
            current.append(item)
        else:
            lines.append(sorted(current, key=lambda r: left_x(r[0])))
            current = [item]
    lines.append(sorted(current, key=lambda r: left_x(r[0])))
    return lines


def line_text(line, sep=" "):
    parts = [collapse_spaced(clean_text(t)) for _, t, _ in line if t.strip()]
    return sep.join(p for p in parts if p)


def line_conf(line):
    return sum(c for _, _, c in line) / len(line) if line else 0.0


# ══════════════════════════════════════════════════════════════════════════════
#  FIELD EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def extract_email(lines):
    hits = []
    for line in lines:
        full = re.sub(r"\s+", "", line_text(line))
        m = _EMAIL_RE.search(full)
        if m:
            hits.append((line_conf(line), re.sub(r"\s+", "", m.group()).lower()))
        for _, t, c in line:
            t2 = re.sub(r"\s+", "", t)
            m2 = _EMAIL_RE.search(t2)
            if m2:
                hits.append((c, re.sub(r"\s+", "", m2.group()).lower()))
    return max(hits, key=lambda x: x[0])[1] if hits else ""


def extract_phones(lines):
    found = []
    for line in lines:
        txt = fix_digits(line_text(line))
        for m in _PHONE_RE.finditer(txt):
            raw = m.group()
            digits = re.sub(r"\D", "", raw)
            if PHONE_MIN_DIGITS <= len(digits) <= PHONE_MAX_DIGITS:
                norm = re.sub(r"(?<=\d) (?=\d)", "", raw)
                norm = re.sub(r"\s{2,}", " ", norm).strip()
                if norm and norm not in found:
                    found.append(norm)
    return found


def extract_website(lines):
    for line in lines:
        m = _WEB_RE.search(line_text(line))
        if m:
            url = m.group().strip()
            return ("http://" + url if not url.startswith("http") else url).lower()
    return ""


def extract_name(lines, raw_tokens):
    if not raw_tokens:
        return ""
    all_bottom_y = [top_y(b) + height_px(b) for b, _, _ in raw_tokens]
    card_h = max(all_bottom_y) if all_bottom_y else 1

    def _ok(bbox, text, conf):
        if conf < OCR_CONF_THRESHOLD:
            return False
        t = text.strip()
        if len(t) < 2:
            return False
        if re.search(r"\d", t) or "@" in t:
            return False
        if _COMPANY_KW.search(t) or _TITLE_KW.search(t) or _ADDR_KW.search(t):
            return False
        return sum(c.isalpha() for c in t) / max(len(t), 1) >= 0.60

    cands = [(b, t, c) for b, t, c in raw_tokens if _ok(b, t, c)]
    if not cands:
        return ""
    cand_lines = group_lines(cands)

    def _score(ln):
        avg_h = float(np.mean([height_px(b) for b, _, _ in ln]))
        avg_c = line_conf(ln)
        avg_y = float(np.mean([top_y(b) for b, _, _ in ln]))
        return avg_h * avg_c * (1 + 0.35 * (1.0 - avg_y / card_h))

    best = max(cand_lines, key=_score)
    parts = []
    for _, t, _ in sorted(best, key=lambda r: left_x(r[0])):
        p = collapse_spaced(clean_text(t))
        if p.isupper() and len(p) > 2:
            p = p.title()
        parts.append(p)
    name = " ".join(parts)
    return re.sub(r"^[^A-Za-z]+|[^A-Za-z.'\-]+$", "", name).strip()


def extract_company(lines):
    scored = []
    for line in lines:
        t = line_text(line).strip()
        if not t:
            continue
        score = 0
        if _COMPANY_KW.search(t):
            score += 10
        if t.isupper() and len(t.split()) >= 2:
            score += 3
        if len(t.split()) >= 2:
            score += 1
        if "@" in t or re.search(r"\d{4,}", t):
            score -= 10
        if _ADDR_KW.search(t):
            score -= 5
        if _TITLE_KW.search(t):
            score -= 4
        if score > 0:
            scored.append((score, line_conf(line), t))
    if not scored:
        return ""
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return scored[0][2]


def extract_job_title(lines):
    for line in lines:
        t = line_text(line).strip()
        if _TITLE_KW.search(t) and not _COMPANY_KW.search(t):
            return t
    return ""


def extract_address(lines):
    parts = []
    for line in lines:
        t = line_text(line).strip()
        if not t:
            continue
        is_addr = bool(_ADDR_KW.search(t))
        if re.search(r"\b\d{4,9}\b", t) and not re.search(r"[+\-]\s*\d{6,}", t):
            is_addr = True
        if re.search(r"[A-Z][a-z]+,\s*[A-Z][a-z]+", t):
            is_addr = True
        if is_addr:
            parts.append(t)
    return ", ".join(parts) if parts else ""


def extract_social(lines):
    r = {"linkedin": "", "twitter": ""}
    for line in lines:
        t = line_text(line)
        if not r["linkedin"]:
            m = _LINKEDIN_RE.search(t)
            if m:
                r["linkedin"] = m.group().strip()
        if not r["twitter"]:
            m = _TWITTER_RE.search(t)
            if m:
                r["twitter"] = m.group().strip()
    return r


# ══════════════════════════════════════════════════════════════════════════════
#  FULL PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
def process_card(image_path, debug=False):
    result = {
        "file": image_path,
        "name": "",
        "job_title": "",
        "company": "",
        "email": "",
        "phones": [],
        "website": "",
        "address": "",
        "linkedin": "",
        "twitter": "",
        "_debug": {},
    }
    img = resize_for_ocr(load_image(image_path))
    img, skew = deskew(img)
    proc, is_dark, mode = preprocess_image(img)
    if debug:
        result["_debug"].update({"skew": round(skew, 2), "dark": is_dark, "mode": mode})
        cv2.imwrite(
            os.path.join(
                os.path.dirname(os.path.abspath(image_path)),
                "DEBUG_preprocessed_v5.png",
            ),
            proc,
        )
    easy = run_easyocr(proc)
    tess = run_tesseract(proc)
    tokens = merge_ocr(easy, tess)
    kept = [(b, t, c) for b, t, c in tokens if c >= OCR_CONF_THRESHOLD]
    if debug:
        result["_debug"]["total"] = len(tokens)
        result["_debug"]["kept"] = len(kept)
        result["_debug"]["raw"] = [
            {"text": t, "conf": round(c, 3), "y": round(top_y(b), 1)}
            for b, t, c in sorted(tokens, key=lambda r: top_y(r[0]))
        ]
    lines = group_lines(kept)
    if debug:
        result["_debug"]["lines"] = [
            {"n": i + 1, "text": line_text(ln), "conf": round(line_conf(ln), 3)}
            for i, ln in enumerate(lines)
        ]
    result["email"] = extract_email(lines)
    result["phones"] = extract_phones(lines)
    result["website"] = extract_website(lines)
    result["name"] = extract_name(lines, kept)
    result["company"] = extract_company(lines)
    result["job_title"] = extract_job_title(lines)
    result["address"] = extract_address(lines)
    soc = extract_social(lines)
    result["linkedin"] = soc["linkedin"]
    result["twitter"] = soc["twitter"]
    result["file"] = os.path.basename(image_path)
    if debug:
        _annotate(img, tokens, image_path)
    return result


def _annotate(img, tokens, src):
    ann = img.copy()
    for bbox, text, conf in tokens:
        pts = np.array([[int(p[0]), int(p[1])] for p in bbox], np.int32)
        color = (0, 200, 0) if conf >= OCR_CONF_THRESHOLD else (0, 0, 200)
        cv2.polylines(ann, [pts], True, color, 2)
        label = f"{text[:24]}{'…' if len(text)>24 else ''} [{conf:.2f}]"
        pos = (max(pts[0][0], 0), max(pts[0][1] - 5, 12))
        (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.4, 1)
        cv2.rectangle(
            ann, (pos[0], pos[1] - lh - 2), (pos[0] + lw, pos[1] + 2), (20, 20, 20), -1
        )
        cv2.putText(
            ann, label, pos, cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1, cv2.LINE_AA
        )
    out = os.path.join(os.path.dirname(os.path.abspath(src)), "DEBUG_annotated_v5.png")
    cv2.imwrite(out, ann)
    print(f"  [DEBUG] Annotated → {out}")


# ══════════════════════════════════════════════════════════════════════════════
#  TERMINAL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
def print_result(r):
    w = 60
    print()
    print("═" * w)
    print(f"  FILE      : {r['file']}")
    print("─" * w)
    print(f"  NAME      : {r['name']      or '—'}")
    print(f"  JOB TITLE : {r['job_title'] or '—'}")
    print(f"  COMPANY   : {r['company']   or '—'}")
    print(f"  EMAIL     : {r['email']     or '—'}")
    for i, ph in enumerate(r["phones"], 1):
        print(f"  PHONE {'#'+str(i):<4}: {ph}")
    if not r["phones"]:
        print(f"  PHONE     : —")
    print(f"  WEBSITE   : {r['website']   or '—'}")
    print(f"  ADDRESS   : {r['address']   or '—'}")
    print(f"  LINKEDIN  : {r['linkedin']  or '—'}")
    print(f"  TWITTER/X : {r['twitter']   or '—'}")
    print("═" * w)
    if r.get("_debug"):
        d = r["_debug"]
        print(
            f"\n  [DEBUG] skew={d.get('skew')}°  dark={d.get('dark')}  mode={d.get('mode')}"
        )
        print(f"  [DEBUG] tokens total={d.get('total')}  kept={d.get('kept')}")
        print("\n  RAW TOKENS:")
        for tk in d.get("raw", []):
            mark = "  " if tk["conf"] >= OCR_CONF_THRESHOLD else "✗ "
            print(f"  {mark}y={tk['y']:>6.1f}  conf={tk['conf']:.3f}  {tk['text']!r}")
        print("\n  GROUPED LINES:")
        for ln in d.get("lines", []):
            print(f"  Line {ln['n']:02d} (conf={ln['conf']:.2f}) : {ln['text']!r}")
        print()


# ══════════════════════════════════════════════════════════════════════════════
#  SAVE — JSON
# ══════════════════════════════════════════════════════════════════════════════
def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def save_json(results, out_dir):
    path = os.path.join(out_dir, f"cards_{_ts()}.json")
    clean = [{k: v for k, v in r.items() if k != "_debug"} for r in results]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2, ensure_ascii=False)
    print(f"  [JSON] → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SAVE — EXCEL  (replaces CSV)
# ══════════════════════════════════════════════════════════════════════════════
def save_excel(results, out_dir):
    if not OPENPYXL_AVAILABLE:
        print(
            "  [WARN] openpyxl missing — skipping Excel export (pip install openpyxl)"
        )
        return

    path = os.path.join(out_dir, f"cards_{_ts()}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Visiting Cards"

    # ── palette ───────────────────────────────────────────────────────────────
    C_HEADER_BG = "1F4E79"  # deep navy
    C_HEADER_FG = "FFFFFF"  # white
    C_ROW_ODD = "EBF3FB"  # light blue tint
    C_ROW_EVEN = "FFFFFF"  # white
    C_BORDER = "BDD7EE"  # soft blue border
    C_ACCENT_FG = "1F4E79"  # navy text for data

    thin = Side(style="thin", color=C_BORDER)
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── columns ───────────────────────────────────────────────────────────────
    COLS = [
        ("File", "file", 28),
        ("Name", "name", 22),
        ("Job Title", "job_title", 30),
        ("Company", "company", 28),
        ("Email", "email", 32),
        ("Phone 1", "phone_1", 20),
        ("Phone 2", "phone_2", 20),
        ("Website", "website", 30),
        ("Address", "address", 40),
        ("LinkedIn", "linkedin", 35),
        ("Twitter/X", "twitter", 22),
    ]

    # ── header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=C_HEADER_BG)
    header_font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (label, _, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = bdr

    ws.row_dimensions[1].height = 22

    # ── data rows ─────────────────────────────────────────────────────────────
    data_font = Font(name="Arial", size=10, color=C_ACCENT_FG)
    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, r in enumerate(results, start=2):
        ph = r.get("phones", [])
        row_bg = C_ROW_ODD if row_idx % 2 == 0 else C_ROW_EVEN
        row_fill = PatternFill("solid", fgColor=row_bg)

        values = {
            "file": r.get("file", ""),
            "name": r.get("name", ""),
            "job_title": r.get("job_title", ""),
            "company": r.get("company", ""),
            "email": r.get("email", ""),
            "phone_1": ph[0] if ph else "",
            "phone_2": ph[1] if len(ph) > 1 else "",
            "website": r.get("website", ""),
            "address": r.get("address", ""),
            "linkedin": r.get("linkedin", ""),
            "twitter": r.get("twitter", ""),
        }

        for col_idx, (_, key, _) in enumerate(COLS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values[key])
            cell.fill = row_fill
            cell.font = data_font
            cell.alignment = data_align
            cell.border = bdr

        ws.row_dimensions[row_idx].height = 18

    # ── column widths ─────────────────────────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── freeze header + auto-filter ───────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

    wb.save(path)
    print(f"  [XLSX] → {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
_IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp", ".heic"}


def collect_paths(target):
    p = Path(target)
    if p.is_file():
        return [str(p)]
    if p.is_dir():
        return sorted(str(f) for f in p.iterdir() if f.suffix.lower() in _IMG_EXTS)
    return []


def gui_pick():
    try:
        from tkinter import Tk, filedialog

        root = Tk()
        root.withdraw()
        paths = filedialog.askopenfilenames(
            title="Select Visiting Card Image(s)",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif *.webp")],
        )
        root.destroy()
        return list(paths)
    except Exception:
        print("  [WARN] Tkinter not available. Pass the image path as a CLI arg.")
        return []


def main():
    ap = argparse.ArgumentParser(description="Visiting Card OCR Extractor v5.0")
    ap.add_argument(
        "target",
        nargs="?",
        default=None,
        help="Image file or folder (blank → GUI picker)",
    )
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--no-json", action="store_true")
    ap.add_argument("--no-xlsx", action="store_true")
    args = ap.parse_args()

    paths = collect_paths(args.target) if args.target else gui_pick()
    if not paths:
        print("  No images found. Exiting.")
        sys.exit(0 if not args.target else 1)

    print(f"\n  {len(paths)} image(s) to process.\n")
    results = []
    out_dir = os.path.dirname(os.path.abspath(paths[0]))

    for i, path in enumerate(paths, 1):
        print(f"  [{i}/{len(paths)}] {os.path.basename(path)}")
        try:
            r = process_card(path, debug=args.debug)
            print_result(r)
            results.append(r)
        except Exception as e:
            print(f"  [ERROR] {path}: {e}")
            if args.debug:
                import traceback

                traceback.print_exc()

    if results:
        if not args.no_json:
            save_json(results, out_dir)
        if not args.no_xlsx:
            save_excel(results, out_dir)

    print("\n  Done.")


if __name__ == "__main__":
    main()
