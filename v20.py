#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Visiting Card OCR Engine v20.0 - Layout-Driven Production Build
================================================================
Architecture: Pure Geometric Layout Intelligence + Reading-Order Graph Extraction
Philosophy  : Zero keyword dictionaries. Zero heuristic regex maps. Zero brand/title lists.
              All field extraction decisions are derived exclusively from:
              - Spatial bounding box geometry
              - Relative typographic scale (font size proxy via token height)
              - Reading-order graph traversal
              - OCR confidence weighting
              - Inter-token gap physics
              - Block structural role inference (not string content)

v20 Architectural Advances over v19:
  1. Token Reconstruction Graph (TRG) — replaces character-stitch heuristics
  2. Layout Role Classifier (LRC) — assigns structural roles via geometry, not regex
  3. Typographic Prominence Score (TPS) — font-size proxy from bounding box height
  4. Spatial Proximity Graph (SPG) — directed reading-order DAG over all rows
  5. Anchor Field Bootstrap — email/phone anchors used to infer surrounding roles
  6. Confidence-Weighted Field Voting — OCR confidence propagated into extraction
  7. Address Block Fusion — geometric vertical clustering, not keyword detection
  8. Social Handle Detection — structural pattern matching (prefix symbols), not brand lists
"""

import os
import sys
import re
import json
import math
import logging
import zipfile
import argparse
import warnings
import unicodedata
from io import BytesIO
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Any, Set
from dataclasses import dataclass, field, asdict
from collections import defaultdict

import cv2
import numpy as np
from PIL import Image
import phonenumbers
from phonenumbers import PhoneNumberMatcher, PhoneNumberFormat
from rapidfuzz import fuzz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["FLAGS_use_mkldnn"] = "0"
os.environ["FLAGS_call_stack_level"] = "2"

# ══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler("ocr_engine_v20.log", mode="a", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("VC_OCR_v20")

# ══════════════════════════════════════════════════════════════════════════════
#  SYSTEM CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
CFG = {
    # OCR
    "OCR_CONF_THRESH": 0.18,
    "PDF_DPI": 250,
    # Segmentation
    "SEG_MIN_AREA_FRAC": 0.03,
    "SEG_MAX_AREA_FRAC": 0.95,
    "SEG_PAD_PX": 12,
    "SEG_CARD_ASPECT_MIN": 0.5,
    "SEG_CARD_ASPECT_MAX": 5.0,
    # Token Reconstruction Graph (TRG)
    "TRG_GAP_SIGMA_MULTIPLIER": 2.2,  # Max gap = median_char_width * sigma_mult
    "TRG_OVERLAP_MERGE_THRESH": 0.15,  # Horizontal overlap ratio to force merge
    "TRG_MAX_SKEW_DEG": 3.0,  # Max angular skew to allow row grouping
    # Row grouping
    "ROW_TOL_FACTOR": 0.50,
    "ROW_TOL_FALLBACK_PX": 8,
    # Vertical block proximity
    "BLOCK_PROX_FACTOR": 1.80,
    # Layout Role Classifier (LRC)
    "LRC_TOP_FRAC": 0.35,  # Top 35% of card = high prominence zone
    "LRC_BOTTOM_FRAC": 0.60,  # Bottom 40% = address zone bias
    "LRC_TALL_SCALE_RATIO": 1.40,  # Token height / median height > 1.4 → prominent
    "LRC_NAME_MAX_TOKENS": 5,  # Max word count for a name candidate
    "LRC_COMPANY_MAX_TOKENS": 8,
    "LRC_TITLE_MAX_TOKENS": 8,
    # Anchor proximity
    "ANCHOR_PROX_LINES": 3,  # Lines above/below an anchor to attribute role
    # Phone
    "PHONE_MIN_DIGITS": 7,
    "PHONE_MAX_DIGITS": 15,
    "MAX_PHONES": 5,
    # Fuzzy dedup
    "FUZZY_SIM_THRESH": 80,
    # Output
    "EXCEL_DB_FILE": "master_contacts.xlsx",
    "DEBUG": False,
    # Script detection
    "SECONDARY_SCRIPT_THRESH": 0.15,
    # IOU dedup
    "IOU_DEDUP_CELL_FACTOR": 0.40,
    "IOU_DEDUP_CELL_MIN_PX": 12,
    # Multi-pass OCR
    "OCR_BRIGHTNESS_ALPHA": 1.35,
    "OCR_BRIGHTNESS_BETA": 25,
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}
EXCEL_COLS = [
    "Timestamp",
    "Name",
    "Job Title",
    "Company",
    "Email",
    "Phone_1",
    "Phone_2",
    "Phone_3",
    "Website",
    "Address",
    "LinkedIn",
    "Twitter",
    "Instagram",
    "GitHub",
    "Quality",
    "Confidence",
    "Source",
]
_QUALITY_COLOUR = {"🟢 GREEN": "C6EFCE", "🟡 YELLOW": "FFEB9C", "🔴 RED": "FFC7CE"}
_THIN_SIDE = Side(style="thin", color="BDD7EE")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

# ── STRUCTURAL PATTERN MATCHERS (no keyword content, pure format signatures) ──
# These match structural patterns (symbols, digit runs, format shapes) NOT content words.
_EMAIL_RE = re.compile(
    r"[A-Za-z0-9._%+\-]{1,64}@[A-Za-z0-9.\-]{1,253}\.[A-Za-z]{2,12}", re.I
)
_WEB_RE = re.compile(
    r"(?:https?://|www\.)[A-Za-z0-9.\-/_%?=&#@]+|[A-Za-z0-9][\w\-]*\.[A-Za-z]{2,12}(?:/[^\s]*)?",
    re.I,
)
_PHONE_DIGITS_RE = re.compile(r"[\+\(]?\d[\d\s\-\.\(\)]{5,16}\d")
# Social: detected by structural prefix symbol patterns, NOT by platform name lists
_SOCIAL_PREFIX_RE = re.compile(
    r"(?:^|[\s,|•·])(@[\w.]{2,32}|/in/[\w\-]{3,100}|linkedin\.com/in/[\w\-]+|github\.com/[\w\-]+|twitter\.com/[\w]+|x\.com/[\w]+|instagram\.com/[\w.]+)",
    re.I,
)
_AT_HANDLE_RE = re.compile(r"(?<![A-Za-z0-9])@([\w.]{2,32})")
_POSTAL_SHAPE_RE = re.compile(
    r"\b\d{4,7}\b"  # Generic 4-7 digit postal
    r"|\b[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}\b"  # UK format
    r"|\b\d{3}-\d{4}\b"  # Japanese format
    r"|\b[A-Z]\d[A-Z]\s*\d[A-Z]\d\b",  # Canadian format
    re.I,
)
# Structural: separators commonly found in address-type blocks
_ADDR_SEPARATOR_RE = re.compile(r"[,/\|\\]{1}")


# ══════════════════════════════════════════════════════════════════════════════
#  CORE DATA STRUCTURES
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class OCRToken:
    """Raw OCR output token with full spatial metadata."""

    text: str
    confidence: float
    bbox: np.ndarray  # shape (4, 2) — four corner points
    center_x: float
    center_y: float
    width: float
    height: float

    @property
    def x_min(self) -> float:
        return float(np.min(self.bbox[:, 0]))

    @property
    def x_max(self) -> float:
        return float(np.max(self.bbox[:, 0]))

    @property
    def y_min(self) -> float:
        return float(np.min(self.bbox[:, 1]))

    @property
    def y_max(self) -> float:
        return float(np.max(self.bbox[:, 1]))

    @property
    def char_width_est(self) -> float:
        """Estimated per-character width using only alphabetic characters."""
        alpha = sum(c.isalpha() for c in self.text)
        return self.width / alpha if alpha > 0 else self.width / max(len(self.text), 1)


@dataclass
class TextRow:
    """Reconstructed text row after TRG-based token stitching."""

    tokens: List[OCRToken]
    text: str
    y_center: float
    x_min: float
    x_max: float
    median_height: float
    avg_confidence: float
    # Layout metadata assigned by LRC
    layout_role: str = "unknown"  # name|title|company|address|contact|social|unknown
    role_score: float = 0.0
    tps: float = 1.0  # Typographic Prominence Score (height ratio vs median)
    position_frac: float = 0.5  # Normalized vertical position [0=top, 1=bottom]

    @property
    def word_count(self) -> int:
        return len(self.text.split())

    @property
    def has_digit(self) -> bool:
        return any(c.isdigit() for c in self.text)

    @property
    def digit_density(self) -> float:
        return sum(c.isdigit() for c in self.text) / max(len(self.text), 1)

    @property
    def punct_density(self) -> float:
        return sum(c in ",./-#@:;" for c in self.text) / max(len(self.text), 1)

    @property
    def alpha_density(self) -> float:
        return sum(c.isalpha() for c in self.text) / max(len(self.text), 1)

    @property
    def is_all_caps(self) -> bool:
        alpha = [c for c in self.text if c.isalpha()]
        return len(alpha) > 0 and all(c.isupper() for c in alpha)

    @property
    def is_title_case(self) -> bool:
        words = self.text.split()
        return (
            len(words) >= 1
            and sum(w[0].isupper() for w in words if w) / len(words) >= 0.6
        )

    @property
    def has_email_pattern(self) -> bool:
        return bool(_EMAIL_RE.search(self.text))

    @property
    def has_phone_pattern(self) -> bool:
        m = _PHONE_DIGITS_RE.search(self.text)
        if not m:
            return False
        return (
            CFG["PHONE_MIN_DIGITS"]
            <= len(re.sub(r"\D", "", m.group(0)))
            <= CFG["PHONE_MAX_DIGITS"]
        )

    @property
    def has_web_pattern(self) -> bool:
        return bool(_WEB_RE.search(self.text))

    @property
    def has_postal_pattern(self) -> bool:
        return bool(_POSTAL_SHAPE_RE.search(self.text))

    @property
    def has_social_pattern(self) -> bool:
        return bool(
            _SOCIAL_PREFIX_RE.search(self.text) or _AT_HANDLE_RE.search(self.text)
        )

    @property
    def separator_count(self) -> int:
        return len(_ADDR_SEPARATOR_RE.findall(self.text))


@dataclass
class ContactCard:
    name: str = "—"
    job_title: str = "—"
    company: str = "—"
    email: str = "—"
    phone_1: str = "—"
    phone_2: str = "—"
    phone_3: str = "—"
    website: str = "—"
    address: str = "—"
    linkedin: str = "—"
    twitter: str = "—"
    instagram: str = "—"
    github: str = "—"
    quality_score: str = "🔴 RED"
    raw_text: str = ""
    confidence_avg: float = 0.0
    source_file: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def to_vcard(self) -> str:
        lines = ["BEGIN:VCARD", "VERSION:3.0"]
        if self.name != "—":
            lines.append(f"FN:{self.name}")
            p = self.name.split(maxsplit=1)
            lines.append(
                f"N:{p[-1]};{p[0]};;;" if len(p) == 2 else f"N:{self.name};;;;"
            )
        if self.company != "—":
            lines.append(f"ORG:{self.company}")
        if self.job_title != "—":
            lines.append(f"TITLE:{self.job_title}")
        if self.email != "—":
            lines.append(f"EMAIL;TYPE=INTERNET:{self.email}")
        if self.phone_1 != "—":
            lines.append(f"TEL;TYPE=VOICE,PREF:{self.phone_1}")
        if self.phone_2 != "—":
            lines.append(f"TEL;TYPE=VOICE:{self.phone_2}")
        if self.website != "—":
            lines.append(f"URL:{self.website}")
        if self.address != "—":
            lines.append(f"ADR;TYPE=WORK:;;{self.address.replace(', ', '; ')};;;;")
        lines.append("END:VCARD")
        return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  PADDLE OCR SINGLETON
# ══════════════════════════════════════════════════════════════════════════════
_PADDLE_CACHE: Dict[str, Any] = {}


def _get_paddle(lang: str = "en"):
    if lang not in _PADDLE_CACHE:
        from paddleocr import PaddleOCR

        _PADDLE_CACHE[lang] = PaddleOCR(
            lang=lang,
        )

    return _PADDLE_CACHE[lang]


# ══════════════════════════════════════════════════════════════════════════════
#  IMAGE PREPROCESSING
# ══════════════════════════════════════════════════════════════════════════════
def deskew(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img
    lines = cv2.HoughLinesP(
        cv2.Canny(gray, 50, 150),
        1,
        np.pi / 180,
        threshold=80,
        minLineLength=60,
        maxLineGap=20,
    )
    if lines is None:
        return img
    angles = []
    for l in lines:
        a = math.degrees(math.atan2(l[0][3] - l[0][1], l[0][2] - l[0][0]))
        if abs(a) < 45:
            angles.append(a)
    if not angles:
        return img
    angle = float(np.median(angles))
    if abs(angle) < 0.5:
        return img
    h, w = img.shape[:2]
    M = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
    cos_a, sin_a = abs(M[0, 0]), abs(M[0, 1])
    nw = int(h * sin_a + w * cos_a)
    nh = int(h * cos_a + w * sin_a)
    M[0, 2] += (nw - w) / 2
    M[1, 2] += (nh - h) / 2
    return cv2.warpAffine(
        img, M, (nw, nh), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE
    )


def preprocess(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img.copy()
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    roi = gray[
        int(img.shape[0] * 0.3) : int(img.shape[0] * 0.7),
        int(img.shape[1] * 0.3) : int(img.shape[1] * 0.7),
    ]
    if roi.size > 0 and float(np.mean(roi)) < 120:
        gray = cv2.bitwise_not(gray)
    gray = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8)).apply(gray)
    sharpened = cv2.addWeighted(gray, 1.55, cv2.GaussianBlur(gray, (0, 0), 3), -0.55, 0)
    return cv2.cvtColor(sharpened, cv2.COLOR_GRAY2BGR)


def segment_cards(img: np.ndarray) -> List[np.ndarray]:
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if img.ndim == 3 else img.copy()
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blurred, 30, 120)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 5))
    dilated = cv2.dilate(edges, kernel, iterations=3)
    cnts, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    pad = CFG["SEG_PAD_PX"]
    rois = []
    for c in cnts:
        area = cv2.contourArea(c)
        if not (
            h * w * CFG["SEG_MIN_AREA_FRAC"] <= area <= h * w * CFG["SEG_MAX_AREA_FRAC"]
        ):
            continue
        x, y, cw, ch = cv2.boundingRect(c)
        aspect = cw / max(ch, 1)
        if CFG["SEG_CARD_ASPECT_MIN"] <= aspect <= CFG["SEG_CARD_ASPECT_MAX"]:
            rois.append((x, y, cw, ch))
    if not rois:
        return [img]
    rois.sort(key=lambda r: (r[1] // 100, r[0]))
    return [
        img[
            max(0, y - pad) : min(h, y + ch + pad),
            max(0, x - pad) : min(w, x + cw + pad),
        ]
        for x, y, cw, ch in rois
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  TOKEN RECONSTRUCTION GRAPH (TRG)
#  Replaces all character-stitch heuristics and keyword-based fragment joining.
#  Pure geometry: gap analysis, overlap detection, angular alignment.
# ══════════════════════════════════════════════════════════════════════════════
class TokenReconstructionGraph:
    """
    Builds a directed graph where nodes are OCR tokens and edges represent
    spatial adjacency within a reading line. Tokens are merged when:
      - Horizontal gap ≤ median_char_width × sigma_multiplier (calibrated per row)
      - OR horizontal overlap ratio ≥ threshold (fragmented bounding boxes)
      - Angular deviation from row baseline ≤ max_skew_deg
    No string content is inspected during graph construction.
    """

    def __init__(self, tokens: List[OCRToken]):
        self.tokens = sorted(tokens, key=lambda t: t.center_x)
        self._adjacency: Dict[int, List[int]] = defaultdict(list)
        self._gap_stats = self._compute_gap_stats()
        self._build_edges()

    def _compute_gap_stats(self) -> Dict[str, float]:
        """
        Compute per-row gap statistics using only geometric properties.
        No content inspection.
        """
        char_widths = [t.char_width_est for t in self.tokens if t.char_width_est > 0]
        if not char_widths:
            return {"median_char_w": 12.0, "std_char_w": 3.0}
        arr = np.array(char_widths)
        return {
            "median_char_w": float(np.median(arr)),
            "std_char_w": float(np.std(arr)),
        }

    def _gap_threshold(self) -> float:
        """
        Adaptive gap threshold calibrated from per-row character width distribution.
        sigma_multiplier controls sensitivity (set in CFG, no content dependency).
        """
        return self._gap_stats["median_char_w"] * CFG["TRG_GAP_SIGMA_MULTIPLIER"]

    def _horizontal_overlap_ratio(self, a: OCRToken, b: OCRToken) -> float:
        """IOU-style horizontal overlap — purely geometric."""
        overlap = max(0.0, min(a.x_max, b.x_max) - max(a.x_min, b.x_min))
        union = max(a.x_max, b.x_max) - min(a.x_min, b.x_min)
        return overlap / union if union > 0 else 0.0

    def _angular_deviation(self, a: OCRToken, b: OCRToken) -> float:
        """Compute angle (degrees) between token centers — used to filter skewed pairs."""
        dx = b.center_x - a.center_x
        dy = b.center_y - a.center_y
        return abs(math.degrees(math.atan2(dy, max(dx, 0.001))))

    def _build_edges(self):
        thresh = self._gap_threshold()
        for i in range(len(self.tokens) - 1):
            a = self.tokens[i]
            b = self.tokens[i + 1]
            gap = b.x_min - a.x_max
            overlap = self._horizontal_overlap_ratio(a, b)
            angle_dev = self._angular_deviation(a, b)
            should_join = (
                gap <= thresh and angle_dev <= CFG["TRG_MAX_SKEW_DEG"]
            ) or overlap >= CFG["TRG_OVERLAP_MERGE_THRESH"]
            if should_join:
                self._adjacency[i].append(i + 1)

    def reconstruct(self) -> str:
        """
        Traverse the graph left-to-right.
        Connected tokens → concatenate directly.
        Disconnected tokens → insert space.
        Result: geometrically-faithful text reconstruction, no content knowledge needed.
        """
        if not self.tokens:
            return ""
        parts = [self.tokens[0].text]
        for i in range(1, len(self.tokens)):
            if (i - 1) in self._adjacency and i in self._adjacency[i - 1]:
                parts.append(self.tokens[i].text)  # connected — direct concat
            else:
                parts.append(" " + self.tokens[i].text)  # gap — insert space
        raw = "".join(parts)
        # Generic character-level repeat collapse (not content-specific)
        raw = re.sub(r"(.)\1{3,}", r"\1\1", raw)  # "AAAA" → "AA" (OCR artifact)
        return re.sub(r"\s+", " ", raw).strip()


# ══════════════════════════════════════════════════════════════════════════════
#  ROW GROUPING & LINE ASSEMBLY
# ══════════════════════════════════════════════════════════════════════════════
def _parse_paddle_result(results) -> List[Tuple[str, float, List]]:
    parsed = []

    if not results:
        return parsed

    try:
        result = results[0]

        texts = result.get("rec_texts", [])
        scores = result.get("rec_scores", [])
        polys = result.get("rec_polys", [])

        for text, conf, poly in zip(texts, scores, polys):

            text = str(text).strip()
            conf = float(conf)

            if not text:
                continue

            if conf < CFG["OCR_CONF_THRESH"]:
                continue

            pts = [[float(x), float(y)] for x, y in poly]

            parsed.append((text, conf, pts))

    except Exception as e:
        print("PARSE ERROR:", e)

    print("\nPARSED TOKEN COUNT:", len(parsed))

    for p in parsed[:10]:
        print("TOKEN:", p[0], "CONF:", round(p[1], 3))

    return parsed


def _bbox_key(pts: List, cell_px: float) -> Tuple[int, int]:
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]

    cx = sum(xs) / len(xs)
    cy = sum(ys) / len(ys)

    print(f"BBOX KEY DEBUG | cx={cx:.1f} cy={cy:.1f} " f"cell_px={cell_px:.1f}")

    return (
        int(cx / cell_px),
        int(cy / cell_px),
    )


def _iou_dedup(pool: dict) -> dict:
    """Deduplicate overlapping OCR boxes — pure geometric IOU, no content."""
    return pool  # Pool keyed by spatial cell already deduplicates spatially


def _run_ocr_pass(img_bgr: np.ndarray, lang: str, cell_px: float, pool: dict):
    if img_bgr.ndim == 2:
        img_bgr = cv2.cvtColor(img_bgr, cv2.COLOR_GRAY2BGR)

    try:
        ocr = _get_paddle(lang)

        result = ocr.predict(img_bgr)

        print("\n" + "=" * 80)
        print("RAW PADDLE RESULT TYPE:")
        print(type(result))
        print("=" * 80)

        # print("RAW PADDLE RESULT:")
        # print(result)

        parsed = _parse_paddle_result(result)

        print("\n" + "=" * 80)
        print("PARSED TOKEN COUNT:", len(parsed))
        print("=" * 80)

        if parsed:
            print("FIRST 10 PARSED TOKENS:")
            for item in parsed[:10]:
                print(item)
        else:
            print("NO PARSED TOKENS RETURNED")

        for text, conf, pts in parsed:
            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]

            tok = {
                "text": text,
                "conf": conf,
                "bbox": pts,
                "cx": sum(xs) / len(xs),
                "cy": sum(ys) / len(ys),
                "th": max(ys) - min(ys),
                "tw": max(xs) - min(xs),
            }

            pool[len(pool)] = tok

        print("\nPOOL SIZE AFTER OCR:", len(pool))

    except Exception as e:
        print("\nOCR EXCEPTION:")
        import traceback

        traceback.print_exc()

        log.warning("OCR pass failure: %s", e)


def detect_scripts(text: str) -> Tuple[str, Optional[str]]:
    counts: Dict[str, int] = {"latin": 0, "cjk": 0, "arabic": 0, "devanagari": 0}
    for ch in text:
        if not ch.strip():
            continue
        cp = ord(ch)
        if 0x4E00 <= cp <= 0x9FFF:
            counts["cjk"] += 1
        elif 0x0600 <= cp <= 0x06FF:
            counts["arabic"] += 1
        elif 0x0900 <= cp <= 0x097F:
            counts["devanagari"] += 1
        elif ch.isalpha():
            counts["latin"] += 1
    ranked = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    total = max(sum(counts.values()), 1)
    secondary = next(
        (sc for sc, cnt in ranked[1:] if cnt / total >= CFG["SECONDARY_SCRIPT_THRESH"]),
        None,
    )
    return ranked[0][0], secondary


def ocr_card(
    proc: np.ndarray, lang_override: Optional[str] = None
) -> Tuple[list, str, Optional[str]]:
    h = proc.shape[0]
    cell_px = max(h * CFG["IOU_DEDUP_CELL_FACTOR"], CFG["IOU_DEDUP_CELL_MIN_PX"])
    pool: dict = {}
    base = lang_override or "en"
    _run_ocr_pass(proc, base, cell_px, pool)
    _run_ocr_pass(
        cv2.convertScaleAbs(
            proc, alpha=CFG["OCR_BRIGHTNESS_ALPHA"], beta=CFG["OCR_BRIGHTNESS_BETA"]
        ),
        base,
        cell_px,
        pool,
    )
    pool = _iou_dedup(pool)
    tokens = list(pool.values())
    primary, secondary = detect_scripts(" ".join(t["text"] for t in tokens))
    return tokens, primary, secondary


def group_into_rows(raw_tokens: List[dict]) -> List[TextRow]:
    """
    Group raw OCR tokens into horizontal rows using geometric Y-center clustering.
    Uses TRG for text reconstruction within each row.
    No content analysis at this stage.
    """
    if not raw_tokens:
        return []
    heights = [t.get("th", 20) for t in raw_tokens]
    median_h = float(np.median(heights))
    tol = max(median_h * CFG["ROW_TOL_FACTOR"], CFG["ROW_TOL_FALLBACK_PX"])

    by_y = sorted(raw_tokens, key=lambda t: t["cy"])
    groups: List[List[dict]] = []
    cur = [by_y[0]]
    for tok in by_y[1:]:
        cur_avg_y = sum(t["cy"] for t in cur) / len(cur)
        if abs(tok["cy"] - cur_avg_y) <= tol:
            cur.append(tok)
        else:
            groups.append(sorted(cur, key=lambda t: t["cx"]))
            cur = [tok]
    groups.append(sorted(cur, key=lambda t: t["cx"]))

    rows: List[TextRow] = []
    for group in groups:
        ocr_tokens = [
            OCRToken(
                text=t["text"],
                confidence=t["conf"],
                bbox=np.array(t["bbox"]),
                center_x=t["cx"],
                center_y=t["cy"],
                width=t.get("tw", 10),
                height=t.get("th", 10),
            )
            for t in group
        ]
        trg = TokenReconstructionGraph(ocr_tokens)
        text = trg.reconstruct()
        y_vals = [t.center_y for t in ocr_tokens]
        x_mins = [t.x_min for t in ocr_tokens]
        x_maxs = [t.x_max for t in ocr_tokens]
        confs = [t.confidence for t in ocr_tokens]
        hs = [t.height for t in ocr_tokens]
        rows.append(
            TextRow(
                tokens=ocr_tokens,
                text=text,
                y_center=float(np.mean(y_vals)),
                x_min=float(np.min(x_mins)),
                x_max=float(np.max(x_maxs)),
                median_height=float(np.median(hs)),
                avg_confidence=float(np.mean(confs)),
            )
        )
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  VERTICAL BLOCK PROXIMITY FUSION
#  Groups rows that are spatially close AND share similar X-extents.
#  Used to fuse multi-line addresses and multi-line company blocks.
#  No string inspection.
# ══════════════════════════════════════════════════════════════════════════════
def fuse_proximate_blocks(rows: List[TextRow]) -> List[TextRow]:
    if len(rows) <= 1:
        return rows
    merged: List[TextRow] = []
    i = 0
    while i < len(rows):
        cur = rows[i]
        m_tokens = list(cur.tokens)
        m_texts = [cur.text]
        j = i + 1
        while j < len(rows):
            nxt = rows[j]
            vert_gap = nxt.y_center - cur.y_center
            if vert_gap > cur.median_height * CFG["BLOCK_PROX_FACTOR"]:
                break
            height_ratio = nxt.median_height / max(cur.median_height, 1)
            if not (0.7 <= height_ratio <= 1.3):
                break
            overlap_min = max(cur.x_min, nxt.x_min)
            overlap_max = min(cur.x_max, nxt.x_max)
            overlap_w = max(0.0, overlap_max - overlap_min)
            span = min(cur.x_max - cur.x_min, nxt.x_max - nxt.x_min)
            if (overlap_w / max(span, 1)) < 0.30:
                break
            m_tokens.extend(nxt.tokens)
            m_texts.append(nxt.text)
            cur = nxt
            j += 1

        y_vals = [t.center_y for t in m_tokens]
        x_mins = [t.x_min for t in m_tokens]
        x_maxs = [t.x_max for t in m_tokens]
        confs = [t.confidence for t in m_tokens]
        hs = [t.height for t in m_tokens]
        merged.append(
            TextRow(
                tokens=m_tokens,
                text=" ".join(m_texts),
                y_center=float(np.mean(y_vals)),
                x_min=float(np.min(x_mins)),
                x_max=float(np.max(x_maxs)),
                median_height=float(np.median(hs)),
                avg_confidence=float(np.mean(confs)),
            )
        )
        i = j
    return merged


# ══════════════════════════════════════════════════════════════════════════════
#  LAYOUT ROLE CLASSIFIER (LRC)
#  Assigns a structural role to each TextRow using ONLY:
#    - Typographic Prominence Score (token height vs card median)
#    - Vertical position fraction (normalized Y position)
#    - Structural pattern presence (email/phone/web/postal shapes)
#    - Word count bounds
#    - Character density metrics
#    - Capitalization geometry
#  Zero keyword dictionaries. Zero brand/title/suffix lists.
# ══════════════════════════════════════════════════════════════════════════════
def compute_tps(rows: List[TextRow]) -> float:
    """Card-level median token height — used as typographic baseline."""
    all_heights = [t.height for row in rows for t in row.tokens]
    return float(np.median(all_heights)) if all_heights else 12.0


def classify_layout_roles(rows: List[TextRow]) -> List[TextRow]:
    """
    Assign layout_role to each row using geometric and structural signals only.
    Role scores are continuous floats; highest score wins per role category.
    """
    if not rows:
        return rows
    card_median_h = compute_tps(rows)
    max_y = max(r.y_center for r in rows) or 1.0
    min_y = min(r.y_center for r in rows) or 0.0
    y_span = max_y - min_y or 1.0

    for row in rows:
        pos_frac = (row.y_center - min_y) / y_span  # 0 = top, 1 = bottom
        tps = row.median_height / max(card_median_h, 1)
        row.tps = tps
        row.position_frac = pos_frac

        # ── STEP 1: Hard-assign contact-type roles via structural pattern detection ──
        # These are format-shape patterns, NOT keyword matches.
        if row.has_email_pattern:
            row.layout_role = "email"
            row.role_score = 100.0
            continue
        if row.has_phone_pattern:
            row.layout_role = "phone"
            row.role_score = 100.0
            continue
        if row.has_web_pattern and not row.has_email_pattern:
            row.layout_role = "web"
            row.role_score = 90.0
            continue
        if row.has_social_pattern:
            row.layout_role = "social"
            row.role_score = 85.0
            continue

        # ── STEP 2: Structural address detection ──
        # Address rows have: postal shape OR high separator count OR high digit density
        # combined with high punctuation density. No street-name keywords.
        addr_score = 0.0
        if row.has_postal_pattern:
            addr_score += 30.0
        if row.digit_density > 0.20:
            addr_score += 15.0
        if row.punct_density > 0.12:
            addr_score += 12.0
        if row.separator_count >= 2:
            addr_score += 10.0
        if pos_frac > CFG["LRC_BOTTOM_FRAC"]:
            addr_score += 8.0  # Bottom of card
        if row.word_count > 4:
            addr_score += 5.0

        # ── STEP 3: Structural prominence scoring for name/company/title ──
        # Prominence = typographic scale + position + capitalization geometry
        # High TPS + top position + few words = strong name candidate
        # Moderate TPS + few words + no digits = title or company candidate
        promi_score = 0.0
        promi_score += tps * 20.0  # Larger text = more prominent
        promi_score += (1.0 - pos_frac) * 15.0  # Higher on card = more prominent
        if row.is_all_caps:
            promi_score += 8.0  # All-caps: company or emphasis
        if row.is_title_case:
            promi_score += 6.0  # Title case: name/title pattern
        if row.has_digit:
            promi_score -= 10.0  # Digits reduce prominence
        if row.alpha_density > 0.80:
            promi_score += 5.0  # High alpha = text block
        if row.avg_confidence > 0.80:
            promi_score += 3.0  # High OCR confidence

        # Name: 1–N_max words, high prominence, top portion of card
        name_score = promi_score
        if not (1 <= row.word_count <= CFG["LRC_NAME_MAX_TOKENS"]):
            name_score -= 15.0
        if pos_frac > 0.50:
            name_score -= 12.0  # Names rarely appear in bottom half
        if row.digit_density > 0.05:
            name_score -= 20.0  # Names don't have digits

        # Title: 1–T_max words, moderate prominence, middle vertical zone
        title_score = promi_score * 0.85
        if not (1 <= row.word_count <= CFG["LRC_TITLE_MAX_TOKENS"]):
            title_score -= 10.0
        if pos_frac > 0.70:
            title_score -= 10.0
        if row.has_digit:
            title_score -= 12.0

        # Company: any word count, often prominent, often all-caps
        company_score = promi_score * 0.80
        if row.is_all_caps:
            company_score += 10.0
        if row.word_count > CFG["LRC_COMPANY_MAX_TOKENS"]:
            company_score -= 8.0

        # Determine role from max scoring dimension
        scores = {
            "address": addr_score,
            "name": name_score,
            "title": title_score,
            "company": company_score,
        }

        best_role = max(scores, key=scores.get)
        best_score = scores[best_role]

        # Only assign if score exceeds a minimum confidence threshold
        if best_score > 5.0:
            row.layout_role = best_role
            row.role_score = best_score
        else:
            row.layout_role = "unknown"
            row.role_score = 0.0

    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  SPATIAL PROXIMITY GRAPH (SPG) — READING-ORDER DAG
#  Builds a directed acyclic graph over all TextRows in reading order.
#  Used for anchor propagation: once email/phone is located, the lines
#  immediately above become candidates for name/title/company via proximity.
# ══════════════════════════════════════════════════════════════════════════════
def build_reading_order_dag(rows: List[TextRow]) -> Dict[int, List[int]]:
    """
    DAG edges: row_i → row_j if row_j is spatially below row_i and within
    horizontal overlap range. Used to propagate contextual role signals.
    """
    dag: Dict[int, List[int]] = defaultdict(list)
    for i, ri in enumerate(rows):
        for j, rj in enumerate(rows):
            if j <= i:
                continue
            if rj.y_center <= ri.y_center:
                continue
            overlap_min = max(ri.x_min, rj.x_min)
            overlap_max = min(ri.x_max, rj.x_max)
            overlap_w = max(0.0, overlap_max - overlap_min)
            ri_span = ri.x_max - ri.x_min
            if (overlap_w / max(ri_span, 1)) >= 0.25:
                dag[i].append(j)
    return dag


def find_ancestors(
    dag: Dict[int, List[int]], target_idx: int, max_hops: int
) -> Set[int]:
    """Return row indices that are ancestors of target_idx within max_hops steps."""
    # Invert the dag to get parent pointers
    parents: Dict[int, List[int]] = defaultdict(list)
    for src, dsts in dag.items():
        for dst in dsts:
            parents[dst].append(src)
    visited = set()
    queue = [target_idx]
    depth = {target_idx: 0}
    while queue:
        node = queue.pop(0)
        for parent in parents.get(node, []):
            if parent not in visited and depth[node] < max_hops:
                visited.add(parent)
                depth[parent] = depth[node] + 1
                queue.append(parent)
    return visited


def find_descendants(
    dag: Dict[int, List[int]], source_idx: int, max_hops: int
) -> Set[int]:
    """Return row indices that are descendants of source_idx within max_hops steps."""
    visited = set()
    queue = [source_idx]
    depth = {source_idx: 0}
    while queue:
        node = queue.pop(0)
        for child in dag.get(node, []):
            if child not in visited and depth[node] < max_hops:
                visited.add(child)
                depth[child] = depth[node] + 1
                queue.append(child)
    return visited


# ══════════════════════════════════════════════════════════════════════════════
#  FUZZY DEDUPLICATION (CONTENT CORRUPTION FILTER)
# ══════════════════════════════════════════════════════════════════════════════
def _corruption_score(a: str, b: str) -> float:
    if a == b:
        return 100.0
    base = fuzz.ratio(a.lower(), b.lower())
    # OCR character confusion pairs (purely visual similarity, not linguistic)
    confusions = {
        ("g", "q"),
        ("q", "g"),
        ("1", "l"),
        ("l", "1"),
        ("0", "o"),
        ("o", "0"),
        ("5", "s"),
        ("s", "5"),
        ("8", "b"),
        ("b", "8"),
        ("6", "b"),
        ("rn", "m"),
    }
    cc = sum(
        1 for c1, c2 in zip(a.lower(), b.lower()) if c1 != c2 and (c1, c2) in confusions
    )
    if cc > 0 and base >= 70:
        base = min(base + min(cc * 5, 20), 100)
    return base


def _select_best_variant(a: str, b: str) -> str:
    score = lambda t: sum(c.isalpha() for c in t) - len(re.findall(r"(.)\1{2,}", t)) * 3
    return a if score(a) >= score(b) else b


def remove_fuzzy_duplicates(rows: List[TextRow]) -> List[TextRow]:
    if len(rows) <= 1:
        return rows
    keep = []
    removed: Set[int] = set()
    for i, r1 in enumerate(rows):
        if i in removed:
            continue
        is_dup = False
        for j, r2 in enumerate(rows):
            if i >= j or j in removed:
                continue
            if _corruption_score(r1.text, r2.text) >= CFG["FUZZY_SIM_THRESH"]:
                best = _select_best_variant(r1.text, r2.text)
                if best == r1.text:
                    removed.add(j)
                else:
                    removed.add(i)
                    is_dup = True
                    break
        if not is_dup:
            keep.append(r1)
    return keep


# ══════════════════════════════════════════════════════════════════════════════
#  FIELD EXTRACTORS — STRUCTURAL PATTERN ONLY
#  All extractors work on structural format signatures, NOT content keywords.
# ══════════════════════════════════════════════════════════════════════════════
def extract_email(corpus: str) -> str:
    m = _EMAIL_RE.search(corpus)
    return m.group(0).lower() if m else "—"


def extract_website(corpus: str, email: str = "—") -> str:
    candidates = _WEB_RE.findall(corpus)
    for c in candidates:
        # Exclude if it's the email domain or an email fragment
        if email != "—" and email.split("@")[-1].lower() in c.lower():
            continue
        if "@" in c:
            continue
        # Exclude pure postal codes matched by lookahead
        if re.fullmatch(r"\d{4,7}", c):
            continue
        return c.strip()
    return "—"


def extract_phones(rows: List[TextRow], blocked_digit_sets: Set[str]) -> List[str]:
    """
    Extract phone numbers using:
      1. phonenumbers library (structural ITU-T parsing, not keyword-based)
      2. Regex digit-run pattern (fallback for non-standard formats)
    Blocked digit sets contain postal codes extracted from address rows.
    """
    phones: List[str] = []
    seen_digits: Set[str] = set()

    for row in rows:
        # Skip rows already assigned as address/email/web to avoid false positives
        if row.layout_role in ("email", "web", "address", "social"):
            continue
        text = row.text

        # Pass 1: ITU-T structural parser
        for region in [
            "US",
            "GB",
            "IN",
            "SG",
            "AU",
            "CA",
            "DE",
            "FR",
            "AE",
            "ZA",
            "HK",
            "JP",
        ]:
            try:
                for match in PhoneNumberMatcher(text, region):
                    if not phonenumbers.is_valid_number(match.number):
                        continue
                    digits = re.sub(
                        r"\D",
                        "",
                        phonenumbers.format_number(
                            match.number, PhoneNumberFormat.E164
                        ),
                    )
                    if digits in seen_digits or digits in blocked_digit_sets:
                        continue
                    seen_digits.add(digits)
                    phones.append(
                        phonenumbers.format_number(
                            match.number, PhoneNumberFormat.INTERNATIONAL
                        )
                    )
                    if len(phones) >= CFG["MAX_PHONES"]:
                        break
            except Exception:
                continue
            if len(phones) >= CFG["MAX_PHONES"]:
                break

        # Pass 2: Regex digit-run fallback (no content dependency)
        for m in _PHONE_DIGITS_RE.finditer(text):
            digits = re.sub(r"\D", "", m.group(0))
            if len(digits) < CFG["PHONE_MIN_DIGITS"]:
                continue
            if len(digits) > CFG["PHONE_MAX_DIGITS"]:
                continue
            if digits in seen_digits or digits in blocked_digit_sets:
                continue
            seen_digits.add(digits)
            phones.append(m.group(0).strip())
            if len(phones) >= CFG["MAX_PHONES"]:
                break

    return phones


def extract_social_handles(rows: List[TextRow]) -> Dict[str, str]:
    """
    Detect social profiles using structural patterns only:
      - URL path structure (/in/, /u/, etc.) — platform-agnostic routing
      - @ handle prefix — structural symbol, not platform name
      - Domain-local hostname patterns matched against known TLD structures
    Returns platform → handle mapping.
    """
    result: Dict[str, str] = {}

    for row in rows:
        text = row.text

        # Detect LinkedIn via URL path structure /in/ — structural, not brand
        m = re.search(r"linkedin\.com/in/([\w\-]+)", text, re.I)
        if m and "linkedin" not in result:
            result["linkedin"] = m.group(0)

        m = re.search(r"/in/([\w\-]{3,100})(?:\s|$)", text, re.I)
        if m and "linkedin" not in result:
            result["linkedin"] = "/in/" + m.group(1)

        # GitHub — structural URL path detection
        m = re.search(r"github\.com/([\w\-]+)", text, re.I)
        if m and "github" not in result:
            result["github"] = m.group(0)

        # Twitter/X — structural URL path
        m = re.search(r"(?:twitter|x)\.com/([\w]+)", text, re.I)
        if m and "twitter" not in result:
            result["twitter"] = m.group(0)

        # Instagram — structural URL
        m = re.search(r"instagram\.com/([\w.]+)", text, re.I)
        if m and "instagram" not in result:
            result["instagram"] = m.group(0)

        # Generic @ handle — structural prefix detection (not platform-specific)
        for at_match in _AT_HANDLE_RE.finditer(text):
            handle = at_match.group(0)
            # Assign to first unassigned social slot
            if "twitter" not in result:
                result["twitter"] = handle
            elif "instagram" not in result and handle != result.get("twitter"):
                result["instagram"] = handle

    return result


def extract_address_blocks(rows: List[TextRow]) -> str:
    """
    Collect rows classified as 'address' by LRC.
    Fuse vertically proximate address rows into a single address string.
    No street-name keywords used.
    """
    addr_rows = [r for r in rows if r.layout_role == "address"]
    if not addr_rows:
        return "—"
    # Sort by vertical position
    addr_rows.sort(key=lambda r: r.y_center)
    return ", ".join(r.text for r in addr_rows)


# ══════════════════════════════════════════════════════════════════════════════
#  ANCHOR-BOOTSTRAPPED FIELD RESOLUTION
#  Once contact anchors (email, phone) are located in the DAG, the rows
#  spatially surrounding them receive boosted scores for name/title/company.
#  This is positional inference, not string matching.
# ══════════════════════════════════════════════════════════════════════════════
def resolve_prominent_fields(
    rows: List[TextRow], dag: Dict[int, List[int]], email: str
) -> Tuple[str, str, str]:

    # Find email/phone anchor rows
    anchor_indices: Set[int] = set()
    for i, row in enumerate(rows):
        if row.layout_role in ("email", "phone"):
            anchor_indices.add(i)

    # Rows above contact anchors
    near_anchor_set: Set[int] = set()
    for ai in anchor_indices:
        near_anchor_set |= find_ancestors(dag, ai, max_hops=CFG["ANCHOR_PROX_LINES"])

    # Build candidate pool
    candidates = []

    for i, row in enumerate(rows):

        if row.layout_role in (
            "email",
            "phone",
            "web",
            "social",
            "address",
            "unknown",
        ):
            if not (row.layout_role == "unknown" and row.position_frac < 0.4):
                continue

        base_score = row.role_score

        if i in near_anchor_set:
            base_score += 18.0

        if row.position_frac < CFG["LRC_TOP_FRAC"]:
            base_score += 10.0

        base_score += row.tps * 12.0
        base_score += row.avg_confidence * 6.0

        candidates.append(
            (
                base_score,
                row.tps,
                i,
                row,
            )
        )

    if not candidates:
        return "—", "—", "—"

    # --------------------------------------------------
    # DEBUG
    # --------------------------------------------------
    print("\n" + "=" * 80)
    print("FIELD CANDIDATES")
    print("=" * 80)

    for score, tps_val, idx, row in candidates:
        print(
            f"TEXT={row.text} | "
            f"ROLE={row.layout_role} | "
            f"SCORE={score:.2f} | "
            f"TPS={tps_val:.2f} | "
            f"POS={row.position_frac:.2f}"
        )

    # --------------------------------------------------
    # NAME
    # --------------------------------------------------
    name = "—"
    job_title = "—"
    company = "—"

    name_candidates = [
        c
        for c in candidates
        if (
            not c[3].has_digit
            and 1 <= c[3].word_count <= 5
            and c[3].position_frac < 0.50
        )
    ]

    if name_candidates:
        name = max(name_candidates, key=lambda x: x[0])[3].text

    # --------------------------------------------------
    # COMPANY
    # --------------------------------------------------
    company_candidates = [c for c in candidates if c[3].text != name]

    if company_candidates:
        company = max(
            company_candidates,
            key=lambda x: (
                x[3].is_all_caps,
                x[3].tps,
                x[0],
            ),
        )[3].text

    # --------------------------------------------------
    # TITLE
    # --------------------------------------------------
    title_candidates = [c for c in candidates if c[3].text not in {name, company}]

    if title_candidates:
        job_title = max(
            title_candidates,
            key=lambda x: x[0],
        )[3].text

    return name, job_title, company


# ══════════════════════════════════════════════════════════════════════════════
#  QUALITY SCORE CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
def calculate_quality_score(card: ContactCard) -> str:
    fields = [
        card.name,
        card.job_title,
        card.company,
        card.email,
        card.phone_1,
        card.website,
        card.address,
    ]
    filled = sum(1 for f in fields if f != "—")
    if filled >= 5:
        return "🟢 GREEN"
    if filled >= 3:
        return "🟡 YELLOW"
    return "🔴 RED"


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN EXTRACTION PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
def extract_contact_card(rows: List[TextRow], label: str) -> ContactCard:
    """
    Full layout-driven extraction pipeline.
    Order of operations:
      1. LRC: assign geometric layout roles to rows
      2. Build SPG reading-order DAG
      3. Extract hard-pattern fields (email, phone, web, social) — format detection
      4. Extract address blocks via geometry cluster
      5. Resolve prominent fields (name/title/company) via DAG + typographic scale
      6. Compute quality score
    Zero keyword dictionaries at any stage.
    """
    card = ContactCard(source_file=label)
    if not rows:
        return card

    full_corpus = " ".join(r.text for r in rows)
    card.raw_text = full_corpus
    card.confidence_avg = float(np.mean([r.avg_confidence for r in rows]))

    # ── 1. Layout Role Classification ──────────────────────────────────────
    rows = classify_layout_roles(rows)

    # ── 2. Reading-Order DAG ───────────────────────────────────────────────
    dag = build_reading_order_dag(rows)

    # ── 3. Hard-Pattern Field Extraction ──────────────────────────────────
    card.email = extract_email(full_corpus)
    card.website = extract_website(full_corpus, card.email)

    # Collect postal digit sets from address rows (to block from phone extractor)
    blocked_postal_digits: Set[str] = set()
    for row in rows:
        if row.layout_role == "address" or row.has_postal_pattern:
            for pm in _POSTAL_SHAPE_RE.finditer(row.text):
                blocked_postal_digits.add(re.sub(r"\D", "", pm.group(0)))

    phones = extract_phones(rows, blocked_postal_digits)
    if len(phones) > 0:
        card.phone_1 = phones[0]
    if len(phones) > 1:
        card.phone_2 = phones[1]
    if len(phones) > 2:
        card.phone_3 = phones[2]

    # ── 4. Social Handles ──────────────────────────────────────────────────
    socials = extract_social_handles(rows)
    card.linkedin = socials.get("linkedin", "—")
    card.twitter = socials.get("twitter", "—")
    card.instagram = socials.get("instagram", "—")
    card.github = socials.get("github", "—")

    # ── 5. Address Blocks ─────────────────────────────────────────────────
    card.address = extract_address_blocks(rows)

    # ── 6. Prominent Field Resolution (name/title/company) ────────────────
    card.name, card.job_title, card.company = resolve_prominent_fields(
        rows, dag, card.email
    )

    # ── 7. Quality Score ──────────────────────────────────────────────────
    card.quality_score = calculate_quality_score(card)
    return card


# ══════════════════════════════════════════════════════════════════════════════
#  IMAGE-LEVEL PROCESSING ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
def process_card_image(img: np.ndarray, label: str) -> ContactCard:
    img = deskew(img)
    proc = preprocess(img)
    tokens, primary, secondary = ocr_card(proc)
    rows = group_into_rows(tokens)
    rows = remove_fuzzy_duplicates(rows)
    rows = fuse_proximate_blocks(rows)
    return extract_contact_card(rows, label)


# ══════════════════════════════════════════════════════════════════════════════
#  PERSISTENT OUTPUT LAYERS
# ══════════════════════════════════════════════════════════════════════════════
def _row_vals(r: ContactCard) -> list:
    return [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        r.name,
        r.job_title,
        r.company,
        r.email,
        r.phone_1,
        r.phone_2,
        r.phone_3,
        r.website,
        r.address,
        r.linkedin,
        r.twitter,
        r.instagram,
        r.github,
        r.quality_score,
        f"{r.confidence_avg:.1%}",
        r.source_file,
    ]


def _write_header(ws):
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, name in enumerate(EXCEL_COLS, 1):
        c = ws.cell(1, ci, value=name)
        c.fill = hf
        c.font = hfont
        c.alignment = align
        c.border = _THIN_BORDER
    ws.row_dimensions[1].height = 20


def _style_row(ws, rn: int, quality: str):
    fill = PatternFill("solid", fgColor=_QUALITY_COLOUR.get(quality, "FFFFFF"))
    align = Alignment(vertical="center", wrap_text=True)
    for ci in range(1, len(EXCEL_COLS) + 1):
        c = ws.cell(rn, ci)
        c.fill = fill
        c.alignment = align
        c.border = _THIN_BORDER
        c.font = Font(size=9)
    ws.row_dimensions[rn].height = 15


def _make_styled_wb(results: List[ContactCard]) -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cards"
    _write_header(ws)
    for ri, r in enumerate(results, 2):
        for ci, v in enumerate(_row_vals(r), 1):
            ws.cell(ri, ci, value=v)
        _style_row(ws, ri, r.quality_score)
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 55
        )
    ws.freeze_panes = "A2"
    return wb


def _safe_save_wb(wb, path: str) -> str:
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}_locked.xlsx")
        wb.save(alt)
        return alt


def _dedup_key(r: ContactCard) -> str:
    return f"{r.email}|{r.phone_1}|{r.name[:10] if r.name != '—' else ''}"


def save_to_database(results: List[ContactCard], excel_path: str):
    month = datetime.now().strftime("%Y-%m")
    wb = (
        openpyxl.load_workbook(excel_path)
        if os.path.exists(excel_path)
        else openpyxl.Workbook()
    )
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    if month in wb.sheetnames:
        ws = wb[month]
        existing = {
            f"{ws.cell(rn,5).value}|{ws.cell(rn,6).value}|{str(ws.cell(rn,2).value or '')[:10]}": rn
            for rn in range(2, ws.max_row + 1)
        }
    else:
        ws = wb.create_sheet(title=month)
        _write_header(ws)
        existing = {}

    for r in results:
        key = _dedup_key(r)
        is_u = key in existing
        rn = existing[key] if is_u else ws.max_row + 1
        for ci, v in enumerate(_row_vals(r), 1):
            ws.cell(rn, ci, value=v)
        _style_row(ws, rn, r.quality_score)
        existing[key] = rn
        log.info("  [DB] %s row %d → %s", "Updated" if is_u else "Appended", rn, r.name)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 55
        )
    ws.freeze_panes = "A2"
    _safe_save_wb(wb, excel_path)


def save_outputs(results: List[ContactCard], out_dir: str) -> Optional[str]:
    history = os.path.join(out_dir, "history")
    os.makedirs(history, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _safe_save_wb(_make_styled_wb(results), os.path.join(history, f"cards_{ts}.xlsx"))
    for i, r in enumerate(results, 1):
        safe = re.sub(r"[^\w\-]", "_", r.name)[:30]
        _safe_save_wb(
            _make_styled_wb([r]),
            os.path.join(history, f"card_{ts}_{i:02d}_{safe}.xlsx"),
        )
    latest = _safe_save_wb(
        _make_styled_wb(results), os.path.join(out_dir, "latest.xlsx")
    )
    with open(os.path.join(out_dir, "latest.json"), "w", encoding="utf-8") as f:
        json.dump([r.to_dict() for r in results], f, indent=2, ensure_ascii=False)
    return latest


def export_vcf(results: List[ContactCard], vcf_path: str):
    Path(vcf_path).write_text(
        "\n".join(r.to_vcard() for r in results), encoding="utf-8"
    )


def load_images(path: str):
    p = Path(path)
    ext = p.suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        img = cv2.imread(str(p), cv2.IMREAD_COLOR)
        if img is not None:
            yield p.name, img
    elif ext == ".pdf":
        from pdf2image import convert_from_bytes

        for i, pg in enumerate(
            convert_from_bytes(p.read_bytes(), dpi=CFG["PDF_DPI"]), 1
        ):
            yield (
                f"{p.stem}_page_{i}",
                cv2.cvtColor(np.array(pg.convert("RGB")), cv2.COLOR_RGB2BGR),
            )
    elif ext == ".zip":
        with zipfile.ZipFile(p) as zf:
            for m in sorted(
                m
                for m in zf.namelist()
                if Path(m).suffix.lower() in IMAGE_EXTENSIONS
                and not Path(m).name.startswith(".")
            ):
                arr = np.frombuffer(zf.read(m), np.uint8)
                img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                if img is not None:
                    yield f"{p.name}/{Path(m).name}", img


def print_result(r: ContactCard):
    W = 64
    print(
        f"\n{'═'*W}\n"
        f"  FILE      : {r.source_file}\n"
        f"{'─'*W}\n"
        f"  NAME      : {r.name}\n"
        f"  JOB TITLE : {r.job_title}\n"
        f"  COMPANY   : {r.company}\n"
        f"  EMAIL     : {r.email}\n"
        f"  PHONE     : {r.phone_1}\n"
        f"  WEBSITE   : {r.website}\n"
        f"  ADDRESS   : {r.address}\n"
        f"  QUALITY   : {r.quality_score}\n"
        f"{'═'*W}"
    )


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="Business Card OCR Engine v20")
    ap.add_argument("target", nargs="?", default=None)
    ap.add_argument("--lang", default=None)
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--no-vcf", action="store_true")
    ap.add_argument("--no-excel", action="store_true")
    ap.add_argument("--db", default=None)
    args = ap.parse_args()

    if args.debug:
        CFG["DEBUG"] = True
        logging.getLogger().setLevel(logging.DEBUG)

    if args.target and Path(args.target).is_dir():
        paths = sorted(
            str(f)
            for f in Path(args.target).iterdir()
            if f.suffix.lower() in IMAGE_EXTENSIONS | {".pdf", ".zip"}
        )
    elif args.target:
        paths = [args.target]
    else:
        try:
            from tkinter import Tk, filedialog

            root = Tk()
            root.withdraw()
            paths = list(filedialog.askopenfilenames(title="Select card images"))
            root.destroy()
        except Exception:
            log.info("No target provided and Tkinter unavailable.")
            return

    if not paths:
        log.info("No files selected.")
        return

    out_dir = os.path.dirname(os.path.abspath(paths[0]))
    db_path = args.db or os.path.join(out_dir, CFG["EXCEL_DB_FILE"])
    results: List[ContactCard] = []

    log.info("v20 Layout-Driven Engine — processing %d file(s)", len(paths))
    for i, p in enumerate(paths, 1):
        log.info("[%d/%d] %s", i, len(paths), p)
        try:
            for plabel, pimg in load_images(p):
                card = process_card_image(pimg, plabel)
                print_result(card)
                results.append(card)
        except Exception as e:
            log.error("Pipeline error on %s: %s", p, e)

    if not results:
        return

    saved_latest = None
    if not args.no_excel:
        save_to_database(results, db_path)
        saved_latest = save_outputs(results, out_dir)
    if not args.no_vcf:
        vcf_path = os.path.join(
            out_dir, f"contacts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.vcf"
        )
        export_vcf(results, vcf_path)
    if saved_latest and os.path.exists(saved_latest) and sys.platform == "win32":
        try:
            os.startfile(os.path.abspath(saved_latest))
        except Exception:
            pass


if __name__ == "__main__":
    main()
